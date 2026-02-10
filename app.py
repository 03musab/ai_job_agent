# --- Standard Library Imports ---
import base64
import concurrent.futures
import csv
import datetime
import hashlib
import json
import logging
import os
import pickle
import random
import re
import signal
import sqlite3
import sys
import threading
import smtplib
import socket
import time
import traceback
from collections import Counter
from dataclasses import dataclass, field
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import unquote_plus
from urllib.parse import quote_plus, urljoin
from io import BytesIO

# --- Third-Party Imports ---
import docx
import gevent.monkey
import pandas as pd
import requests
import schedule
from bs4 import BeautifulSoup
from celery import Celery
from celery.schedules import crontab
from flask import Flask, flash, jsonify, redirect, render_template, request, url_for, Response, stream_with_context
from flask_login import (LoginManager, UserMixin, current_user, login_required,
                         login_user, logout_user)
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import undetected_chromedriver as uc
from selenium import webdriver
from selenium.common.exceptions import (NoSuchElementException, TimeoutException,
                                        WebDriverException as SeleniumWebDriverException)
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
# Selenium WebDriver-specific imports
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.service import Service as FirefoxService
from werkzeug.security import check_password_hash, generate_password_hash
from flask import send_from_directory, send_file
from werkzeug.utils import secure_filename

# Try to import ReportLab for PDF generation
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# Try to import dns.resolver for SMTP checks
try:
    import dns.resolver
except ImportError:
    logger.warning("dnspython module not found. SMTP email verification will be limited.")

import threading
import traceback

# Set Cerebras API Key
os.environ['CEREBRAS_API_KEY'] = "csk-yjmhny5wcyh5dmt4wf9f5mp3k6w4cvkerw2vrh4ceyxh46vr"

# --- Local Imports ---
from resume_service import ResumeParsingService
from skill_gap_service import SkillGapService
from roadmap_service import RoadmapService
from ai_utils import safe_ai_request
from learning_models import LearningPath, LearningModule, UserLearningProgress
from extensions import db

if os.environ.get('ENABLE_GEVENT_PATCH', '0') == '1':
    gevent.monkey.patch_all(ssl=False)

# --- Logging Configuration ---
if not logging.getLogger().handlers:
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- Configurable scoring and scraping constants ---
RELEVANCE_WEIGHTS = {
    'term_in_title': 30,
    'term_in_desc': 10,
    'skill': 5,
    'tool': 4,
    'soft_skill': 3,
    'title_word': 2,
    'exp_match': 20,
    'exp_partial': 10,
    'location': 15,
    'job_type': 10,
}
RECENCY_MULTIPLIERS = {'week': 1.2, 'month': 1.05}
RELEVANCE_MIN_THRESHOLD = float(os.environ.get('RELEVANCE_MIN_THRESHOLD', '10'))
SCRAPE_MAX_WORKERS = int(os.environ.get('SCRAPE_MAX_WORKERS', '5'))
LINKEDIN_DETAIL_MAX_WORKERS = int(os.environ.get('LINKEDIN_DETAIL_MAX_WORKERS', '2'))

# --- Helpers ---

def normalize_for_hash(text: str) -> str:
    """Normalize text for consistent hashing (trim, lowercase, collapse whitespace)."""
    if text is None:
        return ''
    return re.sub(r'\s+', ' ', text.strip().lower())

# --- Celery Configuration ---
def make_celery(app):
    """Create and configure Celery instance."""
    celery = Celery(app.import_name)
    celery.conf.update(app.config)

    # Configure periodic tasks directly within Celery's configuration
    # The beat_schedule below automatically triggers scraping twice a day.
    # Comment it out to prevent automatic scraping on startup.
    # celery.conf.beat_schedule = {
    #     'run-job-hunt-daily-morning': {
    #         'task': 'app.scheduled_job_hunt_for_all_users', # Reference the task by its full path
    #         'schedule': crontab(hour=9, minute=0), # 9:00 AM daily
    #         'args': (), # No arguments for this task
    #         'options': {'queue': 'celery'} # Ensure it uses the default queue
    #     },
    #     'run-job-hunt-daily-evening': {
    #         'task': 'app.scheduled_job_hunt_for_all_users', # Reference the task by its full path
    #         'schedule': crontab(hour=18, minute=0), # 6:00 PM daily
    #         'args': (),
    #         'options': {'queue': 'celery'}
    #     },
    # }
    celery.conf.timezone = 'UTC' # Or your desired timezone, e.g., 'Asia/Kolkata'

    return celery

# Initialize Flask app (top-level, so it's imported by all processes)
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Database Configuration for Learning Paths (SQLite for now, switch to Postgres later)
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///learning_paths.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)
with app.app_context():
    db.create_all()
    # Migration check for topics_data column in learning_modules
    try:
        with db.engine.connect() as conn:
            try:
                conn.execute(db.text("SELECT topics_data FROM learning_modules LIMIT 1"))
            except Exception:
                logger.info("Migrating learning_modules: adding topics_data column")
                conn.execute(db.text("ALTER TABLE learning_modules ADD COLUMN topics_data JSON"))
                conn.commit()
    except Exception as e:
        logger.warning(f"Migration check failed: {e}")

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True) # Ensure the upload folder exists
CSV_UPLOAD_FOLDER = os.path.join(UPLOAD_FOLDER, 'csv_repository')
os.makedirs(CSV_UPLOAD_FOLDER, exist_ok=True)

# Use new style Celery config keys (lowercase without CELERY_ prefix), with fallback to old env vars
broker_url = os.environ.get('broker_url') or os.environ.get('CELERY_BROKER_URL', 'redis://localhost:6379/0')
result_backend = os.environ.get('result_backend') or os.environ.get('CELERY_RESULT_BACKEND', 'redis://localhost:6379/0')
app.config['broker_url'] = broker_url
app.config['result_backend'] = result_backend
# Enforce secure SECRET_KEY usage
if os.environ.get('FLASK_ENV') == 'production' and app.config['SECRET_KEY'] == 'dev-secret-key-change-in-production':
    logger.critical("SECRET_KEY is using an insecure default in production. Set SECRET_KEY env.")
    raise RuntimeError("Insecure SECRET_KEY in production")
elif app.config['SECRET_KEY'] == 'dev-secret-key-change-in-production':
    logger.warning("Using default SECRET_KEY; do not use in production.")

# Initialize Celery (top-level, so it's imported by all processes)
celery = make_celery(app)

# --- Health Check Endpoint for Scraping Readiness ---
@app.route('/health')
def health():
    try:
        # Increase timeout and add a single retry to avoid false negatives under load
        res = celery.control.ping(timeout=5.0)
        if not res:
            time.sleep(0.5)
            res = celery.control.ping(timeout=5.0)
        if res:
            return jsonify({'status': 'ok', 'message': 'System ready. You can now click Apply to start scraping.'})
        else:
            return jsonify({'status': 'error', 'message': 'Celery worker not responding.'}), 503
    except Exception as e:
        logger.exception("Health check failed")
        return jsonify({'status': 'error', 'message': 'Service unavailable'}), 503

@app.route('/task_status/<task_id>')
@login_required
def task_status(task_id):
    """Endpoint to check the status of a Celery task."""
    task = run_job_hunt_task.AsyncResult(task_id)
    if task.state == 'PENDING':
        response = {
            'state': task.state,
            'progress': 0,
            'status': 'Pending...'
        }
    elif task.state != 'FAILURE':
        response = {
            'state': task.state,
            'progress': task.info.get('progress', 0),
            'status': task.info.get('status', '')
        }
    else: # Something went wrong in the background
        response = {
            'state': task.state,
            'progress': 100,
            'status': f"Task failed: {str(task.info)}"
        }
    return jsonify(response)

@app.route('/contact_task_status/<task_id>')
@login_required
def contact_task_status(task_id):
    """Endpoint to check the status of a contact scraping Celery task."""
    task = scrape_contacts_task.AsyncResult(task_id)
    if task.state == 'PENDING':
        response = {
            'state': task.state,
            'progress': 0,
            'status': 'Pending...'
        }
    elif task.state != 'FAILURE':
        response = {
            'state': task.state,
            'progress': task.info.get('progress', 0),
            'status': task.info.get('status', '')
        }
    else: # Something went wrong
        response = {
            'state': task.state,
            'progress': 100,
            'status': f"Task failed: {str(task.info)}"
        }
    return jsonify(response)


# --- News Cache ---
NEWS_CACHE = {
    'data': None,
    'timestamp': 0,
    'expiry': 900  # 15 minutes
}

@app.route('/api/job-news')
@login_required
def get_job_news():
    """Fetches job market news from NewsAPI with caching."""
    global NEWS_CACHE
    
    # Check cache
    now = time.time()
    if NEWS_CACHE['data'] and (now - NEWS_CACHE['timestamp'] < NEWS_CACHE['expiry']):
        return jsonify({'status': 'success', 'data': NEWS_CACHE['data'], 'source': 'cache'})

    api_key = os.environ.get('NEWS_API_KEY')
    if not api_key:
        # Fallback/Demo data if no key provided
        return jsonify({
            'status': 'error', 
            'message': 'API Key not configured',
            'data': []
        })

    try:
        # Keywords for tech job trends
        query = '("technology jobs" OR "AI hiring" OR "software developer" OR "remote work" OR "tech industry")'
        
        url = 'https://newsapi.org/v2/everything'
        params = {
            'q': query,
            'language': 'en',
            'sortBy': 'publishedAt',
            'pageSize': 8,
            'apiKey': api_key
        }
        
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        articles = []
        for article in data.get('articles', []):
            # Simple validation
            if article.get('title') and article.get('url'):
                articles.append({
                    'title': article['title'],
                    'url': article['url'],
                    'source': article.get('source', {}).get('name', 'Unknown'),
                    'publishedAt': article.get('publishedAt', ''),
                    'urlToImage': article.get('urlToImage')
                })
        
        # Update cache
        NEWS_CACHE['data'] = articles
        NEWS_CACHE['timestamp'] = now
        
        return jsonify({'status': 'success', 'data': articles, 'source': 'api'})
        
    except Exception as e:
        logger.error(f"NewsAPI Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# Server-Sent Events endpoint for task progress streaming
@app.route('/task_progress_stream/<task_id>')
@login_required
def task_progress_stream(task_id):
    """Stream task progress as SSE events. Clients should connect with EventSource.
    The server polls Celery AsyncResult for state and meta and emits JSON payloads
    with keys: percent (int 0-100), state (string), milestone (optional), message (optional).
    """
    def generate():
        # initial client retry directive
        yield 'retry: 1000\n\n'
        try:
            async_result = celery.AsyncResult(task_id)
        except Exception:
            payload = {'percent': 0, 'state': 'UNKNOWN', 'milestone': 'init', 'message': 'invalid task id'}
            yield f"data: {json.dumps(payload)}\n\n"
            return

        while True:
            try:
                async_result = celery.AsyncResult(task_id)
                state = getattr(async_result, 'state', None) or 'PENDING'
                info = async_result.info or {}
            except Exception:
                state = 'UNKNOWN'
                info = {}

            # Determine percent from meta if present
            percent = None
            if isinstance(info, dict):
                if 'progress' in info:
                    try:
                        percent = int(float(info.get('progress') or 0))
                    except Exception:
                        percent = None
                elif 'percent' in info:
                    try:
                        percent = int(float(info.get('percent') or 0))
                    except Exception:
                        percent = None

            # Fallback mapping for states
            if percent is None:
                if state == 'PENDING':
                    percent = 0
                elif state in ('RETRY', 'STARTED'):
                    percent = 10
                elif state == 'PROGRESS':
                    percent = 30
                elif state == 'SUCCESS':
                    percent = 100
                elif state == 'FAILURE':
                    percent = 100
                else:
                    percent = 5

            milestone = info.get('milestone') if isinstance(info, dict) else None
            message = info.get('message') if isinstance(info, dict) else None

            payload = {
                'percent': int(max(0, min(100, percent))) if percent is not None else 0,
                'state': state,
                'milestone': milestone or None,
                'message': message or None,
            }

            yield f"data: {json.dumps(payload)}\n\n"

            # Terminate stream on final states
            if state in ('SUCCESS', 'FAILURE') or async_result.ready():
                break

            time.sleep(0.7)

    headers = {
        'Content-Type': 'text/event-stream',
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',
    }
    return Response(stream_with_context(generate()), headers=headers)

# --- Flask-Login Configuration ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' # Redirect to login page if not authenticated

# --- Driver Paths Configuration ---
DRIVER_PATHS = {
    'edge': r"C:\WebDrivers\msedgedriver.exe",
    'chrome': r"C:\WebDrivers\chromedriver.exe",
    'firefox': r"C:\WebDrivers\geckodriver.exe",
    'chrome_pa': None, # Placeholder for headless Linux environments
    'firefox_pa': None # Placeholder for headless Linux environments
}

# --- Browser Binary Paths (for Windows, specify exact locations if not default) ---
# IMPORTANT: DOUBLE-CHECK THESE PATHS ON YOUR SYSTEM.
CHROME_BINARY_PATH_WINDOWS = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
EDGE_BINARY_PATH_WINDOWS = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
CHROME_BINARY_PATH_PA = '/usr/bin/google-chrome'


# --- Google API Scopes for Gmail and Calendar ---
SCOPES = [
    'https://www.googleapis.com/auth/gmail.send',
    'https://www.googleapis.com/auth/calendar'
]

# --- Admin Email for Critical Alerts ---
ADMIN_ALERT_EMAIL = os.environ.get('ADMIN_ALERT_EMAIL', '')


# --- Add a list of common User-Agent strings to mimic different browsers ---
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.1 Safari/605.1.15',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
]

# --- Description sanitizer and Jinja filter ---

def clean_description(text: str, title: str = None, company: str = None) -> str:
    if not text:
        return ''
    try:
        # Split into lines to allow targeted filtering
        lines = re.split(r'[\r\n]+', text)
        cleaned = []
        patterns = [
            re.compile(r'(₹|rs\.?|lpa|ctc|per\s*annum|/year|/month|stipend)', re.IGNORECASE),
            re.compile(r'\b(actively\s*hiring)\b', re.IGNORECASE),
            re.compile(r'\b(work\s*from\s*home|remote)\b', re.IGNORECASE),
            re.compile(r'\b\d+\s*(?:year|week|day|month)s?\s*ago\b', re.IGNORECASE),
            re.compile(r'\b\d+\s*year\(s\)\b', re.IGNORECASE),
            re.compile(r'\bfresher\b', re.IGNORECASE),
        ]
        t_low = (title or '').strip().lower()
        c_low = (company or '').strip().lower()
        for line in lines:
            l = line.strip()
            if not l:
                continue
            # Skip lines that are exactly the title or company
            if t_low and l.lower() == t_low:
                continue
            if c_low and l.lower() == c_low:
                continue
            # Skip noisy lines
            if any(p.search(l) for p in patterns):
                continue
            cleaned.append(l)
        # Collapse multiple spaces and join back with newlines
        result = '\n'.join(re.sub(r'\s+', ' ', c) for c in cleaned)
        return result.strip()
    except Exception:
        # Fallback: basic whitespace normalization
        return re.sub(r'\s+', ' ', text).strip()

# Register as Jinja filter
@app.template_filter('clean_desc')
def jinja_clean_desc_filter(text: str) -> str:
    return clean_description(text)

@app.template_filter('fromjson')
def fromjson_filter(json_string):
    """Custom Jinja filter to parse a JSON string into a Python object."""
    if json_string is None:
        return {}  # Return an empty dict to prevent errors on attribute access
    try:
        return json.loads(json_string)
    except (json.JSONDecodeError, TypeError):
        logger.warning(f"Could not decode JSON string in template: {json_string[:100]}")
        return {} # Return an empty dict on error

@app.context_processor
def inject_now():
    return {'now': datetime.datetime.utcnow()}


@dataclass
class Job:
    """Enhanced Job data structure to store scraped job details.""" # noqa: E501
    title: str
    company: str
    location: str
    salary: str
    link: str
    description: str
    keywords: List[str]
    skills: List[str]
    experience: str
    job_type: str
    posted_date: str
    source: str
    deadline: str
    relevance_score: float = 0.0

    min_experience_years: Optional[int] = None
    max_experience_years: Optional[int] = None
    extracted_tools: List[str] = field(default_factory=list)
    extracted_soft_skills: List[str] = field(default_factory=list)

    user_feedback: Optional[str] = None

    def to_dict(self):
        """Converts the Job object to a dictionary for easier storage/reporting."""
        return {
            'id': getattr(self, 'id', None),
            'title': self.title,
            'company': self.company,
            'location': self.location,
            'salary': self.salary,
            'link': self.link,
            'description': self.description[:500] + '...' if len(self.description) > 500 else self.description,
            'keywords': ', '.join(self.keywords),
            'skills': ', '.join(self.skills),
            'experience': self.experience,
            'job_type': self.job_type,
            'posted_date': self.posted_date,
            'source': self.source,
            'deadline': self.deadline,
            'relevance_score': self.relevance_score,
            'min_experience_years': self.min_experience_years,
            'max_experience_years': self.max_experience_years,
            'extracted_tools': ', '.join(self.extracted_tools),
            'extracted_soft_skills': ', '.join(self.extracted_soft_skills),
            'user_feedback': self.user_feedback
        }

class User(UserMixin):
    def __init__(self, id, username, password_hash, email_recipients='', send_excel_attachment=True):
        self.id = id
        self.username = username
        self.password_hash = password_hash
        self.email_recipients = email_recipients
        self.send_excel_attachment = send_excel_attachment

    @staticmethod
    def get(user_id):
        db = JobDatabase()
        conn = sqlite3.connect(db.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, username, password_hash, email_recipients, send_excel_attachment FROM users WHERE id = ?", (user_id,))
        user_data = cursor.fetchone()
        conn.close()
        if user_data:
            return User(*user_data)
        return None

    @staticmethod
    def get_by_username(username):
        db = JobDatabase()
        conn = sqlite3.connect(db.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, username, password_hash, email_recipients, send_excel_attachment FROM users WHERE username = ?", (username,))
        user_data = cursor.fetchone()
        conn.close()
        if user_data:
            return User(*user_data)
        return None

@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)


class JobDatabase:
    """SQLite database for job tracking and deduplication and user data."""

    def __init__(self, db_path="jobs.db"):
        self.db_path = db_path
        self.init_db()

    def init_db(self):
        """Initializes the SQLite database table(s) if they don't exist, and performs schema migrations."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # --- Users Table ---
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                email_recipients TEXT DEFAULT '',
                send_excel_attachment BOOLEAN DEFAULT TRUE
            )
        ''')
        # Perform ALTER TABLE for new user columns if they don't exist
        cursor.execute("PRAGMA table_info(users)")
        user_columns = [col[1] for col in cursor.fetchall()]
        if 'send_excel_attachment' not in user_columns:
            try: cursor.execute("ALTER TABLE users ADD COLUMN send_excel_attachment BOOLEAN DEFAULT TRUE"); conn.commit(); logger.info("Added send_excel_attachment column to users table.")
            except sqlite3.OperationalError as e: logger.warning(f"send_excel_attachment column already exists or error altering users table: {e}")

        # --- Jobs Table ---
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_details (
                user_id INTEGER PRIMARY KEY,
                full_name TEXT,
                email TEXT,
                phone TEXT,
                address TEXT,
                linkedin_url TEXT,
                github_url TEXT,
                portfolio_url TEXT,
                resume_path TEXT,
                cover_letter_template TEXT,
                browser_profile_name TEXT DEFAULT 'Default',
                browser_profile_path TEXT DEFAULT '',
                resume_parsed_data TEXT,
                willing_to_relocate TEXT DEFAULT 'no',
                authorized_to_work TEXT DEFAULT 'no',
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        ''')

        # Perform ALTER TABLE for new user_details columns if they don't exist
        cursor.execute("PRAGMA table_info(user_details)")
        user_details_columns = [col[1] for col in cursor.fetchall()]
        if 'browser_profile_path' not in user_details_columns:
            try:
                cursor.execute("ALTER TABLE user_details ADD COLUMN browser_profile_path TEXT DEFAULT ''")
                conn.commit()
                logger.info("Added browser_profile_path column to user_details table.")
            except sqlite3.OperationalError as e:
                logger.warning(f"browser_profile_path column already exists or error altering user_details table: {e}")
        if 'browser_profile_name' not in user_details_columns:
            try:
                cursor.execute("ALTER TABLE user_details ADD COLUMN browser_profile_name TEXT DEFAULT 'Default'")
                conn.commit()
                logger.info("Added browser_profile_name column to user_details table.")
            except sqlite3.OperationalError as e:
                logger.warning(f"browser_profile_name column already exists or error altering user_details table: {e}")
        if 'resume_parsed_data' not in user_details_columns:
            try:
                cursor.execute("ALTER TABLE user_details ADD COLUMN resume_parsed_data TEXT")
                conn.commit()
                logger.info("Added resume_parsed_data column to user_details table.")
            except sqlite3.OperationalError as e:
                logger.warning(f"resume_parsed_data column already exists or error altering user_details table: {e}")
        if 'willing_to_relocate' not in user_details_columns:
            try:
                cursor.execute("ALTER TABLE user_details ADD COLUMN willing_to_relocate TEXT DEFAULT 'no'")
                conn.commit()
                logger.info("Added willing_to_relocate column to user_details table.")
            except sqlite3.OperationalError as e:
                logger.warning(f"willing_to_relocate column already exists or error altering user_details table: {e}")
        if 'authorized_to_work' not in user_details_columns:
            try:
                cursor.execute("ALTER TABLE user_details ADD COLUMN authorized_to_work TEXT DEFAULT 'no'")
                conn.commit()
                logger.info("Added authorized_to_work column to user_details table.")
            except sqlite3.OperationalError as e:
                logger.warning(f"authorized_to_work column already exists or error altering user_details table: {e}")
        if 'assisted_apply_mode' not in user_details_columns:
            try:
                cursor.execute("ALTER TABLE user_details ADD COLUMN assisted_apply_mode TEXT DEFAULT 'personal'")
                conn.commit()
                logger.info("Added assisted_apply_mode column to user_details table.")
            except sqlite3.OperationalError as e:
                logger.warning(f"assisted_apply_mode column already exists or error altering user_details table: {e}")

        cursor.execute("PRAGMA table_info(jobs)")
        jobs_existing_columns = [col[1] for col in cursor.fetchall()]

        # Robust check for table existence
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='jobs'")
        if cursor.fetchone() is None:
            logger.info("Jobs table does not exist, creating it with all columns.")
            cursor.execute('''
                CREATE TABLE jobs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_hash TEXT UNIQUE,
                    title TEXT,
                    company TEXT,
                    location TEXT,
                    salary TEXT,
                    link TEXT,
                    description TEXT,
                    keywords TEXT,
                    skills TEXT,
                    experience TEXT,
                    job_type TEXT,
                    posted_date TEXT,
                    source TEXT,
                    relevance_score REAL,
                    found_date TEXT,
                    applied BOOLEAN DEFAULT FALSE,
                    status TEXT DEFAULT 'new',
                    user_id INTEGER,
                    notes TEXT DEFAULT '',
                    min_experience_years INTEGER,
                    max_experience_years INTEGER,
                    extracted_tools TEXT DEFAULT '',
                    extracted_soft_skills TEXT DEFAULT '',
                    user_feedback TEXT DEFAULT '',
                    deadline TEXT,
                    FOREIGN KEY(user_id) REFERENCES users(id)
                )
            ''')
        else: # Table exists, check for new columns and alter
            if 'notes' not in jobs_existing_columns:
                try: cursor.execute("ALTER TABLE jobs ADD COLUMN notes TEXT DEFAULT ''"); conn.commit(); logger.info("Added notes column to jobs table.")
                except sqlite3.OperationalError as e: logger.warning(f"notes column already exists or error altering jobs table: {e}")
            if 'min_experience_years' not in jobs_existing_columns:
                try: cursor.execute("ALTER TABLE jobs ADD COLUMN min_experience_years INTEGER"); conn.commit(); logger.info("Added min_experience_years column to jobs table.")
                except sqlite3.OperationalError as e: logger.warning(f"min_experience_years column already exists or error altering jobs table: {e}")
            if 'max_experience_years' not in jobs_existing_columns:
                try: cursor.execute("ALTER TABLE jobs ADD COLUMN max_experience_years INTEGER"); conn.commit(); logger.info("Added max_experience_years column to jobs table.")
                except sqlite3.OperationalError as e: logger.warning(f"max_experience_years column already exists or error altering jobs table: {e}")
            if 'extracted_tools' not in jobs_existing_columns:
                try: cursor.execute("ALTER TABLE jobs ADD COLUMN extracted_tools TEXT DEFAULT ''"); conn.commit(); logger.info("Added extracted_tools column to jobs table.")
                except sqlite3.OperationalError as e: logger.warning(f"extracted_tools column already exists or error altering jobs table: {e}")
            if 'extracted_soft_skills' not in jobs_existing_columns:
                try: cursor.execute("ALTER TABLE jobs ADD COLUMN extracted_soft_skills TEXT DEFAULT ''"); conn.commit(); logger.info("Added extracted_soft_skills column to jobs table.")
                except sqlite3.OperationalError as e: logger.warning(f"extracted_soft_skills column already exists or error altering jobs table: {e}")
            if 'user_feedback' not in jobs_existing_columns:
                try:
                    cursor.execute("ALTER TABLE jobs ADD COLUMN user_feedback TEXT DEFAULT ''")
                    conn.commit(); logger.info("Added user_feedback column to jobs table.")
                except sqlite3.OperationalError as e:
                    logger.warning(f"user_feedback column already exists or error altering jobs table: {e}")
            if 'deadline' not in jobs_existing_columns:
                try:
                    cursor.execute("ALTER TABLE jobs ADD COLUMN deadline TEXT")
                    conn.commit(); logger.info("Added deadline column to jobs table.")
                except sqlite3.OperationalError as e:
                    logger.warning(f"deadline column already exists or error altering jobs table: {e}")

        # --- Search Profiles Table ---
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS search_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                profile_name TEXT NOT NULL,
                search_terms TEXT NOT NULL,
                location TEXT,
                experience TEXT,
                job_type TEXT,
                UNIQUE(user_id, profile_name),
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        ''')

        # --- User Custom Scores Table (for AI-powered relevance) ---
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_custom_scores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                keyword TEXT NOT NULL, -- The specific keyword/skill/tool (always stored lowercase)
                keyword_type TEXT NOT NULL, -- e.g., 'high_value', 'skill', 'tool', 'soft_skill'
                score_multiplier REAL DEFAULT 1.0, -- Multiplier for this keyword's default score
                UNIQUE(user_id, keyword, keyword_type), -- Ensure unique entry per user/keyword/type
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        ''')

        # --- User Job Feedback Table (for feedback history) ---
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_job_feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                job_id INTEGER NOT NULL,
                feedback_type TEXT NOT NULL, -- 'like' or 'dislike'
                timestamp TEXT NOT NULL,
                UNIQUE(user_id, job_id), -- Only one feedback per job per user
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(job_id) REFERENCES jobs(id)
            )
        ''')

        conn.commit()
        conn.close()
        logger.info("Database schema check/update complete.")
        if os.environ.get('RUN_JOB_HASH_MIGRATION', '0') == '1':
            logger.info("RUN_JOB_HASH_MIGRATION=1 detected; starting job_hash migration...")
            self.migrate_job_hashes_normalized()

    def add_user(self, username, password_hash):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)", (username, password_hash))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            return None # Username already exists
        finally:
            conn.close()

    def add_job(self, job: Job, user_id: Optional[int] = None):
        """Adds a job to the database with deduplication and user association."""
        # The hash must be unique per user for the same job to allow multiple users to track it.
        norm = f"{normalize_for_hash(job.title)}|{normalize_for_hash(job.company)}|{normalize_for_hash(job.location)}|{user_id}"
        job_hash = hashlib.md5(norm.encode('utf-8')).hexdigest()

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT OR IGNORE INTO jobs
                (job_hash, title, company, location, salary, link, description,
                 keywords, skills, experience, job_type, posted_date, source,
                 relevance_score, found_date, user_id, status, notes, deadline,
                 min_experience_years, max_experience_years, extracted_tools, extracted_soft_skills, user_feedback)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                job_hash,  # job_hash
                job.title,  # title
                job.company,  # company
                job.location,  # location
                job.salary,  # salary
                job.link,  # link
                job.description,  # description
                ', '.join(job.keywords),  # keywords
                ', '.join(job.skills),  # skills
                job.experience,  # experience
                job.job_type,  # job_type
                job.posted_date,  # posted_date
                job.source,  # source
                job.relevance_score,  # relevance_score
                datetime.datetime.now().isoformat(),  # found_date
                user_id,  # user_id
                'new',  # status
                '',  # notes
                job.deadline,  # deadline
                job.min_experience_years,  # min_experience_years
                job.max_experience_years,  # max_experience_years
                ', '.join(job.extracted_tools),  # extracted_tools
                ', '.join(job.extracted_soft_skills),  # extracted_soft_skills
                job.user_feedback if job.user_feedback else ''  # user_feedback
            ))
            conn.commit()
            if cursor.rowcount > 0:
                logger.debug(f"Added new job to DB: {job.title} at {job.company} for user {user_id}")
                
                # Create calendar reminder for deadline if available
                if job.deadline and job.deadline != 'N/A':
                    try:
                        calendar_manager = GoogleCalendarManager()
                        job_data = job.to_dict()
                        job_data['id'] = cursor.lastrowid
                        calendar_manager.create_deadline_reminder(job_data, job.deadline)
                        logger.info(f"Created deadline reminder for job: {job.title}")
                    except Exception as e:
                        logger.warning(f"Failed to create deadline reminder: {e}")
                
                return True
            else:
                logger.debug(f"Job already exists in DB (skipped): {job.title} at {job.company} for user {user_id}")
                return False
        except sqlite3.OperationalError as e:
            logger.error(f"SQLite Operational Error (likely schema mismatch during job insert): {e}.")
            # Log with the corrected order for accurate debugging
            correct_values = (job_hash, job.title, job.company, job.location, job.salary, job.link, job.description, ', '.join(job.keywords), ', '.join(job.skills), job.experience, job.job_type, job.posted_date, job.source, job.relevance_score, datetime.datetime.now().isoformat(), user_id, 'new', '', job.deadline, job.min_experience_years, job.max_experience_years, ', '.join(job.extracted_tools), ', '.join(job.extracted_soft_skills), job.user_feedback if job.user_feedback else '')
            logger.error("Attempting to insert values: %s", correct_values)
            return False
        except Exception as e:
            logger.error(f"Error adding job to database: {e}")
            return False
        finally:
            conn.close()

    def get_jobs_for_user(self, user_id: int, limit=100, offset=0,
                             sort_by='found_date', sort_order='DESC',
                             search_query: Optional[str] = None,
                             status_filter: Optional[str] = None,
                             location_filter: Optional[str] = None,
                             job_type_filter: Optional[str] = None) -> Tuple[List[Dict], int]:
        """Retrieves jobs for a specific user from the database, with filters and pagination."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        query_columns = """
            id, job_hash, title, company, location, salary, link, description,
            keywords, skills, experience, job_type, posted_date, source,
            relevance_score, found_date, applied, status, user_id, notes, deadline,
            min_experience_years, max_experience_years, extracted_tools, extracted_soft_skills, user_feedback
        """
        query = f"SELECT {query_columns} FROM jobs WHERE user_id = ?"
        params = [user_id]

        if search_query:
            # Handle special search formats like "id:123"
            if search_query.startswith('id:'):
                job_id = search_query.split(':', 1)[1].strip()
                if job_id.isdigit():
                    query += " AND id = ?"
                    params.append(int(job_id))
                else:
                    # Invalid ID format, no results
                    query += " AND 1 = 0"
            else:
                # Regular text search
                search_pattern = f"%{search_query}%"
                query += " AND (title LIKE ? OR company LIKE ? OR description LIKE ? OR notes LIKE ? OR keywords LIKE ? OR skills LIKE ? OR extracted_tools LIKE ? OR extracted_soft_skills LIKE ?)"
                params.extend([search_pattern, search_pattern, search_pattern, search_pattern, search_pattern, search_pattern, search_pattern, search_pattern])

        if status_filter and status_filter != 'all':
            query += " AND status = ?"
            params.append(status_filter)

        if location_filter:
            location_pattern = f"%{location_filter}%"
            query += " AND location LIKE ?"
            params.append(location_pattern)

        if job_type_filter and job_type_filter != 'all':
            query += " AND job_type = ?"
            params.append(job_type_filter)

        count_query = f"SELECT COUNT(*) FROM ({query})"
        cursor.execute(count_query, params)
        total_count = cursor.fetchone()[0]

        # Safely construct ORDER BY clause
        allowed_sort_columns = ['found_date', 'title', 'location', 'relevance_score']
        if sort_by not in allowed_sort_columns:
            sort_by = 'found_date' # Default to safe column
        sort_order = 'ASC' if sort_order.upper() == 'ASC' else 'DESC' # Ensure valid order

        query += f" ORDER BY {sort_by} {sort_order} LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        cursor.execute(query, params)
        jobs_raw = cursor.fetchall()

        column_names = [description[0] for description in cursor.description]
        jobs = [dict(zip(column_names, job_tuple)) for job_tuple in jobs_raw]

        conn.close()
        logger.info(f"Retrieved {len(jobs)} jobs for user {user_id} (Total: {total_count}).")
        return jobs, total_count

    def save_search_profile(self, user_id: int, profile_name: str, search_terms: str,
                            location: str, experience: str, job_type: str) -> Optional[int]:
        """Saves a search profile for a user. Returns profile ID on success, None on failure."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO search_profiles
                (user_id, profile_name, search_terms, location, experience, job_type)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (user_id, profile_name, search_terms, location, experience, job_type))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            logger.warning(f"Attempted to save duplicate profile name '{profile_name}' for user {user_id}.")
            return None
        except Exception as e:
            logger.error(f"Error saving search profile: {e}")
            return None
        finally:
            conn.close()

    def get_search_profiles(self, user_id: int) -> List[Dict]:
        """Retrieves all saved search profiles for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, profile_name, search_terms, location, experience, job_type FROM search_profiles WHERE user_id = ?", (user_id,))
        profiles_raw = cursor.fetchall()
        column_names = [description[0] for description in cursor.description]
        conn.close()

        profiles = [dict(zip(column_names, profile_tuple)) for profile_tuple in profiles_raw]
        return profiles

    def get_search_profile_by_id(self, profile_id: int, user_id: int) -> Optional[Dict]:
        """Retrieves a single search profile by ID and user ID."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, profile_name, search_terms, location, experience, job_type FROM search_profiles WHERE id = ? AND user_id = ?", (profile_id, user_id))
        profile_data = cursor.fetchone()
        column_names = [description[0] for description in cursor.description]
        conn.close()
        if profile_data:
            return dict(zip(column_names, profile_data))
        return None

    def delete_search_profile(self, profile_id: int, user_id: int) -> bool:
        """Deletes a search profile for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM search_profiles WHERE id = ? AND user_id = ?", (profile_id, user_id))
        conn.commit()
        rows_affected = cursor.rowcount
        conn.close()
        return rows_affected > 0

    def update_user_settings(self, user_id: int, email_recipients: str, send_excel_attachment: bool) -> bool:
        """Updates user's email notification settings."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE users SET
                    email_recipients = ?, send_excel_attachment = ?
                WHERE id = ?
            ''', (email_recipients, int(bool(send_excel_attachment)), user_id))
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error updating user settings for user {user_id}: {e}")
            return False
        finally:
            conn.close()

    def get_user_settings(self, user_id: int) -> Optional[Dict]:
        """Retrieves a user's email notification settings."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT email_recipients, send_excel_attachment FROM users WHERE id = ?", (user_id,))
        settings_data = cursor.fetchone()
        conn.close()
        if settings_data: return {'email_recipients': settings_data[0], 'send_excel_attachment': bool(settings_data[1])}
        return None

    def update_job_status_and_notes(self, job_id: int, user_id: int, status: str, notes: str) -> bool:
        """Updates the application status and notes for a specific job."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE jobs SET status = ?, notes = ? WHERE id = ? AND user_id = ?
            ''', (status, notes, job_id, user_id))
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error updating job status/notes for job {job_id}, user {user_id}: {e}")
            return False
        finally:
            conn.close()

    def record_job_feedback(self, user_id: int, job_id: int, feedback_type: str) -> bool:
        """Records user feedback for a job and updates the job's feedback status."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE jobs SET user_feedback = ? WHERE id = ? AND user_id = ?", (feedback_type, job_id, user_id))
            cursor.execute("INSERT OR REPLACE INTO user_job_feedback (user_id, job_id, feedback_type, timestamp) VALUES (?, ?, ?, ?)",
                           (user_id, job_id, feedback_type, datetime.datetime.now().isoformat()))
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error recording job feedback for user {user_id}, job {job_id}: {e}")
            return False
        finally:
            conn.close()

    def delete_job(self, job_id: int, user_id: int) -> bool:
        """Deletes a specific job for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM jobs WHERE id = ? AND user_id = ?", (job_id, user_id))
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error deleting job {job_id} for user {user_id}: {e}")
            return False
        finally:
            conn.close()

    def delete_jobs(self, job_ids: List[int], user_id: int) -> int:
        """Deletes multiple jobs for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            placeholders = ','.join(['?'] * len(job_ids))
            cursor.execute(f"DELETE FROM jobs WHERE id IN ({placeholders}) AND user_id = ?", (*job_ids, user_id))
            deleted_count = cursor.rowcount
            conn.commit()
            return deleted_count
        except Exception as e:
            logger.error(f"Error deleting multiple jobs for user {user_id}: {e}")
            return 0
        finally:
            conn.close()

    def delete_all_jobs(self, user_id: int) -> int:
        """Deletes all jobs for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM jobs WHERE user_id = ?", (user_id,))
            deleted_count = cursor.rowcount
            conn.commit()
            return deleted_count
        except Exception as e:
            logger.error(f"Error deleting all jobs for user {user_id}: {e}")
            return 0
        finally:
            conn.close()

    def get_job_status_counts(self, user_id: int) -> Dict[str, int]:
        """Retrieves the count of jobs for each status for a specific user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Initialize counts to ensure all statuses are present, even if they are 0.
        status_counts = {
            'new': 0,
            'applied': 0,
            'interview': 0,
            'rejected': 0,
            'offered': 0,
            'archived': 0
        }

        try:
            cursor.execute("SELECT status, COUNT(*) FROM jobs WHERE user_id = ? AND status IS NOT NULL GROUP BY status", (user_id,))
            rows = cursor.fetchall()
            for status, count in rows:
                if status in status_counts:
                    status_counts[status] = count
            return status_counts
        except Exception as e:
            logger.error(f"Error getting job status counts for user {user_id}: {e}")
            return status_counts # Return default counts on error
        finally:
            conn.close()

    def update_custom_score(self, user_id: int, keyword: str, keyword_type: str, change_multiplier: float) -> None:
        """Adjusts the custom relevance score multiplier for a specific keyword for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT score_multiplier FROM user_custom_scores WHERE user_id = ? AND keyword = ? AND keyword_type = ?",
                           (user_id, keyword, keyword_type))
            row = cursor.fetchone()

            if row:
                new_multiplier = row[0] * (1 + change_multiplier)
                new_multiplier = max(0.5, min(2.0, new_multiplier)) # Clamp multiplier to prevent extreme values
                cursor.execute("UPDATE user_custom_scores SET score_multiplier = ? WHERE user_id = ? AND keyword = ? AND keyword_type = ?",
                               (new_multiplier, user_id, keyword, keyword_type))
            else:
                initial_multiplier = 1.0 * (1 + change_multiplier)
                initial_multiplier = max(0.5, min(2.0, initial_multiplier))
                cursor.execute("INSERT INTO user_custom_scores (user_id, keyword, keyword_type, score_multiplier) VALUES (?, ?, ?, ?)",
                               (user_id, keyword, keyword_type, initial_multiplier))
            conn.commit()
        except Exception as e:
            logger.error(f"Error updating custom score for user {user_id}, keyword '{keyword}' type '{keyword_type}': {e}")
        finally:
            conn.close()

    def get_custom_scores(self, user_id: int) -> Dict[Tuple[str, str], float]:
        """Retrieves all custom score multipliers for a given user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT keyword, keyword_type, score_multiplier FROM user_custom_scores WHERE user_id = ?", (user_id,))
        rows = cursor.fetchall()
        conn.close()

        custom_scores = {}
        for row in rows:
            keyword, keyword_type, multiplier = row
            custom_scores[(keyword, keyword_type)] = multiplier
        return custom_scores

    def migrate_job_hashes_normalized(self) -> None:
        """One-time migration to recompute job_hash using normalized title/company/location and user_id."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            # Fetch user_id to make the hash user-specific
            cursor.execute("SELECT id, title, company, location, job_hash, user_id FROM jobs")
            rows = cursor.fetchall()
            updated = 0
            for id_, title, company, location, old_hash, user_id in rows:
                # Recompute hash with user_id
                norm = f"{normalize_for_hash(title)}|{normalize_for_hash(company)}|{normalize_for_hash(location)}|{user_id}"
                new_hash = hashlib.md5(norm.encode('utf-8')).hexdigest()
                if new_hash != old_hash:
                    try:
                        cursor.execute("UPDATE jobs SET job_hash = ? WHERE id = ?", (new_hash, id_))
                        updated += 1
                    except sqlite3.IntegrityError:
                        logger.warning(f"Skipping job id {id_} due to hash collision during migration (likely a true duplicate for the same user).")
                        continue
            conn.commit()
            logger.info(f"Job hash migration complete. Updated {updated} rows.")
        except sqlite3.OperationalError as e:
            if "no such column: user_id" in str(e):
                logger.warning("Skipping job_hash migration: 'user_id' column not found in 'jobs' table. This is expected on very old schemas.")
            else:
                logger.error(f"Job hash migration failed with an operational error: {e}")
        except Exception as e:
            logger.error(f"Job hash migration failed: {e}")
        finally:
            conn.close()

    def get_user_details(self, user_id: int) -> Optional[Dict]:
        """Retrieves a user's application details."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row # This allows accessing columns by name
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM user_details WHERE user_id = ?", (user_id,))
        details_data = cursor.fetchone()
        conn.close()
        if details_data:
            return dict(details_data)
        return None

    def save_user_details(self, user_id: int, details: Dict) -> bool:
        """Saves or updates a user's application details."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            # Using INSERT OR REPLACE (UPSERT) for simplicity
            cursor.execute('''
                INSERT OR REPLACE INTO user_details (
                    user_id, full_name, email, phone, address,
                    linkedin_url, github_url, portfolio_url,
                    resume_path, cover_letter_template, browser_profile_path, browser_profile_name,
                    resume_parsed_data, willing_to_relocate, authorized_to_work, assisted_apply_mode
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                user_id,
                details.get('full_name'), details.get('email'), details.get('phone'),
                details.get('address'), details.get('linkedin_url'), details.get('github_url'),
                details.get('portfolio_url'), details.get('resume_path'),
                details.get('cover_letter_template'),
                details.get('browser_profile_path'), details.get('browser_profile_name', 'Default'),
                details.get('resume_parsed_data'),
                details.get('willing_to_relocate'),
                details.get('authorized_to_work'),
                details.get('assisted_apply_mode', 'personal')
            ))
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error saving user details for user {user_id}: {e}")
            return False
        finally:
            conn.close()

    def delete_user_details(self, user_id: int) -> bool:
        """Deletes a user's application details from the user_details table."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM user_details WHERE user_id = ?", (user_id,))
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error deleting user details for user {user_id}: {e}")
            return False
        finally:
            conn.close()

class OutreachDatabase:
    """Manages the SQLite database for all outreach activities."""

    def __init__(self, db_path="outreach.db"):
        self.db_path = db_path
        self.init_db()

    def init_db(self):
        """Initializes the outreach database tables if they don't exist."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Contacts Table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                full_name TEXT NOT NULL,
                title TEXT,
                company TEXT,
                email TEXT,
                linkedin_url TEXT UNIQUE,
                source TEXT, -- e.g., 'LinkedIn', 'Manual'
                status TEXT DEFAULT 'new', -- e.g., 'new', 'contacted', 'replied'
                notes TEXT,
                added_date TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        ''')

        # Global Hiring Contacts Table (for uploaded CSV data)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS global_hiring_contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT,
                title TEXT,
                company TEXT,
                email TEXT,
                linkedin_url TEXT UNIQUE,
                source TEXT DEFAULT 'Uploaded CSV',
                added_date TEXT NOT NULL
            )
        ''')
        # This UNIQUE constraint on linkedin_url will prevent duplicate entries
        # when we use 'INSERT OR IGNORE' or 'to_sql(if_exists='append')'.

        # Templates Table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                template_name TEXT NOT NULL,
                subject TEXT,
                body TEXT NOT NULL,
                type TEXT DEFAULT 'email', -- 'email', 'linkedin_message', 'text'
                created_date TEXT NOT NULL,
                UNIQUE(user_id, template_name),
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        ''')

        # Add 'tone' column to templates table if it doesn't exist
        cursor.execute("PRAGMA table_info(templates)")
        template_columns = [col[1] for col in cursor.fetchall()]
        if 'tone' not in template_columns:
            try:
                cursor.execute("ALTER TABLE templates ADD COLUMN tone TEXT DEFAULT 'direct'")
                conn.commit()
                logger.info("Added tone column to templates table.")
            except sqlite3.OperationalError as e:
                logger.warning(f"tone column already exists or error altering templates table: {e}")

        # Campaigns Table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS campaigns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                campaign_name TEXT NOT NULL,
                description TEXT,
                status TEXT DEFAULT 'draft', -- 'draft', 'active', 'completed'
                created_date TEXT NOT NULL,
                UNIQUE(user_id, campaign_name),
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        ''')

        # Outreach Logs Table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS outreach_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                campaign_id INTEGER NOT NULL,
                contact_id INTEGER NOT NULL,
                template_id INTEGER,
                channel TEXT NOT NULL, -- 'email', 'linkedin', etc.
                sent_date TEXT NOT NULL,
                status TEXT DEFAULT 'sent', -- 'sent', 'opened', 'replied', 'ghosted'
                status_date TEXT,
                notes TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(campaign_id) REFERENCES campaigns(id),
                FOREIGN KEY(contact_id) REFERENCES contacts(id),
                FOREIGN KEY(template_id) REFERENCES templates(id)
            )
        ''')

        conn.commit()
        conn.close()
        logger.info("Outreach database schema check/update complete.")

    def get_outreach_stats(self, user_id: int) -> Dict[str, Any]:
        """Retrieves key statistics for the outreach dashboard."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        stats = {}
        try:
            # Only count the user's personal contacts for their dashboard.
            cursor.execute("SELECT COUNT(*) FROM contacts WHERE user_id = ?", (user_id,))
            stats['contact_count'] = cursor.fetchone()[0]

            return stats
        finally:
            conn.close()

    def add_campaign(self, user_id: int, data: Dict) -> Optional[int]:
        """Adds a new campaign to the database."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO campaigns (user_id, campaign_name, description, status, created_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                user_id,
                data['campaign_name'],
                data.get('description', ''),
                data.get('status', 'draft'),
                datetime.datetime.now().isoformat()
            ))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            logger.warning(f"Campaign '{data['campaign_name']}' already exists for user {user_id}.")
            return None
        finally:
            conn.close()

    def update_campaign_status(self, campaign_id: int, user_id: int, status: str) -> bool:
        """Updates the status of a campaign (e.g., to 'active')."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE campaigns SET status = ? WHERE id = ? AND user_id = ?", (status, campaign_id, user_id))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()

    def delete_campaign(self, campaign_id: int, user_id: int) -> bool:
        """Deletes a campaign."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM campaigns WHERE id = ? AND user_id = ?", (campaign_id, user_id))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()

    def get_templates_by_user(self, user_id: int) -> List[Dict]:
        """Retrieves all templates for a specific user."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT * FROM templates WHERE user_id = ? ORDER BY created_date DESC", (user_id,))
            templates = [dict(row) for row in cursor.fetchall()]
            return templates
        finally:
            conn.close()

    def add_template(self, user_id: int, data: Dict) -> Optional[int]:
        """Adds a new template to the database."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO templates (user_id, template_name, subject, body, type, tone, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                user_id,
                data['template_name'],
                data.get('subject', ''),
                data['body'],
                data.get('type', 'email'),
                data.get('tone', 'direct'),
                datetime.datetime.now().isoformat()
            ))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            logger.warning(f"Template '{data['template_name']}' already exists for user {user_id}.")
            return None
        finally:
            conn.close()

    def get_template_by_id(self, template_id: int, user_id: int) -> Optional[Dict]:
        """Retrieves a single template by its ID for a specific user."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT * FROM templates WHERE id = ? AND user_id = ?", (template_id, user_id))
            template = cursor.fetchone()
            return dict(template) if template else None
        except Exception as e:
            logger.error(f"Error fetching template {template_id} for user {user_id}: {e}")
            return None
        finally:
            conn.close()

    def update_template(self, template_id: int, user_id: int, data: Dict) -> bool:
        """Updates an existing template."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE templates SET
                    template_name = ?,
                    subject = ?,
                    body = ?,
                    type = ?,
                    tone = ?
                WHERE id = ? AND user_id = ?
            ''', (
                data['template_name'],
                data.get('subject', ''),
                data['body'],
                data.get('type', 'email'),
                data.get('tone', 'direct'),
                template_id,
                user_id
            ))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()

    def delete_template(self, template_id: int, user_id: int) -> bool:
        """Deletes a template from the database."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM templates WHERE id = ? AND user_id = ?", (template_id, user_id))
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Error deleting template {template_id} for user {user_id}: {e}")
            return False
        finally:
            conn.close()

    def add_global_contacts(self, contacts: List[Dict]) -> int:
        """Adds a list of contact dictionaries to the global contacts table, ignoring duplicates."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        inserted_count = 0
        try:
            for contact in contacts:
                try:
                    cursor.execute('''
                        INSERT OR IGNORE INTO global_hiring_contacts (full_name, title, company, email, linkedin_url, source, added_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        contact.get('full_name'), contact.get('title'), contact.get('company'),
                        contact.get('email'), contact.get('linkedin_url'),
                        contact.get('source'), contact.get('added_date')
                    ))
                    if cursor.rowcount > 0:
                        inserted_count += 1
                except sqlite3.IntegrityError:
                    continue
            conn.commit()
            return inserted_count # Return the count of newly inserted rows
        finally:
            conn.close()

    def search_and_copy_global_contacts(self, user_id: int, keywords: List[str]) -> int:
        """
        Searches global contacts for keywords and copies matches to a user's personal contacts.
        Returns the number of new contacts added.
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Build the search query for global contacts
        where_clauses = []
        params = []
        for keyword in keywords:
            like_pattern = f"%{keyword}%"
            where_clauses.append("(full_name LIKE ? OR title LIKE ? OR company LIKE ?)")
            params.extend([like_pattern, like_pattern, like_pattern])
        
        if not where_clauses:
            return 0

        search_query = f"SELECT * FROM global_hiring_contacts WHERE {' OR '.join(where_clauses)}"
        cursor.execute(search_query, params)
        matching_global_contacts = cursor.fetchall()

        added_count = 0
        for global_contact in matching_global_contacts:
            # global_contact is a tuple: (id, full_name, title, company, email, linkedin_url, source, added_date)
            contact_data = {'full_name': global_contact[1], 'title': global_contact[2], 'company': global_contact[3], 'email': global_contact[4], 'linkedin_url': global_contact[5], 'source': 'Internal Search'}
            if self.add_contact(user_id, contact_data):
                added_count += 1
        
        return added_count

    def add_contact(self, user_id: int, data: Dict) -> bool:
        """Adds a new contact, avoiding duplicates based on LinkedIn URL."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT OR IGNORE INTO contacts (user_id, full_name, title, company, email, linkedin_url, source, added_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                user_id,
                data['full_name'],
                data.get('title'),
                data.get('company'),
                data.get('email'),
                data['linkedin_url'],
                data.get('source'),
                datetime.datetime.now().isoformat()
            ))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()

    def get_contacts_by_user(self, user_id: int) -> List[Dict]:
        """Retrieves all contacts for a specific user."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT * FROM contacts WHERE user_id = ? ORDER BY added_date DESC", (user_id,))
            contacts = [dict(row) for row in cursor.fetchall()]
            return contacts
        finally:
            conn.close()

    def clear_global_contacts(self) -> int:
        """Deletes all records from the global_hiring_contacts table."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM global_hiring_contacts")
            deleted_count = cursor.rowcount
            conn.commit()
            return deleted_count
        finally:
            conn.close()

    def delete_contact(self, contact_id: int, user_id: int) -> bool:
        """Deletes a single personal contact for a specific user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM contacts WHERE id = ? AND user_id = ?", (contact_id, user_id))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()

    def delete_all_contacts_for_user(self, user_id: int) -> int:
        """Deletes all personal contacts for a specific user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM contacts WHERE user_id = ?", (user_id,))
            deleted_count = cursor.rowcount
            conn.commit()
            return deleted_count
        finally:
            conn.close()

    def get_campaigns_with_analytics(self, user_id: int) -> List[Dict]:
        """Retrieves all campaigns for a user, enriched with analytics from outreach_logs."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT * FROM campaigns WHERE user_id = ? ORDER BY created_date DESC", (user_id,))
            campaigns = [dict(row) for row in cursor.fetchall()]

            for campaign in campaigns:
                cursor.execute("SELECT status, COUNT(*) FROM outreach_logs WHERE campaign_id = ? GROUP BY status", (campaign['id'],))
                status_counts_raw = cursor.fetchall()
                
                status_counts = {
                    'sent': 0,
                    'opened': 0,
                    'replied': 0,
                    'ghosted': 0
                }
                for status, count in status_counts_raw:
                    if status in status_counts:
                        status_counts[status] = count
                
                total_sent = sum(status_counts.values())
                total_replied = status_counts.get('replied', 0)
                
                campaign['analytics'] = {
                    'total_sent': total_sent,
                    'replied_count': total_replied,
                    'opened_count': status_counts.get('opened', 0),
                    'reply_rate': (total_replied / total_sent * 100) if total_sent > 0 else 0
                }

            return campaigns
        finally:
            conn.close()


class EnhancedJobScraper:
    """Advanced job scraper with multiple sources and strategies."""
    
    def __init__(self, search_terms: List[str], location: Optional[str],
                 experience: Optional[str], job_type: Optional[str], user_id: int, task=None):
        self.search_terms = search_terms
        self.location = location
        self.experience = experience
        self.job_type = job_type
        self.user_id = user_id
        self.task = task
        self.db = JobDatabase()
        self.custom_scores = self.db.get_custom_scores(self.user_id)
        self.session = None  # Avoid shared Session across threads
        
        # Placeholder for common skills/tools/soft skills (can be expanded)
        self.common_skills = ['python', 'java', 'javascript', 'react', 'angular', 'vue', 'nodejs', 'express', 'django', 'flask', 'sql', 'nosql', 'mongodb', 'postgresql', 'mysql', 'aws', 'azure', 'gcp', 'docker', 'kubernetes', 'git', 'github', 'gitlab', 'jenkins', 'ci/cd', 'agile', 'scrum', 'rest api', 'graphql', 'html', 'css', 'redux', 'typescript', 'webpack', 'babel', 'selenium', 'jira', 'confluence', 'tableau', 'power bi', 'excel', 'gcp', 'azure', 'terraform', 'ansible', 'puppet', 'chef', 'splunk', 'elk stack', 'grafana', 'prometheus']
        self.soft_skills = ['communication', 'teamwork', 'problem-solving', 'leadership', 'adaptability', 'time management', 'critical thinking', 'creativity', 'interpersonal', 'collaboration', 'attention to detail', 'analytical']
        self.tools_platforms = ['jira', 'trello', 'asana', 'slack', 'microsoft teams', 'zoom', 'google workspace', 'salesforce', 'zendesk', 'servicenow', 'github actions', 'jenkins', 'gitlab ci']
        
        self.selenium_driver = None

    def _init_selenium_driver(self):
        """Initializes a Selenium WebDriver instance with randomized user-agent."""
        try:
            options = EdgeOptions()
            options.use_chromium = True
            options.add_argument('--headless=new')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            random_user_agent = random.choice(USER_AGENTS)
            options.add_argument(f"user-agent={random_user_agent}")
            
            # Optional: allow disabling via env (default disabled unless explicitly enabled)
            if os.environ.get('ENABLE_SELENIUM', '0') != '1':
                logger.info("Selenium disabled. Set ENABLE_SELENIUM=1 to enable.")
                self.selenium_driver = None
                return False

            driver_path = os.environ.get('EDGE_DRIVER_PATH', DRIVER_PATHS.get('edge'))
            if not driver_path or not os.path.exists(driver_path):
                logger.error("Edge WebDriver not found. Set EDGE_DRIVER_PATH env or update DRIVER_PATHS['edge'].")
                self.selenium_driver = None
                return False

            service = EdgeService(executable_path=driver_path)
            self.selenium_driver = webdriver.Edge(service=service, options=options)
            self.selenium_driver.set_page_load_timeout(30)
            logger.info("✅ Selenium WebDriver initialized successfully.")
            return True
        except FileNotFoundError:
            logger.error(f"WebDriver executable not found for Edge. Please check DRIVER_PATHS configuration.")
            self.selenium_driver = None
            return False
        except Exception as e:
            logger.error(f"An unexpected error occurred during WebDriver initialization: {e}", exc_info=True)
            self.selenium_driver = None
            return False

    def _close_selenium_driver(self):
        """Closes the Selenium WebDriver if it's open."""
        if self.selenium_driver:
            self.selenium_driver.quit()
            self.selenium_driver = None
            logger.info("Selenium WebDriver closed.")

    def _safe_find_element(self, by, value, timeout=10, el=None):
        """Safely finds a single element with an explicit wait."""
        context = el if el else self.selenium_driver
        if context is None: return None
        try:
            return WebDriverWait(context, timeout).until(
                EC.presence_of_element_located((by, value))
            )
        except (TimeoutException, SeleniumWebDriverException):
            logger.warning(f"Element not found within {timeout}s: {by}={value}")
            return None

    def _safe_find_elements(self, by, value, timeout=10, el=None):
        """Safely finds multiple elements with an explicit wait."""
        context = el if el else self.selenium_driver
        if context is None: return []
        try:
            return WebDriverWait(context, timeout).until(
                EC.presence_of_all_elements_located((by, value))
            )
        except (TimeoutException, SeleniumWebDriverException):
            logger.warning(f"No elements found within {timeout}s: {by}={value}")
            return []
            
    def _extract_experience_years(self, text: str) -> Tuple[Optional[int], Optional[int]]:
        """Extracts min and max experience years from a given text string."""
        min_exp, max_exp = None, None
        if not text:
            return min_exp, max_exp
        text_lower = text.lower()

        # Try explicit numeric patterns first
        match_range = re.search(r'(\d+)\s*-\s*(\d+)\s*(?:years?|yrs?)', text_lower)
        if match_range:
            min_exp = int(match_range.group(1))
            max_exp = int(match_range.group(2))
        else:
            match_plus = re.search(r'(\d+)\s*\+\s*(?:years?|yrs?)', text_lower)
            if match_plus:
                min_exp = int(match_plus.group(1))
                max_exp = None
            else:
                match_single = re.search(r'(\d+)\s*(?:years?|yrs?)', text_lower)
                if match_single:
                    min_exp = int(match_single.group(1))
                    max_exp = int(match_single.group(1))
                else:
                    # Heuristic fallbacks from seniority cues
                    if 'entry level' in text_lower or 'junior' in text_lower:
                        min_exp, max_exp = 0, 2
                    elif 'mid level' in text_lower or 'intermediate' in text_lower:
                        min_exp, max_exp = 3, 7
                    elif 'senior' in text_lower or 'lead' in text_lower or 'staff' in text_lower:
                        min_exp, max_exp = 8, None

        return min_exp, max_exp

    def _extract_keywords_from_description(self, description: str, keywords_list: List[str]) -> List[str]:
        """Extracts relevant keywords from job description."""
        found_keywords = []
        desc_lower = description.lower()
        for keyword in keywords_list:
            if keyword.lower() in desc_lower:
                found_keywords.append(keyword)
        return list(set(found_keywords))

    def _get_multiplier(self, key: str, types: List[str], default: float = 1.0) -> float:
        """Return the first matching custom score multiplier for the given key across possible types."""
        for t in types:
            val = self.custom_scores.get((key, t))
            if val is not None:
                return float(val)
        return default

    def _calculate_relevance_score(self, job: Job) -> float:
        """
        Calculates a relevance score for a job based on search terms,
        extracted keywords, experience level, and user-defined custom scores.
        """
        score = 0.0
        job_title_lower = job.title.lower()
        job_description_lower = job.description.lower()

        for term in self.search_terms:
            if term.lower() in job_title_lower:
                score += RELEVANCE_WEIGHTS['term_in_title'] * self._get_multiplier(term.lower(), ['keyword', 'keywords'])
            elif term.lower() in job_description_lower:
                score += RELEVANCE_WEIGHTS['term_in_desc'] * self._get_multiplier(term.lower(), ['keyword', 'keywords'])

        for skill in job.skills:
            score += RELEVANCE_WEIGHTS['skill'] * self._get_multiplier(skill.lower(), ['skill', 'skills'])
        for tool in job.extracted_tools:
            score += RELEVANCE_WEIGHTS['tool'] * self._get_multiplier(tool.lower(), ['tool', 'tools', 'tools_platforms'])
        for soft_skill in job.extracted_soft_skills:
            score += RELEVANCE_WEIGHTS['soft_skill'] * self._get_multiplier(soft_skill.lower(), ['soft_skill', 'soft_skills'])

        for word in re.findall(r'\b\w+\b', job_title_lower):
            score += RELEVANCE_WEIGHTS['title_word'] * self._get_multiplier(word, ['title_word', 'title'])

        if self.experience:
            user_min_exp, user_max_exp = self._extract_experience_years(self.experience)
            if user_min_exp is not None:
                if job.min_experience_years is not None and job.max_experience_years is not None:
                    if max(user_min_exp, job.min_experience_years) <= min(user_max_exp if user_max_exp is not None else float('inf'), job.max_experience_years if job.max_experience_years is not None else float('inf')):
                        score += RELEVANCE_WEIGHTS['exp_match']
                elif job.min_experience_years is not None and job.max_experience_years is None:
                    if user_min_exp >= job.min_experience_years:
                        score += RELEVANCE_WEIGHTS['exp_match']
                elif job.min_experience_years is None and user_max_exp is None:
                    score += RELEVANCE_WEIGHTS['exp_match'] * 0.5 # Partial match for open-ended
                elif job.min_experience_years is None:
                    exp_match_keywords = self._extract_keywords_from_description(job.experience, re.findall(r'\b\w+\b', self.experience.lower()))
                    if exp_match_keywords:
                        score += RELEVANCE_WEIGHTS['exp_partial']

        if self.location and self.location.lower() in job.location.lower():
            score += RELEVANCE_WEIGHTS['location']

        if self.job_type and self.job_type.lower() in job.job_type.lower():
            score += RELEVANCE_WEIGHTS['job_type']

        try:
            if 'ago' in job.posted_date.lower():
                num, unit = job.posted_date.lower().replace('posted', '').replace('ago', '').strip().split(' ')[:2]
                num = int(num)
                if 'hour' in unit:
                    posted_date = datetime.datetime.now() - datetime.timedelta(hours=num)
                elif 'day' in unit:
                    posted_date = datetime.datetime.now() - datetime.timedelta(days=num)
                elif 'week' in unit:
                    posted_date = datetime.datetime.now() - datetime.timedelta(weeks=num)
                elif 'month' in unit:
                    posted_date = datetime.datetime.now() - datetime.timedelta(days=num * 30)
                else:
                    posted_date = datetime.datetime.min
            else:
                posted_date = datetime.datetime.strptime(job.posted_date, '%Y-%m-%d')
        except (ValueError, AttributeError):
            posted_date = datetime.datetime.min

        days_ago = (datetime.datetime.now() - posted_date).days
        if days_ago <= 7:
            score *= RECENCY_MULTIPLIERS['week']
        elif days_ago <= 30:
            score *= RECENCY_MULTIPLIERS['month']

        return round(score, 2)


    def scrape_linkedin_jobs(self, term: str) -> List[Job]:
        """
        Updated LinkedIn scraper: robust, stealthy, paginated, and handles anti-bot measures.
        """
        logger.info(f"[LinkedIn] Scraping for term '{term}' in {self.location or 'All Locations'}...")
        jobs = []
        seen_links = set()
        max_pages = 3  # Limit pages to avoid throttling
        base_url = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
        proxies = None  # Add proxy support if needed

        for page in range(max_pages):
            logger.info(f"[LinkedIn] --- Starting scrape for page {page + 1} of {max_pages} for term '{term}' ---")
            params = {
                'keywords': term,
                'location': self.location,
                'start': page * 25,
                'count': 25
            }
            headers = {'User-Agent': random.choice(USER_AGENTS)}
            try:
                logger.info(f"[LinkedIn] Sending request to LinkedIn for job list (page {page + 1})...")
                time.sleep(random.uniform(2, 5))  # Random delay to avoid detection
                response = requests.get(base_url, params=params, headers=headers, timeout=15, proxies=proxies)

                if response.status_code != 200:
                    logger.warning(f"[LinkedIn] LinkedIn returned status {response.status_code} on page {page + 1}. Stopping further requests.")
                    break

                if b'captcha' in response.content.lower() or b'login' in response.content.lower():
                    logger.warning(f"[LinkedIn] Bot detection or login wall hit on page {page + 1}. Stopping scrape for this term.")
                    break

                try:
                    soup = BeautifulSoup(response.content, 'lxml')
                except Exception:
                    soup = BeautifulSoup(response.content, 'html.parser')

                # More robust selector for job cards
                job_cards = soup.select('div.job-card-container, div.base-card, li.job-result-card')
                if not job_cards:
                    logger.info(f"[LinkedIn] No job cards found on page {page + 1}. Stopping for this term.")
                    break

                logger.info(f"[LinkedIn] Found {len(job_cards)} job cards on page {page + 1}.")
                job_details = []

                for card in job_cards:
                    try:
                        # More robust selectors with fallbacks
                        title_elem = card.select_one('h3.base-search-card__title, h3.job-card-list__title')
                        title = title_elem.text.strip() if title_elem else 'N/A'

                        company_elem = card.select_one('h4.base-search-card__subtitle, h4.job-card-container__company-name')
                        company = company_elem.text.strip() if company_elem else 'N/A'

                        location_elem = card.select_one('span.job-search-card__location, span.job-card-container__metadata-item')
                        location = location_elem.text.strip() if location_elem else 'N/A'

                        link_elem = card.select_one('a.base-card__full-link, a.job-card-container__link')
                        link = link_elem.get('href') if link_elem and link_elem.get('href') else 'N/A'
                        # Add a fallback for the link
                        if link == 'N/A':
                            link_elem = card.find('a', href=True)
                            if link_elem:
                                link = link_elem.get('href')

                        if link in seen_links or link == 'N/A':
                            continue
                        seen_links.add(link)

                        posted_date_elem = card.find('time')
                        posted_date = posted_date_elem.text.strip() if posted_date_elem else 'N/A'

                        job_details.append({
                            'title': title,
                            'company': company,
                            'location': location,
                            'link': link,
                            'posted_date': posted_date
                        })
                    except Exception as e:
                        logger.warning(f"[LinkedIn] Error extracting job card: {e}", exc_info=True)
                        continue

                logger.info(f"[LinkedIn] Now fetching full job descriptions for {len(job_details)} jobs on page {page + 1}...")

                def fetch_description(job):
                    try:
                        logger.info(f"[LinkedIn] Fetching job detail page: {job['link']}")
                        detail_headers = {'User-Agent': random.choice(USER_AGENTS)}
                        time.sleep(random.uniform(2.5, 5.5)) # Increased delay to avoid 429 errors
                        resp = requests.get(job['link'], headers=detail_headers, timeout=15)

                        if resp.status_code == 200:
                            detail_soup = BeautifulSoup(resp.content, 'lxml')
                            desc_elem = detail_soup.find('div', class_=lambda x: x and 'description' in x)
                            description = desc_elem.get_text(separator=' ', strip=True) if desc_elem else 'N/A'
                            logger.info(f"[LinkedIn] Successfully fetched description for job: {job['title']} at {job['company']}")
                        else:
                            description = 'N/A'
                            logger.warning(f"[LinkedIn] Failed to fetch job detail page (status {resp.status_code}) for {job['link']}")
                    except Exception as e:
                        logger.warning(f"[LinkedIn] Error fetching job detail page: {e}")
                        description = 'N/A'

                    job['description'] = description
                    job['source'] = 'LinkedIn'
                    return job

                with concurrent.futures.ThreadPoolExecutor(max_workers=LINKEDIN_DETAIL_MAX_WORKERS) as executor:
                    enriched_jobs = list(executor.map(fetch_description, job_details))

                logger.info(f"[LinkedIn] Finished fetching all job descriptions for page {page + 1}.")
                for job in enriched_jobs:
                    jobs.append(self._create_job_object_from_raw_data(job))

                if len(job_cards) < 10:
                    logger.info(f"[LinkedIn] Less than 10 jobs found on page {page + 1}, assuming last page.")
                    break
            except requests.exceptions.RequestException as e:
                logger.error(f"[LinkedIn] Network error on page {page + 1} for '{term}': {e}", exc_info=True)
                continue
            except Exception as e:
                logger.error(f"[LinkedIn] Unexpected error on page {page + 1} for '{term}': {e}", exc_info=True)
                continue

        logger.info(f"[LinkedIn] ✅ Finished scraping. Total jobs scraped for '{term}': {len(jobs)}.")
        return jobs
    
    def scrape_internshala_jobs(self, term: str) -> List[Job]:
        """
        Enhanced Internshala scraper with improved resilience, better error handling,
        and support for the latest website structure (2025).
        """
        logger.info(f"[Internshala] Starting enhanced scrape for term '{term}' in {self.location or 'All Locations'}...")
        jobs = []
        seen_links = set()
        max_pages = 5  # Increased to get more results
        
        # Enhanced URL generation with better slug handling
        term_slug = self._create_url_slug(term)
        base_url = "https://internshala.com/jobs/"
        
        # Determine URL format based on location
        url_params = self._build_url_params(term_slug)
        
        if self.location:
            is_remote = self.location.lower() in ['remote', 'work from home', 'wfh', 'online']
            
            if is_remote:
                base_url += f"work-from-home/{term_slug}/"
                logger.info(f"[Internshala] Using remote job URL: {base_url}")
            else:
                # Handle location-specific searches
                primary_location = self.location.split(',')[0].strip()
                location_slug = self._create_url_slug(primary_location)
                base_url += f"{term_slug}-in-{location_slug}/"
                logger.info(f"[Internshala] Using location-based URL: {base_url}")
        else:
            base_url += f"{term_slug}/"
            logger.info(f"[Internshala] Using general job URL: {base_url}")
        
        job_details_from_list = []
        
        # Enhanced session with better configuration
        with requests.Session() as session:
            session.headers.update({
                'User-Agent': random.choice(USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                'Accept-Language': 'en-US,en;q=0.9,hi;q=0.8',
                'Accept-Encoding': 'gzip, deflate, br',
                'Cache-Control': 'no-cache',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'same-origin'
            })
            
            # Set initial referer
            referer = 'https://internshala.com/'
            
            for page in range(1, max_pages + 1):
                url = f"{base_url}?page={page}" if page > 1 else base_url
                
                try:
                    logger.info(f"[Internshala] Scraping page {page}: {url}")
                    session.headers['Referer'] = referer
                    
                    # Randomized delay to avoid detection
                    time.sleep(random.uniform(2.5, 5.0))
                    
                    response = session.get(url, timeout=20)
                    referer = url
                    
                    if response.status_code == 404:
                        logger.warning(f"[Internshala] Page not found (404): {url}")
                        break
                    elif response.status_code == 403:
                        logger.warning(f"[Internshala] Access forbidden (403). Possible rate limiting.")
                        time.sleep(10)  # Wait longer if blocked
                        continue
                    elif response.status_code != 200:
                        logger.warning(f"[Internshala] HTTP {response.status_code} on page {page}")
                        break
                    
                    soup = BeautifulSoup(response.content, 'lxml')
                    
                    # Enhanced page validation
                    if self._is_no_results_page(soup):
                        logger.info(f"[Internshala] No results found on page {page}")
                        break
                    
                    # Multiple selectors for job cards (fallback approach)
                    job_cards = self._extract_job_cards(soup)
                    
                    if not job_cards:
                        logger.info(f"[Internshala] No job cards found on page {page}")
                        break
                    
                    logger.info(f"[Internshala] Found {len(job_cards)} job cards on page {page}")
                    
                    # Process each job card
                    for card in job_cards:
                        job_data = self._parse_job_card(card, seen_links)
                        if job_data:
                            job_details_from_list.append(job_data)
                    
                except requests.exceptions.Timeout:
                    logger.error(f"[Internshala] Timeout on page {page}")
                    break
                except requests.exceptions.RequestException as e:
                    logger.error(f"[Internshala] Network error on page {page}: {e}")
                    break
                except Exception as e:
                    logger.error(f"[Internshala] Unexpected error while scraping list page {page}: {e}", exc_info=True)
                    break
    
        if not job_details_from_list:
            logger.info(f"[Internshala] No jobs found for term '{term}'")
            return []
    
        logger.info(f"[Internshala] Found {len(job_details_from_list)} unique jobs. Fetching details...")
    
        # Fetch job details with enhanced concurrency
        with concurrent.futures.ThreadPoolExecutor(max_workers=SCRAPE_MAX_WORKERS) as executor:
            future_to_job = {
                executor.submit(self._fetch_job_details, job_data): job_data 
                for job_data in job_details_from_list
            }
            
            for future in concurrent.futures.as_completed(future_to_job):
                try:
                    enriched_job = future.result()
                    if enriched_job:
                        jobs.append(self._create_job_object_from_raw_data(enriched_job))
                except Exception as e:
                    logger.error(f"[Internshala] Error processing job detail: {e}")
    
        logger.info(f"[Internshala] ✅ Enhanced scrape completed. Total jobs: {len(jobs)}")
        return jobs

    def _create_url_slug(self, text: str) -> str:
        """Create URL-friendly slug from text."""
        slug = text.lower().strip()
        slug = re.sub(r'[^\w\s-]', '', slug)
        slug = re.sub(r'[\s_-]+', '-', slug)
        return slug.strip('-')

    def _build_url_params(self, term_slug: str) -> dict:
        """Build URL parameters for search."""
        return {
            'category': '',
            'type': 'job',
            'location': self.location or '',
            'search': term_slug
        }

    def _is_no_results_page(self, soup: BeautifulSoup) -> bool:
        """Check if page shows no results."""
        no_result_indicators = [
            '.no_internship_found',
            '.no-results',
            '.empty-state',
            '[data-testid="no-results"]'
        ]
        
        for selector in no_result_indicators:
            if soup.select_one(selector):
                return True
        
        # Check for text indicators
        page_text = soup.get_text().lower()
        if any(phrase in page_text for phrase in ['no jobs found', 'no results', 'no internships found']):
            return True
        
        return False

    def _extract_job_cards(self, soup: BeautifulSoup) -> list:
        """Extract job cards using multiple selectors."""
        selectors = [
            '.internship_meta',
            '.job_container',
            '.individual_internship',
            '[data-testid="job-card"]',
            '.search_results .container-fluid > div'
        ]
        
        for selector in selectors:
            cards = soup.select(selector)
            if cards:
                logger.debug(f"[Internshala] Found job cards using selector: {selector}")
                return cards
        
        return []

    def _parse_job_card(self, card: BeautifulSoup, seen_links: set) -> Optional[Dict]:
        """Parse individual job card."""
        try:
            # Enhanced link extraction
            link_selectors = [
                '.profile a',
                '.view_detail_button',
                '.job_title a',
                'a[href*="/job/"]',
                'a[href*="/jobs/detail/"]'
            ]
            
            link_elem = None
            for selector in link_selectors:
                link_elem = card.select_one(selector)
                if link_elem and link_elem.get('href'):
                    break
            
            if not link_elem or not link_elem.get('href'):
                return None
            
            href = link_elem['href']
            link = urljoin('https://internshala.com', href)
            
            if link in seen_links:
                return None
            seen_links.add(link)
            
            # Enhanced title extraction
            title_selectors = ['.profile a', '.job_title', '.heading_4_5 a', 'h3 a', 'h4 a']
            title = self._extract_text_by_selectors(card, title_selectors) or 'N/A'
            
            # Enhanced company extraction
            company_selectors = ['.company_name', '.company a', '.subtitle', '[data-testid="company-name"]']
            company = self._extract_text_by_selectors(card, company_selectors) or 'N/A'
            
            # Enhanced location extraction
            location_selectors = ['#location_names', '.location', '.job_location', '[data-testid="location"]']
            location = self._extract_text_by_selectors(card, location_selectors) or 'N/A'
            
            # Extract salary if available on card
            salary_selectors = ['.stipend', '.salary', '.package', '[data-testid="salary"]']
            salary = self._extract_text_by_selectors(card, salary_selectors) or 'N/A'
            
            return {
                'title': title,
                'company': company,
                'location': location,
                'salary': salary,
                'link': link,
            }
            
        except Exception as e:
            logger.warning(f"[Internshala] Error parsing job card: {e}")
            return None

    def _extract_text_by_selectors(self, element: BeautifulSoup, selectors: List[str]) -> Optional[str]:
        """Extract text using multiple selector fallbacks."""
        for selector in selectors:
            elem = element.select_one(selector)
            if elem:
                text = elem.get_text(strip=True)
                if text and text != 'N/A':
                    return text
        return None

    def _fetch_job_details(self, job_data: Dict) -> Optional[Dict]:
        """Fetch detailed job information."""
        link = job_data['link']
        
        try:
            logger.debug(f"[Internshala] Fetching details: {link}")
            
            # Randomized delay
            time.sleep(random.uniform(1.5, 3.5))
            
            headers = {
                'User-Agent': random.choice(USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Referer': 'https://internshala.com/jobs/'
            }
            
            response = requests.get(link, headers=headers, timeout=20)
            
            if response.status_code != 200:
                logger.warning(f"[Internshala] Failed to fetch {link} (status: {response.status_code})")
                return None
            
            soup = BeautifulSoup(response.content, 'lxml')
            
            # Enhanced description extraction
            description = self._extract_description(soup)
            
            # Enhanced salary extraction (if not already found)
            if job_data.get('salary', 'N/A') == 'N/A':
                salary_selectors = ['.stipend', '.salary_container', '.ctc', '.package_container']
                job_data['salary'] = self._extract_text_by_selectors(soup, salary_selectors) or 'N/A'
            
            # Extract posted date
            posted_date = self._extract_posted_date(soup)
            
            # Extract skills/requirements
            skills = self._extract_skills(soup)
            
            # Extract experience
            experience = self._extract_experience(soup)
            
            # Extract job type
            job_type = self._extract_job_type(soup)
            
            # Extract application deadline
            deadline = self._extract_deadline(soup)
            
            job_data.update({
                'description': description,
                'posted_date': posted_date,
                'skills': skills,
                'experience': experience,
                'job_type': job_type,
                'deadline': deadline,
                'source': 'Internshala'
            })
            
            return job_data
            
        except Exception as e:
            logger.error(f"[Internshala] Error fetching details for {link}: {e}")
            return None

    def _create_job_object_from_raw_data(self, raw_data: Dict) -> Job:
        """Helper to create a Job object and calculate its relevance score."""
        salary = raw_data.get('salary', '')
        experience = raw_data.get('experience', '')

        # Use combined textual context (title, description, raw experience) for extraction
        title_txt = raw_data.get('title', '') or ''
        desc_txt = raw_data.get('description', '') or ''
        exp_txt = raw_data.get('experience', '') or ''
        combined_text = ' '.join([t for t in [title_txt, desc_txt, exp_txt] if t]).strip()
        if combined_text and (not salary or salary == 'N/A'):
            salary = self._extract_salary(combined_text)
        
        job_obj = Job(
            title=raw_data.get('title', ''),
            company=raw_data.get('company', ''),
            location=raw_data.get('location', ''),
            salary=salary,
            link=raw_data.get('link', ''),
            description=raw_data.get('description', ''),
            keywords=self.search_terms,
            skills=raw_data.get('skills') or self._extract_keywords_from_description(combined_text or desc_txt, self.common_skills),
            experience=experience,
            job_type=raw_data.get('job_type', 'Full-time'),
            posted_date=raw_data.get('posted_date', ''),
            source=raw_data.get('source', ''),
            deadline=raw_data.get('deadline', 'N/A')
        )
        # Derive experience years and keyword families from a broader context for robustness
        context_for_years = combined_text or job_obj.description
        job_obj.min_experience_years, job_obj.max_experience_years = self._extract_experience_years(context_for_years)
        job_obj.extracted_tools = self._extract_keywords_from_description(combined_text or job_obj.description, self.tools_platforms)
        job_obj.extracted_soft_skills = self._extract_keywords_from_description(combined_text or job_obj.description, self.soft_skills)
        job_obj.relevance_score = self._calculate_relevance_score(job_obj)
        return job_obj
        
    def _extract_salary(self, text: str) -> str:
        """Extracts salary information from text using regex patterns."""
        if not text:
            return "N/A"
        patterns = [
            # ₹ 30,00,000 - 33,00,000 /year or /month or per annum
            r'₹\s*[\d,]+(?:\s*-\s*[\d,]+)?\s*(?:/\s*year|/\s*month|per\s*annum|pa)?',
            # Rs. 15,00,000 - 20,00,000 per annum
            r'rs\.?\s*[\d,]+(?:\s*-\s*[\d,]+)?\s*(?:/\s*year|/\s*month|per\s*annum|pa)?',
            # 12 - 18 LPA, 12.5-15 lpa
            r'(?:\d+(?:\.\d+)?)\s*-\s*(?:\d+(?:\.\d+)?)\s*(?:lpa|lakhs?)',
            # CTC: ₹ 20,00,000 - 30,00,000
            r'ctc\s*[:\-]?\s*₹?\s*[\d,]+(?:\s*-\s*[\d,]+)?',
        ]
        for pattern in patterns:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                return m.group(0).strip()
        return "N/A"

    def _extract_description(self, soup: BeautifulSoup) -> str:
        """Extract job description with multiple fallbacks."""
        description_selectors = [
            '.internship_details .text-container',
            '.job_description',
            '.description_container',
            '.internship_details',
            '[data-testid="job-description"]'
        ]
        
        for selector in description_selectors:
            elem = soup.select_one(selector)
            if elem:
                description = elem.get_text(separator='\n', strip=True)
                if description and len(description) > 10:
                    return description
        
        return 'N/A'

    def _extract_posted_date(self, soup: BeautifulSoup) -> str:
        """Extract posted date."""
        date_selectors = [
            '.posted_on_container .status-container',
            '.posted_date',
            '.job_posted_date',
            '[data-testid="posted-date"]'
        ]
        
        posted_date = self._extract_text_by_selectors(soup, date_selectors)
        
        if posted_date:
            # Clean up posted date text
            posted_date = re.sub(r'Posted on\s*', '', posted_date, flags=re.IGNORECASE)
            posted_date = posted_date.strip()
        
        return posted_date or 'N/A'

    def _extract_skills(self, soup: BeautifulSoup) -> List[str]:
        """Extract required skills."""
        skills_selectors = [
            '.round_tabs_container .round_tabs',
            '.skills_required .skill',
            '.requirements .skill',
            '[data-testid="skill-tag"]'
        ]
        
        skills = []
        for selector in skills_selectors:
            skill_elements = soup.select(selector)
            if skill_elements:
                skills = [elem.get_text(strip=True) for elem in skill_elements]
                break
        
        return [skill for skill in skills if skill and skill.strip()]

    def _extract_experience(self, soup: BeautifulSoup) -> str:
        """Extract experience requirements."""
        # Look for experience section
        experience_headings = soup.find_all(['h3', 'h4', 'h5'], 
                                          string=re.compile(r'Experience|Experience Required', re.I))
        
        for heading in experience_headings:
            next_elem = heading.find_next_sibling(['div', 'p', 'span'])
            if next_elem:
                experience_text = next_elem.get_text(strip=True)
                if experience_text:
                    return experience_text
        
        # Fallback selectors
        exp_selectors = [
            '.experience_required',
            '.experience_container',
            '[data-testid="experience"]'
        ]
        
        return self._extract_text_by_selectors(soup, exp_selectors) or 'N/A'

    def _extract_job_type(self, soup: BeautifulSoup) -> str:
        """Extract job type (Full-time, Part-time, etc.)."""
        type_selectors = [
            '.job_type',
            '.employment_type',
            '.type_container',
            '[data-testid="job-type"]'
        ]
        
        job_type = self._extract_text_by_selectors(soup, type_selectors)
        
        if not job_type:
            # Infer from URL or content
            url = soup.find('link', {'rel': 'canonical'})
            if url and 'internship' in url.get('href', '').lower():
                return 'Internship'
            else:
                return 'Full-time'  # Default assumption for jobs
        
        return job_type

    def _extract_deadline(self, soup: BeautifulSoup) -> str:
        """Extract application deadline."""
        deadline_selectors = [
            '.application_deadline',
            '.last_date',
            '.deadline_container',
            '[data-testid="deadline"]'
        ]
        
        return self._extract_text_by_selectors(soup, deadline_selectors) or 'N/A'

    def scrape_all_sources(self) -> List[Job]:
        """Scrapes jobs from all configured sources concurrently."""
        all_jobs = []
        scraper_methods = {
            'LinkedIn': self.scrape_linkedin_jobs,
            'Internshala': self.scrape_internshala_jobs
            # To add a new scraper, simply add its name and method here.
            # 'NewSource': self.scrape_newsource_jobs
        }

        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': 'Initializing scrapers for multiple sources...', 'progress': 10})

        total_scrapers = len(self.search_terms) * len(scraper_methods)
        scrapers_done = 0
        progress_start = 15 # Start after initialization
        progress_range = 70 # Scraping happens between 10% and 80%

        with concurrent.futures.ThreadPoolExecutor(max_workers=SCRAPE_MAX_WORKERS) as executor:
            future_to_source = {}
            for name, method in scraper_methods.items():
                for term in self.search_terms:
                    future = executor.submit(method, term)
                    future_to_source[future] = f"{name}-{term}"
            
            for future in concurrent.futures.as_completed(future_to_source):
                source_info = future_to_source[future]
                try:
                    jobs_from_source = future.result()
                    all_jobs.extend(jobs_from_source)
                    logger.info(f"Finished scraping {source_info}. Found {len(jobs_from_source)} jobs.")
                    
                    scrapers_done += 1
                    progress = progress_start + int((scrapers_done / total_scrapers) * progress_range)
                    if self.task: # More detailed progress
                        self.task.update_state(state='PROGRESS', meta={'status': f'Scraped {len(jobs_from_source)} jobs from {source_info.split("-")[0]} for "{source_info.split("-")[1]}"...', 'progress': progress})

                except Exception as exc:
                    logger.error(f'{source_info} generated an exception: {exc}', exc_info=True)
                    
        logger.info(f"Total jobs scraped across all sources before deduplication: {len(all_jobs)}.")
        
        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': f'Deduplicating and scoring {len(all_jobs)} jobs...', 'progress': 85})

        unique_jobs = {}
        for job in all_jobs:
            norm = f"{normalize_for_hash(job.title)}|{normalize_for_hash(job.company)}|{normalize_for_hash(job.location)}"
            key = hashlib.md5(norm.encode('utf-8')).hexdigest()
            if key not in unique_jobs or job.relevance_score > unique_jobs[key].relevance_score:
                unique_jobs[key] = job
        
        filtered_jobs = list(unique_jobs.values())
        filtered_jobs = [job for job in filtered_jobs if job.title and job.title != 'N/A' and job.link and job.link != 'N/A']

        # Location filtering removed: all jobs are included regardless of location.

        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': f'Filtering {len(filtered_jobs)} unique jobs by relevance...', 'progress': 88})

        high_relevance_jobs = [job for job in filtered_jobs if job.relevance_score >= RELEVANCE_MIN_THRESHOLD]
        high_relevance_jobs.sort(key=lambda x: x.relevance_score, reverse=True)
        
        logger.info(f"Jobs count after filtering by relevance_score >= 10: {len(high_relevance_jobs)}.")
        
        source_counts = Counter(job.source for job in high_relevance_jobs)
        for source, count in source_counts.items():
            logger.info(f"Final jobs from {source}: {count}.")
            
        return high_relevance_jobs[:50]

class ApplicationFiller:
    """Automated job application filler for various ATS platforms"""
    
    def __init__(self, driver: webdriver.Remote, user_details: Dict):
        self.driver = driver
        self.user_details = user_details
        self.wait = WebDriverWait(driver, 10)
        # Extract parsed data and preferences for easier access
        self.resume_data = self.user_details.get('parsed_resume', {})
        self.preferences = {
            'willing_to_relocate': self.user_details.get('willing_to_relocate', 'no'),
            'authorized_to_work': self.user_details.get('authorized_to_work', 'no')
        }
        
    def fill(self, url: str) -> bool:
        """Fill application form based on URL pattern"""
        try:
            self.driver.get(url)
            time.sleep(2)  # Allow page to load
            
            # This is a generic flag to indicate if we should try answering questions.
            # Specific implementations for Lever, Greenhouse etc. will call this.
            should_answer_questions = True
            
            if "greenhouse.io" in url:
                return self.fill_greenhouse()
            elif "jobs.lever.co" in url:
                return self.fill_lever()
            elif "workday.com" in url or "myworkdayjobs.com" in url:
                return self.fill_workday()
            elif "bamboohr.com" in url:
                return self.fill_bamboohr()
            elif "smartrecruiters.com" in url:
                return self.fill_smartrecruiters()
            elif "jobvite.com" in url:
                return self.fill_jobvite()
            else:
                # For unknown sites, we still try the generic fill and question answering
                filled_generic = self.fill_generic()
                if filled_generic and should_answer_questions:
                    self.answer_custom_questions()
                logger.info(f"Attempting generic form fill for: {url}")
                return self.fill_generic()
                
        except Exception as e:
            logger.error(f"Error filling application at {url}: {str(e)}")
            return False

    def safe_send_keys(self, element, text: str) -> bool:
        """Safely send keys to an element"""
        try:
            if element and text:
                element.clear()
                element.send_keys(text)
                return True
        except Exception as e:
            logger.warning(f"Failed to send keys: {str(e)}")
        return False

    def find_element_by_multiple_selectors(self, selectors: List[str], timeout: int = 5):
        """Try multiple selectors to find an element"""
        for selector in selectors:
            try:
                if selector.startswith('//'):
                    element = WebDriverWait(self.driver, timeout).until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )
                elif selector.startswith('#'):
                    element = WebDriverWait(self.driver, timeout).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                else: # Default to CSS_SELECTOR for attribute selectors like 'input[name*="..."]'
                    element = WebDriverWait(self.driver, timeout).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                return element
            except TimeoutException:
                continue
        return None

    def fill_greenhouse(self) -> bool:
        """Fill Greenhouse application forms"""
        logger.info("Filling Greenhouse application...")
        
        try:
            # Common Greenhouse field mappings
            field_mappings = {
                'first_name': ['first_name', 'firstName', '#first_name'],
                'last_name': ['last_name', 'lastName', '#last_name'],
                'email': ['email', 'email_address', '#email'],
                'phone': ['phone', 'phone_number', '#phone'],
                'resume': ['resume', 'resume_file', 'input[type="file"]'],
                'cover_letter': ['cover_letter', 'coverLetter', 'cover_letter_text'],
                'linkedin': ['linkedin', 'linkedin_url', '#linkedin']
            }
            
            # Fill basic fields
            for field_key, selectors in field_mappings.items():
                if field_key in self.user_details:
                    element = self.find_element_by_multiple_selectors(selectors)
                    if element:
                        if field_key == 'resume' and element.get_attribute('type') == 'file':
                            element.send_keys(self.user_details[field_key])
                        else:
                            self.safe_send_keys(element, self.user_details[field_key])
            
            self.handle_dropdowns()
            self.answer_custom_questions()
            self.fill_custom_questions()
            return True
        except Exception as e:
            logger.error(f"Error in Greenhouse form fill: {str(e)}")
            return False

    def fill_lever(self) -> bool:
        """Fill Lever application forms"""
        logger.info("Filling Lever application...")
        try:
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="name"]']), self.user_details.get('full_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="email"]']), self.user_details.get('email', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="phone"]']), self.user_details.get('phone', ''))
            resume_upload = self.find_element_by_multiple_selectors(['input[name="resume"]', 'input[type="file"]'])
            if resume_upload and 'resume' in self.user_details:
                resume_upload.send_keys(self.user_details['resume'])
            self.answer_custom_questions()
            return True
        except Exception as e:
            logger.error(f"Error in Lever form fill: {str(e)}")
            return False

    def fill_workday(self) -> bool:
        """Fill Workday application forms"""
        logger.info("Filling Workday application...")
        try:
            time.sleep(3)
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[data-automation-id="firstName"]']), self.user_details.get('first_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[data-automation-id="lastName"]']), self.user_details.get('last_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[data-automation-id="email"]']), self.user_details.get('email', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[data-automation-id="phone"]']), self.user_details.get('phone', ''))
            self.answer_custom_questions()
            return True
        except Exception as e:
            logger.error(f"Error in Workday form fill: {str(e)}")
            return False

    def fill_bamboohr(self) -> bool:
        logger.info("Filling BambooHR application...")
        try:
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name*="firstName"]']), self.user_details.get('first_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name*="lastName"]']), self.user_details.get('last_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name*="email"]']), self.user_details.get('email', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name*="phone"]']), self.user_details.get('phone', ''))
            self.answer_custom_questions()
            return True
        except Exception as e:
            logger.error(f"Error in BambooHR form fill: {str(e)}")
            return False

    def fill_smartrecruiters(self) -> bool:
        logger.info("Filling SmartRecruiters application...")
        try:
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="firstName"]']), self.user_details.get('first_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="lastName"]']), self.user_details.get('last_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="email"]']), self.user_details.get('email', ''))
            self.answer_custom_questions()
            return True
        except Exception as e:
            logger.error(f"Error in SmartRecruiters form fill: {str(e)}")
            return False

    def fill_jobvite(self) -> bool:
        logger.info("Filling Jobvite application...")
        try:
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="firstName"]']), self.user_details.get('first_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="lastName"]']), self.user_details.get('last_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="email"]']), self.user_details.get('email', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name="phone"]']), self.user_details.get('phone', ''))
            self.answer_custom_questions()
            return True
        except Exception as e:
            logger.error(f"Error in Jobvite form fill: {str(e)}")
            return False

    def fill_generic(self) -> bool:
        logger.info("Attempting generic form fill...")
        try:
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name*="first"]', 'input[placeholder*="first name"]']), self.user_details.get('first_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name*="last"]', 'input[placeholder*="last name"]']), self.user_details.get('last_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[name*="full"]', 'input[placeholder*="full name"]']), self.user_details.get('full_name', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[type="email"]', 'input[name*="email"]']), self.user_details.get('email', ''))
            self.safe_send_keys(self.find_element_by_multiple_selectors(['input[type="tel"]', 'input[name*="phone"]']), self.user_details.get('phone', ''))
            return True
        except Exception as e:
            logger.error(f"Error in generic form fill: {str(e)}")
            return False

    def handle_dropdowns(self):
        try:
            if 'experience_level' in self.user_details:
                exp_dropdown = self.find_element_by_multiple_selectors(['select[name*="experience"]', 'select[id*="experience"]'])
                if exp_dropdown: Select(exp_dropdown).select_by_visible_text(self.user_details['experience_level'])
            if 'country' in self.user_details:
                country_dropdown = self.find_element_by_multiple_selectors(['select[name*="country"]', 'select[id*="country"]'])
                if country_dropdown: Select(country_dropdown).select_by_visible_text(self.user_details['country'])
        except Exception as e:
            logger.warning(f"Error handling dropdowns: {str(e)}")

    def fill_custom_questions(self):
        try:
            for textarea in self.driver.find_elements(By.TAG_NAME, "textarea"):
                label = self.get_field_label(textarea)
                if label and 'custom_answers' in self.user_details:
                    answer = self.user_details.get('custom_answers', {}).get(label.lower(), self.user_details.get('default_answer', ''))
                    self.safe_send_keys(textarea, answer)
        except Exception as e:
            logger.warning(f"Error filling custom questions: {str(e)}")

    def get_field_label(self, element):
        try:
            field_id = element.get_attribute('id')
            if field_id: return self.driver.find_element(By.CSS_SELECTOR, f'label[for="{field_id}"]').text.strip()
            return element.find_element(By.XPATH, '..').text.strip()
        except: return None

    def answer_custom_questions(self):
        """Finds and answers custom questions on a page."""
        logger.info("[AssistApply] Searching for custom questions to answer...")
        # Generic selectors for question containers. LinkedIn uses the 'jobs-easy-apply-form-section__grouping' class.
        question_containers = self.driver.find_elements(By.XPATH, "//div[contains(@class, 'form-question') or contains(@class, 'jobs-easy-apply-form-section__grouping') or contains(@class, 'fb-form-element')]")
        
        if not question_containers:
            logger.info("[AssistApply] No custom question containers found.")
            return

        for container in question_containers:
            try:
                label_el = container.find_element(By.TAG_NAME, 'label')
                question_text = label_el.text
                if not question_text:
                    continue

                answer = self._generate_answer(question_text)
                if answer is None:
                    logger.info(f"[AssistApply] No answer generated for question: '{question_text}'")
                    continue

                # Find the input/select/textarea associated with the label
                input_id = label_el.get_attribute('for')
                if input_id:
                    input_el = self.driver.find_element(By.ID, input_id)
                else: # Fallback for inputs not linked by 'for'
                    input_el = container.find_element(By.CSS_SELECTOR, 'input, select, textarea')

                # Now fill the element based on its type
                if input_el.tag_name == 'select':
                    self._select_answer(input_el, answer)
                elif input_el.get_attribute('type') in ['radio', 'checkbox']:
                    self._check_answer(container, answer)
                else: # Text input or textarea
                    self.safe_send_keys(input_el, answer)
                
                logger.info(f"[AssistApply] Answered '{question_text}' with '{answer}'")
                time.sleep(0.5) # Small delay after answering

            except NoSuchElementException:
                continue # Skip if a container doesn't have the expected structure
            except Exception as e:
                logger.warning(f"[AssistApply] Error processing a question container: {e}")

    def _generate_answer(self, question_text: str) -> Optional[str]:
        """Generates an answer based on question text and user data."""
        q_lower = question_text.lower()

        # Years of experience
        match = re.search(r'(?:how many|how much) years of experience.* with ([\w\s\+\#\.\-]+)\??', q_lower)
        if match:
            skill = match.group(1).strip().replace('.', r'\.') # Sanitize for regex
            return self._calculate_years_for_skill(skill)

        # Relocation
        if 'relocate' in q_lower or 'commute' in q_lower or 'relocation' in q_lower:
            return 'Yes' if self.preferences['willing_to_relocate'] == 'yes' else 'No'

        # Work Authorization & Sponsorship
        # Check for sponsorship first, as it's a more specific keyword.
        if 'sponsorship' in q_lower or 'visa' in q_lower:
            # If authorized, you don't need sponsorship.
            return 'No' if self.preferences['authorized_to_work'] == 'yes' else 'Yes'
        
        if 'legally authorized' in q_lower or 'work authorization' in q_lower or 'work permit' in q_lower or 'right to work' in q_lower:
            # This question is "Are you authorized?".
            if self.preferences['authorized_to_work'] == 'yes':
                return 'Yes'
            else:
                # Return a more specific string to help match dropdowns like "No, I require sponsorship".
                return 'No, I require sponsorship'

        # EEO Questions - Default to "Decline to self-identify" for privacy and safety.
        if 'gender' in q_lower or 'race' in q_lower or 'ethnicity' in q_lower or 'veteran' in q_lower or 'disability' in q_lower:
            return 'Decline to self-identify'

        return None

    def _calculate_years_for_skill(self, skill_name: str) -> str:
        """Calculates total years of experience for a skill from parsed resume data."""
        work_experience = self.resume_data.get('work_experience', [])
        if not work_experience:
            return "0"

        # This is a simplified date parser for demonstration
        def parse_date(date_str):
            if 'present' in date_str.lower(): return datetime.datetime.now()
            for fmt in ('%b %Y', '%B %Y', '%Y'):
                try: return datetime.datetime.strptime(date_str, fmt)
                except ValueError: pass
            return None

        relevant_periods = []
        for job in work_experience:
            context = (job.get('title', '') + ' ' + job.get('description', '')).lower()
            if re.search(r'\b' + re.escape(skill_name.lower()) + r'\b', context):
                start = parse_date(job.get('start_date', ''))
                end = parse_date(job.get('end_date', 'Present'))
                if start and end:
                    relevant_periods.append((start, end))
        
        if not relevant_periods: return "0"

        relevant_periods.sort(key=lambda x: x[0])
        merged = []
        for start, end in relevant_periods:
            if not merged or start > merged[-1][1]: merged.append([start, end])
            else: merged[-1][1] = max(merged[-1][1], end)
        
        total_days = sum((end - start).days for start, end in merged)
        return str(round(total_days / 365.25))

    def _select_answer(self, select_el: webdriver.remote.webelement.WebElement, answer: str):
        """Selects an answer in a dropdown, trying partial text match."""
        try:
            select = Select(select_el)
            for option in select.options:
                if answer.lower() in option.text.lower():
                    select.select_by_visible_text(option.text); return
        except Exception as e: logger.warning(f"Could not select '{answer}' in dropdown: {e}")

    def _check_answer(self, container: webdriver.remote.webelement.WebElement, answer: str):
        """Finds and clicks a radio button or checkbox within a container."""
        try:
            options = container.find_elements(By.CSS_SELECTOR, 'input[type="radio"], input[type="checkbox"]')
            for option in options:
                if answer.lower() in option.find_element(By.XPATH, '..').text.lower():
                    option.click(); return
        except Exception as e: logger.warning(f"Could not check '{answer}' radio/checkbox: {e}")

    def submit_application(self) -> bool:
        try:
            submit_selectors = [
                'input[type="submit"]',
                'button[type="submit"]',
                'button[name*="submit"]',
                '.submit-btn',
                '[data-automation-id="submitApplication"]',
                '//button[contains(text(), "Submit")]',
                '//button[contains(text(), "Apply")]'
            ]
            submit_btn = self.find_element_by_multiple_selectors(submit_selectors)
            if submit_btn and submit_btn.is_enabled():
                submit_btn.click()
                logger.info("Application submitted successfully!")
                return True
            else:
                logger.warning("Submit button not found or not enabled")
                return False
        except Exception as e:
            logger.error(f"Error submitting application: {str(e)}")
            return False


class GoogleCalendarManager:
    """Enhanced Google Calendar integration for job application tracking."""
    def __init__(self):
        self.calendar_service = self.authenticate_calendar()
        self.job_alerts_calendar_id = None
        self.user_name = None
        self._initialize_calendar()
    
    def authenticate_calendar(self):
        """Authenticate with Google Calendar API using shared credentials."""
        # Use the same authentication as Gmail since we share the token
        gmail_service = SmartEmailer().authenticate_gmail()
        if not gmail_service:
            return None
        
        # Get the credentials from the gmail service
        creds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        
        if not creds or not creds.valid:
            logger.error("No valid credentials for Calendar API")
            return None
        
        try:
            return build('calendar', 'v3', credentials=creds)
        except Exception as e:
            logger.error(f"Failed to build Calendar service: {e}")
            return None
    
    def _initialize_calendar(self):
        """Initialize the Job Alerts calendar and get user info."""
        if not self.calendar_service:
            return
        
        try:
            # Get user's name from their primary calendar
            primary_calendar = self.calendar_service.calendars().get(calendarId='primary').execute()
            self.user_name = primary_calendar.get('summary', 'User')
            
            # Find or create "Job Alerts" calendar
            self.job_alerts_calendar_id = self._get_or_create_job_alerts_calendar()
            
        except Exception as e:
            logger.error(f"Failed to initialize calendar: {e}")
    
    def _get_or_create_job_alerts_calendar(self) -> Optional[str]:
        """Find existing Job Alerts calendar or create a new one."""
        try:
            # List all calendars to find existing "Job Alerts" calendar
            calendar_list = self.calendar_service.calendarList().list().execute()
            
            for calendar in calendar_list.get('items', []):
                if calendar.get('summary') == 'Job Alerts':
                    logger.info(f"Found existing Job Alerts calendar: {calendar['id']}")
                    return calendar['id']
            
            # Create new "Job Alerts" calendar if not found
            calendar_body = {
                'summary': 'Job Alerts',
                'description': f'Automated job application tracking for {self.user_name}',
                'timeZone': 'UTC'
            }
            
            created_calendar = self.calendar_service.calendars().insert(body=calendar_body).execute()
            calendar_id = created_calendar['id']
            
            # Set calendar color to a nice blue
            calendar_list_entry = {
                'id': calendar_id,
                'colorId': '9'  # Blue color
            }
            self.calendar_service.calendarList().patch(
                calendarId=calendar_id, 
                body=calendar_list_entry
            ).execute()
            
            logger.info(f"Created new Job Alerts calendar: {calendar_id}")
            return calendar_id
            
        except Exception as e:
            logger.error(f"Failed to get/create Job Alerts calendar: {e}")
            return 'primary'  # Fallback to primary calendar
    
    def _get_existing_event_id(self, job_id: int, event_type: str) -> Optional[str]:
        """Check if an event already exists for this job to prevent duplicates."""
        if not self.job_alerts_calendar_id:
            return None
        
        try:
            # Search for existing events with this job ID in the description
            search_query = f"Job ID: {job_id}"
            events = self.calendar_service.events().list(
                calendarId=self.job_alerts_calendar_id,
                q=search_query,
                maxResults=50
            ).execute()
            
            for event in events.get('items', []):
                description = event.get('description', '')
                if f"Job ID: {job_id}" in description and event_type.lower() in event.get('summary', '').lower():
                    return event.get('id')
            
            return None
            
        except Exception as e:
            logger.error(f"Failed to check for existing events: {e}")
            return None
    
    def _delete_existing_event(self, event_id: str):
        """Delete an existing calendar event."""
        try:
            self.calendar_service.events().delete(
                calendarId=self.job_alerts_calendar_id,
                eventId=event_id
            ).execute()
            logger.info(f"Deleted existing calendar event: {event_id}")
        except Exception as e:
            logger.error(f"Failed to delete existing event {event_id}: {e}")
    
    def create_interview_event(self, job_data: Dict) -> Optional[str]:
        """Create a calendar event when job status changes to 'interview'."""
        if not self.calendar_service or not self.job_alerts_calendar_id:
            logger.warning("Calendar service not available. Skipping interview event creation.")
            return None
        
        try:
            job_id = job_data.get('id')
            
            # Check for existing interview event and delete it
            existing_event_id = self._get_existing_event_id(job_id, 'interview')
            if existing_event_id:
                self._delete_existing_event(existing_event_id)
            
            # Default to next business day at 10 AM if no specific time provided
            interview_date = datetime.datetime.now() + datetime.timedelta(days=1)
            # Skip to Monday if it's weekend
            while interview_date.weekday() > 4:  # 0-4 is Mon-Fri
                interview_date += datetime.timedelta(days=1)
            
            interview_date = interview_date.replace(hour=10, minute=0, second=0, microsecond=0)
            end_time = interview_date + datetime.timedelta(hours=1)
            
            event = {
                'summary': f'🎯 Interview: {job_data.get("title", "Job")} at {job_data.get("company", "Company")}',
                'description': self._build_interview_description(job_data),
                'start': {
                    'dateTime': interview_date.isoformat(),
                    'timeZone': 'UTC',
                },
                'end': {
                    'dateTime': end_time.isoformat(),
                    'timeZone': 'UTC',
                },
                'reminders': {
                    'useDefault': False,
                    'overrides': [
                        {'method': 'popup', 'minutes': 24 * 60},  # 1 day before
                        {'method': 'popup', 'minutes': 60},       # 1 hour before
                    ],
                },
                'colorId': '9',  # Blue color for interviews
            }
            
            result = self.calendar_service.events().insert(
                calendarId=self.job_alerts_calendar_id, 
                body=event
            ).execute()
            logger.info(f"Created interview calendar event in Job Alerts calendar: {result.get('id')}")
            return result.get('id')
            
        except Exception as e:
            logger.error(f"Failed to create interview calendar event: {e}")
            return None
    
    def create_deadline_reminder(self, job_data: Dict, deadline_str: str) -> Optional[str]:
        """Create a calendar reminder for application deadlines."""
        if not self.calendar_service or not self.job_alerts_calendar_id or not deadline_str or deadline_str == 'N/A':
            return None
        
        try:
            job_id = job_data.get('id')
            
            # Check for existing deadline event and delete it
            existing_event_id = self._get_existing_event_id(job_id, 'deadline')
            if existing_event_id:
                self._delete_existing_event(existing_event_id)
            
            # Parse deadline string (you might need to improve this parsing)
            deadline_date = self._parse_deadline(deadline_str)
            if not deadline_date:
                return None
            
            # Create all-day event for the deadline
            event = {
                'summary': f'⏰ Application Deadline: {job_data.get("title", "Job")} at {job_data.get("company", "Company")}',
                'description': self._build_deadline_description(job_data),
                'start': {
                    'date': deadline_date.strftime('%Y-%m-%d'),
                },
                'end': {
                    'date': deadline_date.strftime('%Y-%m-%d'),
                },
                'reminders': {
                    'useDefault': False,
                    'overrides': [
                        {'method': 'popup', 'minutes': 24 * 60 * 3},  # 3 days before
                        {'method': 'popup', 'minutes': 24 * 60},      # 1 day before
                    ],
                },
                'colorId': '11',  # Red color for deadlines
            }
            
            result = self.calendar_service.events().insert(
                calendarId=self.job_alerts_calendar_id, 
                body=event
            ).execute()
            logger.info(f"Created deadline reminder in Job Alerts calendar: {result.get('id')}")
            return result.get('id')
            
        except Exception as e:
            logger.error(f"Failed to create deadline reminder: {e}")
            return None
    
    def create_followup_reminder(self, job_data: Dict, status: str) -> Optional[str]:
        """Create follow-up reminders based on job status and company response patterns."""
        if not self.calendar_service or not self.job_alerts_calendar_id:
            return None
        
        try:
            job_id = job_data.get('id')
            
            # Check for existing follow-up event and delete it
            existing_event_id = self._get_existing_event_id(job_id, 'follow-up')
            if existing_event_id:
                self._delete_existing_event(existing_event_id)
            
            followup_date = self._calculate_followup_date(status)
            if not followup_date:
                return None
            
            followup_time = followup_date.replace(hour=9, minute=0, second=0, microsecond=0)
            end_time = followup_time + datetime.timedelta(minutes=30)
            
            event = {
                'summary': f'📞 Follow-up: {job_data.get("title", "Job")} at {job_data.get("company", "Company")}',
                'description': self._build_followup_description(job_data, status),
                'start': {
                    'dateTime': followup_time.isoformat(),
                    'timeZone': 'UTC',
                },
                'end': {
                    'dateTime': end_time.isoformat(),
                    'timeZone': 'UTC',
                },
                'reminders': {
                    'useDefault': False,
                    'overrides': [
                        {'method': 'popup', 'minutes': 60},  # 1 hour before
                    ],
                },
                'colorId': '5',  # Yellow color for follow-ups
            }
            
            result = self.calendar_service.events().insert(
                calendarId=self.job_alerts_calendar_id, 
                body=event
            ).execute()
            logger.info(f"Created follow-up reminder in Job Alerts calendar: {result.get('id')}")
            return result.get('id')
            
        except Exception as e:
            logger.error(f"Failed to create follow-up reminder: {e}")
            return None
    
    def _build_interview_description(self, job_data: Dict) -> str:
        """Build detailed description for interview calendar event."""
        skills_list = job_data.get('skills', '').split(',')[:5] if job_data.get('skills') else []
        skills_text = ', '.join([skill.strip() for skill in skills_list]) if skills_list else 'Review job requirements'
        
        description = f"""🎯 INTERVIEW PREPARATION CHECKLIST for {self.user_name}

📋 Job Details:
• Position: {job_data.get('title', 'N/A')}
• Company: {job_data.get('company', 'N/A')}
• Location: {job_data.get('location', 'N/A')}
• Experience Level: {job_data.get('experience', 'N/A')}
• Job Type: {job_data.get('job_type', 'N/A')}
• Relevance Score: {job_data.get('relevance_score', 'N/A')}%

🔍 Pre-Interview Research (Complete 24-48 hours before):
□ Company mission, values, and recent news/press releases
□ Research interviewer(s) on LinkedIn
□ Review company's products/services and competitors
□ Understand the role's responsibilities and requirements
□ Prepare 3-5 thoughtful questions about the role/company
□ Practice common behavioral and technical questions

💼 Materials to Prepare:
□ 3-5 printed copies of updated resume
□ Portfolio/work samples relevant to the role
□ List of professional references with contact info
□ Notebook and pen for taking notes
□ Business cards (if you have them)
□ Directions and parking information

⚡ Key Skills & Experience to Highlight:
{skills_text}

🗣️ STAR Method Examples to Prepare:
□ Situation: Challenging project or problem you faced
□ Task: What you needed to accomplish
□ Action: Specific steps you took to address it
□ Result: Positive outcome and what you learned

❓ Questions to Ask Them:
□ "What does success look like in this role after 6 months?"
□ "What are the biggest challenges facing the team/department?"
□ "How would you describe the company culture?"
□ "What opportunities are there for professional development?"
□ "What are the next steps in the interview process?"

📝 Your Notes:
{job_data.get('notes', 'No additional notes')}

🔗 Original Job Posting: {job_data.get('link', 'N/A')}
📧 Follow-up: Send thank-you email within 24 hours after interview

---
Generated by AI JobSnap for {self.user_name}
Job ID: {job_data.get('id', 'N/A')}
"""
        return description
    
    def _build_deadline_description(self, job_data: Dict) -> str:
        """Build description for deadline reminder."""
        return f"""⏰ APPLICATION DEADLINE TODAY for {self.user_name}!

📋 Job Details:
• Position: {job_data.get('title', 'N/A')}
• Company: {job_data.get('company', 'N/A')}
• Location: {job_data.get('location', 'N/A')}
• Experience Required: {job_data.get('experience', 'N/A')}
• Job Type: {job_data.get('job_type', 'N/A')}
• Relevance Score: {job_data.get('relevance_score', 'N/A')}%

✅ FINAL APPLICATION CHECKLIST - Complete Today:
□ Resume tailored specifically for {job_data.get('company', 'this company')}
□ Custom cover letter highlighting relevant experience and enthusiasm
□ Portfolio/work samples that demonstrate relevant skills
□ Professional references list (if requested)
□ Review all application requirements one final time
□ Proofread everything for typos and formatting
□ Submit application before 11:59 PM today!

🎯 Key Skills to Emphasize:
{', '.join(job_data.get('skills', '').split(',')[:3]) if job_data.get('skills') else 'Review job requirements'}

💡 Last-Minute Tips:
• Use keywords from the job description in your application
• Quantify your achievements with specific numbers/results
• Show genuine interest in the company and role
• Follow application instructions exactly as specified

📝 Your Notes:
{job_data.get('notes', 'No additional notes')}

🔗 Apply Here: {job_data.get('link', 'N/A')}

⚠️ URGENT: This is your final reminder - deadline is TODAY!

---
Generated by AI JobSnap for {self.user_name}
Job ID: {job_data.get('id', 'N/A')}
"""
    
    def _build_followup_description(self, job_data: Dict, status: str) -> str:
        """Build description for follow-up reminder."""
        status_descriptions = {
            'applied': {
                'action': 'Send polite follow-up email asking about application timeline and next steps',
                'template': f"Hi [Hiring Manager], I hope this message finds you well. I recently submitted my application for the {job_data.get('title', '')} position at {job_data.get('company', '')} and wanted to follow up on the status. I remain very enthusiastic about this opportunity and would appreciate any updates on your timeline for next steps. Thank you for your time and consideration.",
                'tips': '• Keep it brief and professional\n• Show continued interest\n• Ask about timeline, not decision\n• Include your contact information'
            },
            'interview': {
                'action': 'Send thank-you note within 24 hours and reiterate your strong interest',
                'template': f"Dear [Interviewer Name], Thank you for taking the time to meet with me yesterday about the {job_data.get('title', '')} position. I enjoyed our conversation about [specific topic discussed] and am even more excited about the opportunity to contribute to {job_data.get('company', '')}. Please let me know if you need any additional information. I look forward to hearing about next steps.",
                'tips': '• Send within 24 hours\n• Reference specific conversation points\n• Reiterate key qualifications\n• Keep it concise but personal'
            },
            'offered': {
                'action': 'Respond to job offer professionally - accept, negotiate, or request more time',
                'template': f"Dear [Hiring Manager], Thank you for extending the offer for the {job_data.get('title', '')} position at {job_data.get('company', '')}. I am excited about this opportunity and would like to [accept/discuss terms/request additional time to consider]. [Include specific next steps or questions]. I appreciate your time and look forward to your response.",
                'tips': '• Respond promptly (within 2-3 business days)\n• Be gracious regardless of your decision\n• If negotiating, be specific and reasonable\n• Confirm details in writing'
            },
            'rejected': {
                'action': 'Request constructive feedback and maintain professional relationship for future opportunities',
                'template': f"Dear [Hiring Manager], Thank you for informing me about your decision regarding the {job_data.get('title', '')} position. While I'm disappointed, I understand these decisions are difficult. I would greatly appreciate any feedback you could share about my application or interview to help me improve. I remain interested in {job_data.get('company', '')} and would welcome the opportunity to be considered for future roles that match my background.",
                'tips': '• Stay positive and professional\n• Ask for specific feedback\n• Express continued interest in company\n• Keep the door open for future opportunities'
            }
        }
        
        status_info = status_descriptions.get(status, {
            'action': 'Follow up on application status',
            'template': f"Hi [Name], I wanted to follow up on my application for the {job_data.get('title', '')} position. I remain very interested in this opportunity and would appreciate any updates on the timeline. Thank you for your time and consideration.",
            'tips': '• Keep it professional and brief'
        })
        
        return f"""📞 FOLLOW-UP REMINDER - {status.upper()} STATUS

📋 Job Details:
• Position: {job_data.get('title', 'N/A')}
• Company: {job_data.get('company', 'N/A')}
• Location: {job_data.get('location', 'N/A')}
• Current Status: {status.title()}
• Relevance Score: {job_data.get('relevance_score', 'N/A')}%

🎯 Action Required:
{status_info['action']}

📧 Email Template:
{status_info['template']}

💡 Best Practices:
{status_info['tips']}

📝 Your Notes:
{job_data.get('notes', 'No notes available')}

🔗 Job Link: {job_data.get('link', 'N/A')}

---
Generated by AI JobSnap for {self.user_name}
"""
    
    def _parse_deadline(self, deadline_str: str) -> Optional[datetime.datetime]:
        """Parse deadline string into datetime object."""
        if not deadline_str or deadline_str == 'N/A':
            return None
        
        try:
            # Try common date formats
            formats = [
                '%Y-%m-%d',
                '%m/%d/%Y',
                '%d/%m/%Y',
                '%B %d, %Y',
                '%b %d, %Y',
                '%d %B %Y',
                '%d %b %Y'
            ]
            
            for fmt in formats:
                try:
                    return datetime.datetime.strptime(deadline_str.strip(), fmt)
                except ValueError:
                    continue
            
            # If no format matches, try to extract date from text
            import re
            date_match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', deadline_str)
            if date_match:
                month, day, year = date_match.groups()
                return datetime.datetime(int(year), int(month), int(day))
            
            logger.warning(f"Could not parse deadline: {deadline_str}")
            return None
            
        except Exception as e:
            logger.error(f"Error parsing deadline '{deadline_str}': {e}")
            return None
    
    def _calculate_followup_date(self, status: str) -> Optional[datetime.datetime]:
        """Calculate appropriate follow-up date based on status."""
        now = datetime.datetime.now()
        
        followup_delays = {
            'applied': 14,    # 2 weeks after application
            'interview': 1,   # 1 day after interview (thank you note)
            'offered': 3,     # 3 days to respond to offer
            'rejected': 30,   # 1 month later for feedback/networking
        }
        
        days = followup_delays.get(status)
        if days is None:
            return None
        
        followup_date = now + datetime.timedelta(days=days)
        
        # Skip weekends for follow-ups
        while followup_date.weekday() > 4:  # 0-4 is Mon-Fri
            followup_date += datetime.timedelta(days=1)
        
        return followup_date


class SmartEmailer:
    """Enhanced email system with better formatting and attachments using Gmail API."""
    def __init__(self):
        self.gmail_service = self.authenticate_gmail()

    def authenticate_gmail(self):
        creds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        
        # Check if we have the required scopes
        if creds and creds.valid:
            # Verify we have both Gmail and Calendar scopes
            if not all(scope in (creds.scopes or []) for scope in SCOPES):
                logger.info("Token missing required scopes. Re-authenticating...")
                creds = None
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    # FIX: Use google.auth.transport.requests.Request() for token refresh
                    creds.refresh(Request())
                except Exception as e:
                    logger.warning(f"Failed to refresh token, re-authenticating: {e}")
                    creds = None
            
            if not creds:
                try:
                    flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                    creds = flow.run_local_server(port=0)
                except Exception as e:
                    logger.critical(f"Failed to authenticate with Google APIs: {e}")
                    return None
        
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)
        
        return build('gmail', 'v1', credentials=creds)

    def create_excel_report(self, jobs: List[Job]) -> str:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
        from openpyxl.utils import get_column_letter
        filename = f"jobs_report_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        job_data = [job.to_dict() for job in jobs]
        # Remove 'id' and 'source' columns if present and add row number
        df = pd.DataFrame(job_data)
        if 'id' in df.columns:
            df = df.drop(columns=['id'])
        if 'source' in df.columns:
            df = df.drop(columns=['source'])
        df.insert(0, 'No.', range(1, len(df) + 1))
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Jobs', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Jobs']
                # --- Add Title Row ---
                title = f"Job Report – {datetime.date.today().strftime('%B %d, %Y')} (Total: {len(jobs)})"
                worksheet.insert_rows(1)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)
                title_cell = worksheet.cell(row=1, column=1)
                title_cell.value = title
                title_cell.font = Font(size=16, bold=True, color="FFFFFF")
                title_cell.fill = GradientFill(stop=("4F81BD", "1E90FF"))
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                # --- Header Styling ---
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = GradientFill(stop=("4F81BD", "1E90FF"))
                for cell in worksheet[2]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                worksheet.freeze_panes = worksheet['A3']
                # --- Column Widths and Alternating Row Colors ---
                for i, column in enumerate(worksheet.columns, 1):
                    max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column)
                    worksheet.column_dimensions[get_column_letter(i)].width = min(max_length + 4, 50)
                alt_fill = PatternFill("solid", fgColor="F2F2F2")
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                    if (row[0].row % 2) == 1:
                        for cell in row:
                            cell.fill = alt_fill
                # --- Borders for all cells ---
                thin = Side(border_style="thin", color="CCCCCC")
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # --- Highlight High/Medium/Low Relevance ---
                rel_col = None
                for idx, cell in enumerate(worksheet[2], 1):
                    if cell.value and 'relevance' in str(cell.value).lower():
                        rel_col = idx
                        break
                if rel_col:
                    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                        try:
                            val = float(row[rel_col-1].value)
                            if val >= 60:
                                for cell in row:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                                row[rel_col-1].value = f"🟢 {val}"
                            elif val >= 30:
                                row[rel_col-1].fill = PatternFill("solid", fgColor="FFF2CC")
                                row[rel_col-1].value = f"🟡 {val}"
                            else:
                                row[rel_col-1].fill = PatternFill("solid", fgColor="F8CBAD")
                                row[rel_col-1].value = f"🔴 {val}"
                        except:
                            continue
                # --- Make Links Clickable and Button Style ---
                link_col = None
                for idx, cell in enumerate(worksheet[2], 1):
                    if cell.value and 'link' in str(cell.value).lower():
                        link_col = idx
                        break
                if link_col:
                    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                        cell = row[link_col-1]
                        if cell.value and cell.value.startswith('http'):
                            cell.hyperlink = cell.value
                            cell.value = 'Apply Now →'
                            cell.font = Font(bold=True, color="1565C0")
                            cell.fill = PatternFill("solid", fgColor="E3F2FD")
                            cell.alignment = Alignment(horizontal="center")
                # --- Highlight Salary Column ---
                salary_col = None
                for idx, cell in enumerate(worksheet[2], 1):
                    if cell.value and 'salary' in str(cell.value).lower():
                        salary_col = idx
                        break
                if salary_col:
                    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                        cell = row[salary_col-1]
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
                # --- Improve Header Names ---
                header_map = {
                    'title': 'Job Title',
                    'company': 'Company',
                    'location': 'Location',
                    'salary': 'Salary / Stipend',
                    'link': 'Job Link',
                    'description': 'Description',
                    'keywords': 'Keywords',
                    'skills': 'Skills',
                    'experience': 'Experience',
                    'job_type': 'Job Type',
                    'posted_date': 'Posted Date',
                    'relevance_score': 'Relevance Score',
                    'min_experience_years': 'Min Exp (yrs)',
                    'max_experience_years': 'Max Exp (yrs)',
                    'extracted_tools': 'Tools',
                    'extracted_soft_skills': 'Soft Skills',
                    'user_feedback': 'User Feedback',
                }
                for idx, cell in enumerate(worksheet[2], 1):
                    if cell.value in header_map:
                        cell.value = header_map[cell.value]
                # --- Add Footer Row ---
                footer_row = worksheet.max_row + 1
                worksheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=worksheet.max_column)
                footer_cell = worksheet.cell(row=footer_row, column=1)
                footer_cell.value = f"End of Report – {datetime.date.today().strftime('%B %d, %Y')}"
                footer_cell.font = Font(italic=True, color="888888")
                footer_cell.alignment = Alignment(horizontal="center")
            logger.info(f"Excel report created: {filename}.")
            return filename
        except Exception as e:
            logger.error(f"Error creating Excel report: {e}")
            return ""

    def build_smart_html_email(self, jobs: List[Job]) -> str:
        if not jobs:
            logger.info("No jobs to include in the email report. Sending 'no jobs' email.")
            return """
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <style>
                    body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 0; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }
                    .container { max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 10px 30px rgba(0,0,0,0.1); }
                    .header { background: linear-gradient(45deg, #FF6B6B, #4ECDC4); color: white; padding: 30px; text-align: center; }
                    .content { padding: 30px; text-align: center; }
                    .emoji { font-size: 48px; margin-bottom: 20px; }
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>🔍 Daily Job Hunt Report</h1>
                        <p>No matches found today</p>
                    </div>
                    <div class="content">
                        <div class="emoji">😴</div>
                        <h2>No new jobs found matching your criteria today</h2>
                        <p>Don't worry! The JobSnap is still running and will continue searching for opportunities that match your profile.</p>
                    </div>
                </div>
            </body>
            </html>
            """
        
        total_jobs = len(jobs)
        avg_relevance = sum(job.relevance_score for job in jobs) / len(jobs) if jobs else 0
        high_relevance_jobs = [job for job in jobs if job.relevance_score >= 60]
        
        company_counts = Counter(job.company for job in jobs)
        salary_jobs = [job for job in jobs if job.salary != 'N/A']
        remote_jobs = [job for job in jobs if 'remote' in job.location.lower() or 'remote' in job.job_type.lower()]
        source_counts = Counter(job.source for job in jobs)
        skill_counts = Counter(s for job in jobs for s in job.skills)
        experience_levels = Counter()
        for job in jobs:
            exp_key = job.experience.lower()
            if 'fresher' in exp_key or 'entry' in exp_key:
                experience_levels['Entry Level'] += 1
            elif 'senior' in exp_key or 'lead' in exp_key or 'manager' in exp_key:
                experience_levels['Senior Level'] += 1
            else:
                experience_levels['Mid Level'] += 1

        top_companies = sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:6]
        top_skills = sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)[:8]
        top_sources = sorted(source_counts.items(), key=lambda x: x[1], reverse=True)
        top_jobs = sorted(jobs, key=lambda x: x.relevance_score, reverse=True)[:10]

        html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Daily Job Intelligence Report</title>
            <style>
                * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                body {{ 
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; 
                    line-height: 1.6; 
                    color: #2c3e50; 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    padding: 20px;
                }}
                .email-container {{ 
                    max-width: 800px; 
                    margin: 0 auto; 
                    background: #ffffff; 
                    border-radius: 16px; 
                    overflow: hidden; 
                    box-shadow: 0 20px 40px rgba(0,0,0,0.1);
                }}
                .header {{ 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white; 
                    text-align: center; 
                    padding: 40px 20px; 
                    position: relative;
                    overflow: hidden;
                }}
                .header::before {{
                    content: '';
                    position: absolute;
                    top: -50%;
                    left: -50%;
                    width: 200%;
                    height: 200%;
                    background: url('data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100"><circle cx="50" cy="50" r="2" fill="rgba(255,255,255,0.1)"/></svg>') repeat;
                    animation: float 20s infinite linear;
                }}
                @keyframes float {{ from {{ transform: translateX(-50px) translateY(-50px); }} to {{ transform: translateX(50px) translateY(50px); }} }}
                .header h1 {{ font-size: 2.5rem; margin-bottom: 10px; position: relative; z-index: 1; }}
                .header p {{ font-size: 1.1rem; opacity: 0.9; position: relative; z-index: 1; }}
                
                .summary-stats {{ 
                    display: grid; 
                    grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); 
                    gap: 20px; 
                    padding: 30px; 
                    background: #f8f9fa;
                }}
                .stat-card {{ 
                    background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
                    border-radius: 12px; 
                    padding: 20px; 
                    text-align: center; 
                    border: 1px solid rgba(0,0,0,0.05);
                    transition: transform 0.2s ease, box-shadow 0.2s ease;
                    position: relative;
                    overflow: hidden;
                }}
                .stat-card::before {{
                    content: '';
                    position: absolute;
                    top: 0;
                    left: 0;
                    right: 0;
                    height: 3px;
                    background: linear-gradient(90deg, #667eea, #764ba2, #f093fb);
                }}
                .stat-card:hover {{ transform: translateY(-2px); box-shadow: 0 10px 25px rgba(0,0,0,0.1); }}
                .stat-number {{ 
                    font-size: 2rem; 
                    font-weight: 700; 
                    background: linear-gradient(135deg, #667eea, #764ba2);
                    -webkit-background-clip: text;
                    -webkit-text-fill-color: transparent;
                    background-clip: text;
                    margin-bottom: 5px;
                }}
                .stat-label {{ font-size: 0.9rem; color: #6c757d; font-weight: 500; }}
                
                .content-section {{ padding: 30px; }}
                .insights {{ margin-bottom: 30px; }}
                .insights h2 {{ 
                    font-size: 1.8rem; 
                    margin-bottom: 20px; 
                    color: #2c3e50;
                    display: flex;
                    align-items: center;
                    gap: 10px;
                }}
                .insights h3 {{ 
                    font-size: 1.2rem; 
                    margin: 20px 0 10px 0; 
                    color: #495057;
                    display: flex;
                    align-items: center;
                    gap: 8px;
                }}
                
                .tag-container {{ 
                    display: flex; 
                    flex-wrap: wrap; 
                    gap: 8px; 
                    margin-bottom: 20px; 
                }}
                .tag {{ 
                    background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%);
                    color: #1565c0;
                    padding: 6px 14px; 
                    border-radius: 20px; 
                    font-size: 0.85rem; 
                    font-weight: 500; 
                    border: 1px solid rgba(21, 101, 192, 0.2);
                    transition: all 0.2s ease;
                }}
                .tag:hover {{ 
                    background: linear-gradient(135deg, #bbdefb 0%, #90caf9 100%);
                    transform: translateY(-1px);
                }}
                
                .jobs-section h2 {{ 
                    font-size: 1.8rem; 
                    margin-bottom: 25px; 
                    color: #2c3e50;
                    display: flex;
                    align-items: center;
                    gap: 10px;
                }}
                .job-card {{ 
                    background: #ffffff;
                    border: 1px solid #e9ecef;
                    border-radius: 12px; 
                    padding: 25px; 
                    margin-bottom: 20px; 
                    transition: all 0.3s ease;
                    position: relative;
                    overflow: hidden;
                }}
                .job-card::before {{
                    content: '';
                    position: absolute;
                    top: 0;
                    left: 0;
                    right: 0;
                    height: 4px;
                    background: linear-gradient(90deg, #667eea, #764ba2);
                    transform: scaleX(0);
                    transition: transform 0.3s ease;
                }}
                .job-card:hover {{ 
                    box-shadow: 0 10px 30px rgba(102, 126, 234, 0.15);
                    transform: translateY(-2px);
                    border-color: #667eea;
                }}
                .job-card:hover::before {{ transform: scaleX(1); }}
                
                .job-title {{ 
                    font-size: 1.3rem; 
                    font-weight: 600; 
                    margin-bottom: 8px;
                }}
                .job-title a {{ 
                    color: #667eea; 
                    text-decoration: none; 
                    transition: color 0.2s ease;
                }}
                .job-title a:hover {{ color: #764ba2; }}
                
                .job-company {{ 
                    font-weight: 600; 
                    color: #495057; 
                    font-size: 1.1rem; 
                    margin-bottom: 12px;
                }}
                .job-meta {{ 
                    display: flex; 
                    flex-wrap: wrap; 
                    gap: 15px; 
                    margin: 12px 0; 
                    font-size: 0.9rem; 
                    color: #6c757d; 
                }}
                .job-meta-item {{ 
                    display: flex; 
                    align-items: center; 
                    gap: 5px;
                    background: #f8f9fa;
                    padding: 4px 10px;
                    border-radius: 15px;
                    font-weight: 500;
                }}
                .job-description {{ 
                    font-size: 0.95rem; 
                    margin: 15px 0; 
                    color: #495057; 
                    line-height: 1.6;
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    border-left: 4px solid #667eea;
                }}
                
                .btn {{ 
                    display: inline-block; 
                    padding: 12px 24px; 
                background: linear-gradient(135deg, #134e5e 0%, #71b280 50%, #a8e6cf 100%);
                    color: white; 
                    text-decoration: none; 
                    border-radius: 25px; 
                    font-size: 0.9rem; 
                    font-weight: 600;
                    margin-top: 15px; 
                    transition: all 0.3s ease;
                    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
                }}
                .btn:hover {{ 
                    transform: translateY(-2px);
                    box-shadow: 0 8px 25px rgba(102, 126, 234, 0.4);
                    text-decoration: none;
                    color: white;
                }}
                
                .footer {{ 
                    background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
                    color: white;
                    text-align: center; 
                    padding: 30px 20px; 
                    margin-top: 20px; 
                }}
                .footer h3 {{ margin-bottom: 15px; font-size: 1.3rem; }}
                .footer p {{ margin: 8px 0; opacity: 0.9; }}
                
                .relevance-badge {{
                    position: absolute;
                    top: 15px;
                    right: 15px;
                    background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
                    color: white;
                    padding: 5px 12px;
                    border-radius: 15px;
                    font-size: 0.8rem;
                    font-weight: 600;
                }}
                
                @media (max-width: 600px) {{
                    .email-container {{ margin: 10px; border-radius: 12px; }}
                    .header h1 {{ font-size: 2rem; }}
                    .summary-stats {{ grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 15px; padding: 20px; }}
                    .content-section {{ padding: 20px; }}
                    .job-card {{ padding: 20px; }}
                    .job-meta {{ flex-direction: column; gap: 8px; }}
                    .job-meta-item {{ align-self: flex-start; }}
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                <div class="header">
                    <h1>🧠 Daily Job Intelligence Report</h1>
                    <p>{datetime.date.today().strftime('%B %d, %Y')}</p>
                </div>
                
                <div class="summary-stats">
                    <div class="stat-card">
                        <div class="stat-number">{total_jobs}</div>
                        <div class="stat-label">New Jobs</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{len(high_relevance_jobs)}</div>
                        <div class="stat-label">High Relevance</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{avg_relevance:.0f}%</div>
                        <div class="stat-label">Avg Relevance</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{len(salary_jobs)}</div>
                        <div class="stat-label">With Salary</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{len(remote_jobs)}</div>
                        <div class="stat-label">Remote</div>
                    </div>
                </div>
                
                <div class="content-section">
                    <div class="insights">
                        <h2>📊 Market Insights</h2>
                        
                        <h3>🏢 Top Hiring Companies</h3>
                        <div class="tag-container">
                            {' '.join(f'<span class="tag">{c} ({n})</span>' for c, n in top_companies)}
                        </div>
                        
                        <h3>🛠️ In-Demand Skills</h3>
                        <div class="tag-container">
                            {' '.join(f'<span class="tag">{s} ({n})</span>' for s, n in top_skills)}
                        </div>
                        
                        <h3>📱 Job Sources</h3>
                        <div class="tag-container">
                            {' '.join(f'<span class="tag">{src} ({cnt})</span>' for src, cnt in top_sources)}
                        </div>
                        
                        <h3>👨‍💼 Experience Levels</h3>
                        <div class="tag-container">
                            {' '.join(f'<span class="tag">{lvl} ({cnt})</span>' for lvl, cnt in experience_levels.items())}
                        </div>
                    </div>
                    
                    <div class="jobs-section">
                        <h2>🎯 Top Job Matches</h2>
                        {''.join(f'''
<div class="job-card">
    <div class="relevance-badge">{job.relevance_score:.0f}% Match</div>
    <div class="job-title">
        {f'<a href="{job.link}" target="_blank">{job.title}</a>' if job.title and job.title != 'N/A' and job.link and job.link != 'N/A' else (job.title if job.title and job.title != 'N/A' else '')}
    </div>
    <div class="job-company">{job.company}</div>
    <div class="job-meta">
        {f'<span class="job-meta-item">📍 {job.location}</span>' if job.location and job.location != 'N/A' else ''}
        {f'<span class="job-meta-item">💰 {job.salary}</span>' if job.salary and job.salary != 'N/A' else ''}
        {f'<span class="job-meta-item">👨‍💼 {job.experience}</span>' if job.experience and job.experience != 'N/A' else ''}
        {f'<span class="job-meta-item">🔗 {job.source}</span>' if job.source and job.source != 'N/A' else ''}
        {f'<span class="job-meta-item">📅 {job.posted_date}</span>' if job.posted_date and job.posted_date != 'N/A' else ''}
    </div>
    <div class="job-description">
        {job.description[:300]}{"..." if len(job.description) > 300 else ""}
    </div>
    {f'''<div class="tag-container">
        {' '.join(f'<span class="tag">{s}</span>' for s in (job.skills + job.extracted_tools)[:6])}
    </div>''' if job.skills or job.extracted_tools else ''}
    <a href="{job.link}" class="btn" target="_blank">Apply Now →</a>
</div>
''' for job in top_jobs)}
                    </div>
                </div>
                
                <div class="footer">
                    <h3>🤖 AI-Powered Job Intelligence</h3>
                    <p>Scanned <strong>{total_jobs}</strong> jobs from <strong>{len(top_sources)}</strong> sources</p>
                    <p>Next scan: {(datetime.datetime.now() + datetime.timedelta(hours=12)).strftime('%I:%M %p')}</p>
                    <p style="margin-top: 20px; font-size: 0.9rem; opacity: 0.8;">
                        Built with ❤️ by your Super Job Agent
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        return html

    def send_email(self, subject: str, html_content: str, attachment_path: Optional[str] = None, recipients: List[str] = None):
        """Sends an email with HTML content and an optional attachment using Gmail API."""
        if recipients is None:
            recipients = [ADMIN_ALERT_EMAIL] # Default to admin if no recipients specified

        if not self.gmail_service:
            logger.critical("Email not sent: Gmail service not authenticated.")
            return

        message = MIMEMultipart()
        message['to'] = ', '.join(recipients)
        message['subject'] = subject
        message.attach(MIMEText(html_content, 'html'))
        
        if attachment_path and os.path.exists(attachment_path):
            try:
                with open(attachment_path, 'rb') as f:
                    mime_base = MIMEBase('application', 'octet-stream')
                    mime_base.set_payload(f.read())
                    encoders.encode_base64(mime_base)
                    mime_base.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(attachment_path)}"')
                    message.attach(mime_base)
                logger.info(f"Attached file: {os.path.basename(attachment_path)} to email.")
            except Exception as e:
                logger.error(f"Error attaching file {attachment_path}: {e}")
        
        raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
        try:
            self.gmail_service.users().messages().send(userId='me', body={'raw': raw}).execute()
            logger.info(f"📧 Email sent successfully to {recipients}!")
        except Exception as e:
            logger.error(f"Error sending email to {recipients}: {e}")

class JobHunter:
    """Main class to run the job alert system."""
    def __init__(self, user_id: int, search_terms: List[str], location: str, experience: str = "", job_type: str = "", task=None):
        self.user_id = user_id
        self.search_terms = search_terms
        self.location = location
        self.experience = experience
        self.job_type = job_type
        self.task = task
        self.db = JobDatabase() # This is fine, DB can be initialized here.
        self.scraper = EnhancedJobScraper(search_terms=self.search_terms, location=self.location, experience=self.experience, job_type=self.job_type, user_id=self.user_id, task=self.task)
        self.emailer = SmartEmailer()
        
    def run(self):
        """Orchestrates the entire job search and alert process."""
        logger.info(f"🔍 Starting job hunt for user {self.user_id} with terms: {self.search_terms}, location: {self.location}")
        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': 'Initializing job hunt...', 'progress': 5, 'milestone': 'Initializing'})

        logger.info("Calling scraper.scrape_all_sources()...")
        jobs = self.scraper.scrape_all_sources() # This method now handles progress from 15% to 88%
        logger.info(f"Scraper returned {len(jobs)} jobs.")
        
        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': f'Scraping complete. Found {len(jobs)} potential jobs. Saving to database...', 'progress': 90})
        
        # Store new jobs in the database for the specific user
        new_jobs_count = 0
        for job in jobs:
            if self.db.add_job(job, self.user_id):
                new_jobs_count += 1
        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': f'Saved {new_jobs_count} new jobs to your dashboard.', 'progress': 92, 'milestone': 'Saving'})
        logger.info(f"💾 Added {new_jobs_count} new jobs to the database for user {self.user_id}.")

        # Retrieve user settings for email
        user_settings = self.db.get_user_settings(self.user_id)
        email_recipients = user_settings.get('email_recipients') if user_settings else ADMIN_ALERT_EMAIL
        send_excel_attachment = user_settings.get('send_excel_attachment') if user_settings else True

        if not jobs:
            logger.info(f"📊 Found 0 new jobs matching criteria for user {self.user_id} today. No email will be sent.")
            if self.task:
                self.task.update_state(state='SUCCESS', meta={'status': 'Search complete. No new jobs found.', 'progress': 100})
            # Optionally send a "no jobs found" email
            # html = self.emailer.build_smart_html_email([])
            # self.emailer.send_email(f"No New Job Matches - {datetime.date.today().strftime('%b %d, %Y')}", html)
            return
            
        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': 'Generating your personalized email report...', 'progress': 95, 'milestone': 'Reporting'})

        html = self.emailer.build_smart_html_email(jobs)
        excel_file = ""
        if send_excel_attachment:
            excel_file = self.emailer.create_excel_report(jobs)
        
        subject = f"🧠 {len(jobs)} New Job Matches – {datetime.date.today().strftime('%b %d, %Y')}"
        
        if self.task:
            self.task.update_state(state='PROGRESS', meta={'status': f'Sending report to {email_recipients}...', 'progress': 98})

        # Pass the recipients list directly
        self.emailer.send_email(subject, html, attachment_path=excel_file, recipients=email_recipients.split(','))
        
        
        if os.path.exists(excel_file):
            os.remove(excel_file)
            logger.info(f"Cleaned up Excel report: {excel_file}.")

        logger.info(f"🤖 Super Job Agent finished its run for user {self.user_id}.")

# --- Celery Tasks ---
@celery.task(bind=True, name='app.run_job_hunt_task')
def run_job_hunt_task(self, user_id: int, search_terms: List[str], location: str, experience: str, job_type: str):
    """Celery task to run the job hunt for a specific user and search profile."""
    logger.info(f"Celery task {self.request.id} started for user {user_id} with terms {search_terms} at {location}.")
    try:
        logger.info(f"Creating JobHunter instance for user {user_id}.")
        hunter = JobHunter(user_id=user_id, search_terms=search_terms, location=location, experience=experience, job_type=job_type, task=self) # Correctly passing self
        logger.info(f"Running JobHunter.run() for user {user_id}.")
        hunter.run()
        logger.info(f"Celery task {self.request.id} completed successfully for user {user_id}.")
        return {'status': 'Complete! Your dashboard will now reload.', 'progress': 100}
    except Exception as e:
        logger.error(f"Celery task {self.request.id} failed for user {user_id}: {e}", exc_info=True)
        self.update_state(state='FAILURE', meta={'status': 'Task failed!', 'error': str(e), 'progress': 100})
        # Consider sending an admin alert email here
        # SmartEmailer().send_email("CRITICAL: Job Hunt Task Failed", f"Task for user {user_id} failed with error: {traceback.format_exc()}", recipients=[ADMIN_ALERT_EMAIL])
        raise

# In app.py
@celery.task(name='app.assisted_apply_task')
def assisted_apply_task(user_id: int, job_id: int):
    """
    Launches a NON-headless browser and uses ApplicationFiller to pre-fill a job application.
    Uses a pre-existing browser profile if configured by the user, to maintain login sessions.
    """
    db = JobDatabase()
    conn = sqlite3.connect(db.db_path)
    conn.row_factory = sqlite3.Row # Use row_factory for dict-like access
    cursor = conn.cursor()

    # 1. Fetch user_details and job_link from DB
    cursor.execute("SELECT * FROM user_details WHERE user_id = ?", (user_id,))
    user_details_raw = cursor.fetchone()

    if not user_details_raw:
        logger.error(f"[AssistApply] No user_details found for user_id {user_id}. Aborting. Please add your details to the database.")
        conn.close()
        return {'status': 'error', 'message': 'User details not found in database.'}

    user_details = dict(user_details_raw)

    # Load parsed resume data from the JSON string in the database
    try:
        user_details['parsed_resume'] = json.loads(user_details.get('resume_parsed_data', '{}') or '{}')
    except (json.JSONDecodeError, TypeError):
        user_details['parsed_resume'] = {}
        logger.warning(f"[AssistApply] Could not load/parse resume data for user {user_id}. Answering will be limited.")

    # Remap DB keys to what the ApplicationFiller bot expects
    if user_details.get('linkedin_url'):
        user_details['linkedin'] = user_details['linkedin_url']
    if user_details.get('resume_path'):
        # Ensure the path is absolute for the Celery worker
                # This is a temporary fix for local development, consider a more robust path handling for production
        user_details['resume'] = os.path.abspath(user_details['resume_path'])
    if user_details.get('cover_letter_template'):
        user_details['cover_letter'] = user_details['cover_letter_template']

    # Add first/last name for compatibility with the filler bot
    full_name = user_details.get('full_name', '')
    if full_name:
        name_parts = full_name.split(' ')
        user_details['first_name'] = name_parts[0]
        user_details['last_name'] = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''

    cursor.execute("SELECT link FROM jobs WHERE id = ? AND user_id = ?", (job_id, user_id))
    job_link_raw = cursor.fetchone()
    conn.close()

    if not job_link_raw or not job_link_raw['link']:
        logger.error(f"[AssistApply] No job link found for job_id {job_id}. Aborting.")
        return {'status': 'error', 'message': 'Job link not found.'}

    job_link = job_link_raw['link']
    logger.info(f"[AssistApply] Starting for user {user_id} on job {job_id} -> {job_link}")

    # 2. Launch a NON-headless Selenium browser
    driver = None
    try:
        # Use undetected_chromedriver to avoid bot detection
        options = uc.ChromeOptions()
        options.add_argument('--start-maximized')

        # Explicitly set the browser binary location to prevent errors if auto-detection fails.
        # This is the fix for the "Binary Location Must be a String" TypeError.
        if os.path.exists(CHROME_BINARY_PATH_WINDOWS):
            options.binary_location = CHROME_BINARY_PATH_WINDOWS
        else:
            logger.warning(f"Chrome binary not found at the configured path: {CHROME_BINARY_PATH_WINDOWS}. Relying on auto-detection.")
        
        assisted_apply_mode = user_details.get('assisted_apply_mode', 'personal')
        if assisted_apply_mode == 'personal':
            browser_profile_path_input = user_details.get('browser_profile_path', '').strip()
            browser_profile_name = user_details.get('browser_profile_name', 'Default').strip() # Get the user-configured profile name
            if browser_profile_path_input:
                if browser_profile_path_input.lower().endswith('.exe'):
                    logger.error(f"[AssistApply] CONFIGURATION ERROR: The Browser Profile Path is set to an executable file ('{browser_profile_path_input}'). It should be a directory path ending in 'User Data'. Please correct this in your profile. Falling back to a temporary profile.")
                    browser_profile_path_input = '' # Clear the invalid path to proceed with fallback

            if browser_profile_path_input:
                user_data_dir = browser_profile_path_input

                if os.path.isdir(user_data_dir):
                    lock_file_path = os.path.join(user_data_dir, 'SingletonLock')
                    if os.path.exists(lock_file_path):
                        logger.warning(f"[AssistApply] WARNING: A lock file was found at '{lock_file_path}'. This strongly indicates that Chrome is already running with this profile. Please close all Chrome windows and processes before trying again.")

                    logger.info(f"[AssistApply] Using browser session from User Data Dir: '{user_data_dir}' with Profile: '{browser_profile_name}'")
                    options.user_data_dir = user_data_dir
                    options.profile_directory = browser_profile_name
                else:
                    logger.warning(f"[AssistApply] The derived User Data Directory '{user_data_dir}' (from your input '{browser_profile_path_input}') is not a valid directory. Falling back to a new temporary profile. Please check your profile settings.")
            else:
                logger.info("[AssistApply] No browser profile path configured in your profile. Starting with a new temporary profile. You may need to log in to job sites.")
        else:
            logger.info("[AssistApply] Mode is 'temporary'. Starting with a fresh, temporary browser profile.")

        logger.info("[AssistApply] Attempting to initialize undetected_chromedriver...")
        driver = uc.Chrome(options=options, headless=False, use_subprocess=True)
        
        filler = ApplicationFiller(driver, user_details)
        filler.fill(job_link)
        logger.info(f"[AssistApply] Form filling process initiated for {job_link}. The browser window is now open for your review and final submission.")
    except FileNotFoundError as e:
        logger.critical(f"[AssistApply] CONFIGURATION ERROR: {e}", exc_info=True)
        if driver: driver.quit()
    except SeleniumWebDriverException as e:
        if "session not created" in str(e).lower():
            logger.error(f"[AssistApply] CRITICAL ERROR: Could not create browser session. This often happens if the browser is already running with the same user profile. Please close all Chrome windows and try again.", exc_info=False)
            logger.debug(f"Full SessionNotCreatedException: {e}") # Log the full trace for debugging if needed
            raise  # Re-raise the exception to mark the Celery task as failed
        else:
            logger.error(f"[AssistApply] A browser automation error occurred. This might be due to a driver/browser version mismatch or a browser crash.", exc_info=True)
            raise  # Re-raise for other Selenium errors too
        if driver: driver.quit()
    except Exception as e:
        logger.error(f"An error occurred during assisted_apply_task for job {job_id}: {e}", exc_info=True)
        if driver: driver.quit() # Close browser on error
        raise # Re-raise to fail the task


@celery.task(name='app.scheduled_job_hunt_for_all_users')
def scheduled_job_hunt_for_all_users():
    db = JobDatabase()
    conn = sqlite3.connect(db.db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT id, username FROM users")
    users = cursor.fetchall()
    conn.close()

    for user_id, username in users:
        profiles = db.get_search_profiles(user_id)
        if not profiles:
            logger.info(f"User {username} (ID: {user_id}) has no saved search profiles. Skipping scheduled hunt.")
            continue

        for profile in profiles:
            search_terms = [term.strip() for term in profile['search_terms'].split(',') if term.strip()]
            location = profile['location']
            experience = profile['experience']
            job_type = profile['job_type']
            
            logger.info(f"Scheduling job hunt for user {username} (ID: {user_id}) with profile '{profile['profile_name']}'.")
            # Directly call the task, as this is now a Celery Beat scheduled task
            run_job_hunt_task.delay(user_id, search_terms, location, experience, job_type)


# --- Flask Routes ---
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))  # Logged-in users go to dashboard
    return redirect("https://aijobsnap.vercel.app")

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        password_hash = generate_password_hash(password)
        db = JobDatabase()
        user_id = db.add_user(username, password_hash)
        if user_id:
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Username already exists. Please choose a different one.', 'danger')
    return render_template('register.html', now=datetime.datetime.now())

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db_user = User.get_by_username(username)
        if db_user and check_password_hash(db_user.password_hash, password):
            login_user(db_user)
            flash('Logged in successfully!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Login Unsuccessful. Please check username and password', 'danger')
    return render_template('login.html', now=datetime.datetime.now())

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    db = JobDatabase()
    if request.method == 'POST':
        # Get existing details to not lose the resume path if not updated
        existing_details = db.get_user_details(current_user.id) or {}
        
        # Start with the existing resume path and parsed data
        resume_path = existing_details.get('resume_path')
        resume_parsed_data = existing_details.get('resume_parsed_data')

        details = {
            'full_name': request.form.get('full_name'),
            'email': request.form.get('email'),
            'phone': request.form.get('phone'),
            'address': request.form.get('address'),
            'linkedin_url': request.form.get('linkedin_url'),
            'github_url': request.form.get('github_url'),
            'portfolio_url': request.form.get('portfolio_url'),
            'cover_letter_template': request.form.get('cover_letter_template'),
            'browser_profile_path': request.form.get('browser_profile_path', '').strip(),
            'browser_profile_name': request.form.get('browser_profile_name', 'Default').strip(), # NEW: Get profile name from form
            'willing_to_relocate': request.form.get('willing_to_relocate', 'no'),
            'authorized_to_work': request.form.get('authorized_to_work', 'no'),
            'assisted_apply_mode': request.form.get('assisted_apply_mode', 'personal')
        }

        # Handle resume file upload
        if 'resume' in request.files:
            file = request.files['resume']
            if file and file.filename != '':
                # Save to a user-specific subfolder to avoid name collisions
                user_upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(current_user.id))
                os.makedirs(user_upload_dir, exist_ok=True)
                filename = secure_filename(file.filename)
                file_path = os.path.join(user_upload_dir, filename)
                file.save(file_path)
                resume_path = file_path # Update with the new path

                # Parse the resume using the new service
                parser = ResumeParsingService()
                parsed_data = parser.parse(file_path)
                resume_parsed_data = json.dumps(parsed_data) if parsed_data else None
        
        details['resume_path'] = resume_path
        details['resume_parsed_data'] = resume_parsed_data

        if db.save_user_details(current_user.id, details):
            flash('Profile updated successfully!', 'success')
        else:
            flash('Failed to update profile.', 'danger')
        return redirect(url_for('profile'))

    user_details = db.get_user_details(current_user.id)
    custom_answers = {}
    if user_details and user_details.get('custom_answers'):
        try:
            custom_answers = json.loads(user_details['custom_answers'])
        except (json.JSONDecodeError, TypeError):
            logger.warning(f"Could not parse custom_answers for user {current_user.id}")

    # Check if the profile path is set and valid
    profile_path_is_set = user_details and user_details.get('browser_profile_path')
    return render_template('profile.html', 
                           user_details=user_details, 
                           custom_answers=custom_answers,
                           profile_path_is_set=profile_path_is_set)

@app.route('/outreach_studio')
@login_required
def outreach_studio():
    """Renders the Outreach Studio dashboard."""
    outreach_db = OutreachDatabase()
    # In a real implementation, you'd get this from Redis
    # For now, we'll just use the current time as a placeholder.
    last_scan_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    stats = outreach_db.get_outreach_stats(current_user.id)
    return render_template('outreach_studio.html', last_scan_time=last_scan_time, stats=stats)

@app.route('/learning_path')
@login_required
def learning_path():
    """Renders the Learning Path page."""
    return render_template('learning_path.html')

@app.route('/setup_assisted_apply')
@login_required
def setup_assisted_apply():
    """Renders the Assisted Apply setup wizard page."""
    return render_template('setup_assisted_apply.html')

@app.route('/setup_assisted_apply/start_browser', methods=['POST'])
@login_required
def setup_assisted_apply_start_browser():
    """
    Launches a non-headless Chrome browser for the user to log into LinkedIn.
    This browser will remain open (detached).
    """
    # This function is now deprecated in favor of the re-authentication flow.
    # We will re-use the Google Auth flow which is more robust.
    try:
        temp_profile_dir = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], 'chrome_setup_profile'))
        if not os.path.exists(temp_profile_dir):
            os.makedirs(temp_profile_dir)
        logger.info(f"[Setup Wizard] Using temporary Chrome profile directory: {temp_profile_dir}")

        options = uc.ChromeOptions()
        options.user_data_dir = temp_profile_dir
        options.profile_directory = "Default"

        driver = uc.Chrome(options=options, headless=False, use_subprocess=True)
        driver.get("https://www.linkedin.com/login") # Explicitly navigate to the URL

        logger.info("[Setup Wizard] Launched Chrome browser for user login.")
        return jsonify({'status': 'success', 'message': 'Browser launched. Please log in to LinkedIn in the new window.'})

    except Exception as e:
        logger.error(f"Error launching browser for Assisted Apply setup: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'Failed to launch browser: {e}. Check server logs.'}), 500

@app.route('/setup_assisted_apply/detect', methods=['POST'])
@login_required
def setup_assisted_apply_detect():
    """
    Detects, tests, and saves the user's Chrome browser profile path.
    Assumes a browser was previously launched and the user logged in.
    """
    driver = None
    # The path we will test is the one we created in the 'start_browser' step.
    detected_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], 'chrome_setup_profile'))
    profile_name_for_setup = "Default" # For the setup wizard, we assume the 'Default' profile

    try:
        if not detected_path or not os.path.isdir(detected_path):
            return jsonify({'status': 'error', 'message': f"Could not auto-detect a valid profile path. Detected: {detected_path}"}), 500
        
        logger.info(f"[Setup Wizard] Detected profile path: {detected_path}")

        # --- Step 2: Test the detected profile path ---
        logger.info(f"[Setup Wizard] Testing detected path by launching a new undetected_chromedriver instance...")
        test_options = uc.ChromeOptions()
        user_data_dir = detected_path
        test_options.user_data_dir = user_data_dir
        test_options.profile_directory = profile_name_for_setup

        test_driver = uc.Chrome(options=test_options, headless=True)
        test_driver.get("https://www.linkedin.com/feed/")
        time.sleep(4) # Wait for page to load and redirect if not logged in

        # A simple but effective test: if we are redirected to /login, the test fails.
        is_logged_in = "linkedin.com/feed/" in test_driver.current_url
        test_driver.quit()

        if not is_logged_in:
            logger.warning("[Setup Wizard] Test failed. LinkedIn login page was detected.")
            return jsonify({'status': 'error', 'message': 'Test Failed: The detected profile is not logged into LinkedIn. Please try again.'}), 400

        logger.info("[Setup Wizard] Test successful! User is logged into LinkedIn.")

        # --- Step 3: Save the successful path to the database ---
        db = JobDatabase()
        user_details = db.get_user_details(current_user.id) or {}
        user_details['browser_profile_path'] = user_data_dir # Save the parent 'User Data' directory
        user_details['browser_profile_name'] = profile_name_for_setup # NEW: Save the profile name used during setup
        db.save_user_details(current_user.id, user_details)

        return jsonify({'status': 'success', 'message': 'Setup complete! Your LinkedIn profile is connected.'})

    except Exception as e:
        logger.error(f"Error during Assisted Apply setup: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'An unexpected error occurred: {e}'}), 500

@app.route('/people_radar')
@login_required
def people_radar():
    """Renders the People Radar page for managing contacts."""
    outreach_db = OutreachDatabase()
    contacts = outreach_db.get_contacts_by_user(current_user.id)
    templates = outreach_db.get_templates_by_user(current_user.id)
    return render_template('people_radar.html', contacts=contacts, templates=templates)

@app.route('/setup_assisted_apply/test', methods=['POST'])
@login_required
def test_assisted_apply_connection():
    """Tests the currently saved browser profile path."""
    db = JobDatabase()
    user_details = db.get_user_details(current_user.id)
    profile_path = user_details.get('browser_profile_path') if user_details else None
    profile_name = user_details.get('browser_profile_name', 'Default') if user_details else 'Default' # NEW: Get profile name

    if not profile_path or not os.path.isdir(profile_path):
        return jsonify({'status': 'error', 'message': 'No profile path is configured or the path is invalid.'}), 400

    driver = None
    try:
        options = uc.ChromeOptions()
        options.user_data_dir = profile_path
        options.profile_directory = profile_name

        driver = uc.Chrome(options=options, headless=True)
        
        driver.get("https://www.linkedin.com/feed/")
        time.sleep(4)

        is_logged_in = "linkedin.com/feed/" in driver.current_url
        driver.quit()

        if is_logged_in:
            return jsonify({'status': 'success', 'message': 'Connection successful! You are logged into LinkedIn.'})
        else:
            return jsonify({'status': 'error', 'message': 'Connection test failed. You are not logged into LinkedIn with this profile.'}), 400
    except Exception as e:
        logger.error(f"Error testing profile connection: {e}", exc_info=True)
        if driver: driver.quit()
        return jsonify({'status': 'error', 'message': f'Test failed with an error: {e}'}), 500

@app.route('/open_browser_session', methods=['POST'])
@login_required
def open_browser_session():
    """
    Launches a non-headless Chrome browser using the user's saved profile path.
    This is useful for debugging or manual logins.
    """
    db = JobDatabase()
    user_details = db.get_user_details(current_user.id)
    profile_path = user_details.get('browser_profile_path') if user_details else None
    profile_name = user_details.get('browser_profile_name', 'Default') if user_details else 'Default' # NEW: Get profile name

    if not profile_path or not os.path.isdir(profile_path):
        return jsonify({'status': 'error', 'message': 'No valid browser profile path is configured. Please run the setup wizard.'}), 400

    try:
        options = uc.ChromeOptions()
        options.user_data_dir = profile_path
        options.profile_directory = profile_name

        # For opening a browser for the user, we must not use headless mode.
        # The use_subprocess=True is important for keeping it open.
        driver = uc.Chrome(options=options, headless=False, use_subprocess=True)
        driver.get("https://www.linkedin.com/feed/") # Navigate to a useful page

        return jsonify({'status': 'success', 'message': 'Browser session opened successfully.'})
    except Exception as e:
        logger.error(f"Error opening browser session for user {current_user.id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'Failed to open browser: {e}'}), 500

@app.route('/delete_contact/<int:contact_id>', methods=['POST'])
@login_required
def delete_contact(contact_id):
    """Deletes a single personal contact for the current user."""
    outreach_db = OutreachDatabase()
    if outreach_db.delete_contact(contact_id, current_user.id):
        flash('Contact deleted successfully.', 'success')
    else:
        flash('Failed to delete contact. It may not exist or you may not have permission.', 'danger')
    return redirect(url_for('people_radar'))

@app.route('/delete_all_personal_contacts', methods=['POST'])
@login_required
def delete_all_personal_contacts():
    """Deletes all personal contacts for the current user."""
    try:
        outreach_db = OutreachDatabase()
        count = outreach_db.delete_all_contacts_for_user(current_user.id)
        flash(f'Successfully deleted all {count} of your personal contacts.', 'success')
    except Exception as e:
        logger.error(f"Error clearing personal contacts for user {current_user.id}: {e}", exc_info=True)
        flash(f'An error occurred while clearing your contacts: {e}', 'danger')
    return redirect(url_for('people_radar'))

@app.route('/admin/upload_contacts', methods=['GET', 'POST'])
@login_required
def admin_upload_contacts():
    # In a real app, you would add a check here to ensure only admin users can access this.
    # For now, any logged-in user can access it for demonstration.
    if request.method == 'POST':
        if 'contacts_csv' not in request.files:
            flash('No file part in the request.', 'danger')
            return redirect(request.url)
        file = request.files['contacts_csv']
        if file.filename == '':
            flash('No file selected for uploading.', 'warning')
            return redirect(request.url)
        if file and file.filename.endswith('.csv'):
            try:
                # --- New Robust Approach using standard `csv` library ---
                # Save the file to the repository for persistence and scraper access
                filename = secure_filename(file.filename)
                timestamp = int(time.time())
                save_path = os.path.join(CSV_UPLOAD_FOLDER, f"admin_{timestamp}_{filename}")
                file.save(save_path)
                logger.info(f"Saved admin upload CSV to {save_path}")

                # Read from the saved file
                with open(save_path, 'r', encoding='utf-8-sig') as f:
                    csv_text = f.read()

                reader = csv.DictReader(csv_text.splitlines())
                
                # --- Smarter Header Mapping ---
                # This map allows for flexible CSV column names.
                header_map = {
                    'full_name': ['full_name', 'name', 'contact name', 'contact', 'full name', 'customer name'],
                    'title': ['title', 'job title', 'position', 'role', 'job_title', 'job position'],
                    'company': ['company', 'company name', 'organization', 'employer', 'business', 'corp'],
                    'email': ['email', 'email address', 'e-mail', 'mail', 'email_address', 'contact email'],
                    'phone': ['phone', 'phone number', 'contact phone', 'contact_phone'],
                    'linkedin_url': ['linkedin_url', 'linkedin', 'linkedin profile', 'profile url', 'linkedin_profile', 'social profile']
                }

                contacts_to_add = []
                for row in reader:
                    processed_row = {}
                    # Normalize the keys from the uploaded file once
                    normalized_input_row = {}
                    for key, value in row.items():
                        if key:
                            normalized_input_row[key.lower().strip().replace(' ', '_')] = value

                    # Map the flexible headers to our standard database columns
                    for db_key, possible_headers in header_map.items():
                        for header in possible_headers:
                            if header in normalized_input_row:
                                processed_row[db_key] = normalized_input_row[header]
                                break # Move to the next db_key once found
                    
                    # Generate LinkedIn URL if missing but Email exists (to prevent duplicates)
                    if not processed_row.get('linkedin_url') and processed_row.get('email'):
                        email_hash = hashlib.md5(processed_row['email'].lower().encode()).hexdigest()
                        processed_row['linkedin_url'] = f"csv-upload-{email_hash}"

                    # Only add if we have at least a name or email
                    if processed_row.get('full_name') or processed_row.get('email'):
                        # Add metadata
                        processed_row['source'] = 'Uploaded CSV'
                        processed_row['added_date'] = datetime.datetime.now().isoformat()
                        contacts_to_add.append(processed_row)

                # --- Diagnostic Check ---
                if contacts_to_add:
                    first_contact_preview = contacts_to_add[0]
                    flash(f"Diagnostic: First contact read from CSV -> Name: {first_contact_preview.get('full_name')}, Company: {first_contact_preview.get('company')}", 'info')

                outreach_db = OutreachDatabase()
                count = outreach_db.add_global_contacts(contacts_to_add)
                flash(f'Successfully processed {count} contacts from the CSV. Duplicates were ignored.', 'success')
            except Exception as e:
                logger.error(f"Error processing uploaded CSV: {e}", exc_info=True)
                flash(f'An error occurred while processing the file: {e}', 'danger')
            return redirect(url_for('admin_upload_contacts'))

    return render_template('admin_upload.html')

@app.route('/admin/clear_global_contacts', methods=['POST'])
@login_required
def clear_global_contacts():
    # Add admin check here in a real application
    try:
        outreach_db = OutreachDatabase()
        count = outreach_db.clear_global_contacts()
        flash(f'Successfully deleted {count} global contacts.', 'success')
    except Exception as e:
        logger.error(f"Error clearing global contacts: {e}", exc_info=True)
        flash(f'An error occurred while clearing contacts: {e}', 'danger')
    return redirect(url_for('admin_upload_contacts'))

@app.route('/signal_board')
@login_required
def signal_board():
    """Renders the Signal Board page for tracking campaigns."""
    outreach_db = OutreachDatabase()
    campaigns = outreach_db.get_campaigns_with_analytics(current_user.id)
    return render_template('signal_board.html', campaigns=campaigns)

@app.route('/api/send_message', methods=['POST'])
@login_required
def api_send_message():
    """API endpoint to send a personalized email to a contact."""
    data = request.get_json()
    contact_email = data.get('contact_email')
    subject = data.get('subject')
    body = data.get('body')

    if not all([contact_email, subject, body]):
        return jsonify({'success': False, 'message': 'Missing required fields.'}), 400

    try:
        emailer = SmartEmailer()
        if not emailer.gmail_service:
            return jsonify({'success': False, 'message': 'Gmail not authenticated. Please connect your Google account.'}), 500
        
        emailer.send_email(subject=subject, html_content=body.replace('\n', '<br>'), recipients=[contact_email])
        
        return jsonify({'success': True, 'message': 'Email sent successfully!'})
    except Exception as e:
        logger.error(f"Error sending email via API for user {current_user.id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'An error occurred: {e}'}), 500

@app.route('/message_vault')
@login_required
def message_vault():
    """Renders the Message Vault page for managing templates."""
    outreach_db = OutreachDatabase()
    templates = outreach_db.get_templates_by_user(current_user.id)
    return render_template('message_vault.html', templates=templates)

@app.route('/api/templates', methods=['POST'])
@login_required
def add_template():
    """API endpoint to add a new template."""
    data = request.get_json()
    outreach_db = OutreachDatabase()
    template_id = outreach_db.add_template(current_user.id, data)
    if template_id:
        return jsonify({'success': True, 'message': 'Template saved!', 'template_id': template_id}), 201
    return jsonify({'success': False, 'message': 'A template with this name already exists.'}), 409

@app.route('/api/templates/<int:template_id>', methods=['PUT'])
@login_required
def update_template(template_id):
    """API endpoint to update an existing template."""
    data = request.get_json()
    outreach_db = OutreachDatabase()
    if outreach_db.update_template(template_id, current_user.id, data):
        return jsonify({'success': True, 'message': 'Template updated!'})
    return jsonify({'success': False, 'message': 'Update failed. Template not found or name conflict.'}), 400

@app.route('/api/templates/<int:template_id>', methods=['DELETE'])
@login_required
def delete_template(template_id):
    """API endpoint to delete a template."""
    outreach_db = OutreachDatabase()
    if outreach_db.delete_template(template_id, current_user.id):
        return jsonify({'success': True, 'message': 'Template deleted!'})
    return jsonify({'success': False, 'message': 'Delete failed. Template not found.'}), 404

@app.route('/api/campaigns', methods=['POST'])
@login_required
def create_campaign():
    """API endpoint to create a new campaign."""
    data = request.get_json()
    if not data or not data.get('campaign_name'):
        return jsonify({'success': False, 'message': 'Campaign name is required.'}), 400
    
    outreach_db = OutreachDatabase()
    campaign_id = outreach_db.add_campaign(current_user.id, data)
    if campaign_id:
        return jsonify({'success': True, 'message': 'Campaign created successfully.', 'campaign_id': campaign_id})
    return jsonify({'success': False, 'message': 'Failed to create campaign. Name might be duplicate.'}), 400

@app.route('/api/campaigns/<int:campaign_id>/status', methods=['POST'])
@login_required
def update_campaign_status_route(campaign_id):
    """API endpoint to update campaign status (e.g., set to 'active')."""
    data = request.get_json()
    status = data.get('status')
    if status not in ['draft', 'active', 'completed', 'paused']:
        return jsonify({'success': False, 'message': 'Invalid status.'}), 400
        
    outreach_db = OutreachDatabase()
    if outreach_db.update_campaign_status(campaign_id, current_user.id, status):
        return jsonify({'success': True, 'message': f'Campaign status updated to {status}.'})
    return jsonify({'success': False, 'message': 'Campaign not found.'}), 404

@app.route('/api/campaigns/<int:campaign_id>', methods=['DELETE'])
@login_required
def delete_campaign_route(campaign_id):
    """API endpoint to delete a campaign."""
    outreach_db = OutreachDatabase()
    if outreach_db.delete_campaign(campaign_id, current_user.id):
        return jsonify({'success': True, 'message': 'Campaign deleted.'})
    return jsonify({'success': False, 'message': 'Campaign not found.'}), 404


@app.route('/clear_profile', methods=['POST'])
@login_required
def clear_profile():
    """Deletes all profile information for the current user."""
    db = JobDatabase()
    # The action is considered successful even if there was nothing to delete.
    db.delete_user_details(current_user.id)
    # Return a JSON response for the frontend to handle.
    return jsonify({
        'success': True,
        'message': 'Your profile information has been cleared.'
    })

# --- Import custom scrapers ---
from outreach_scraper import LegalContactResearcher

# --- Enhanced Contact Search Term Mapping ---
HR_SEARCH_TERMS_MAP = {
    # Job Function
    "manager": ["HR Manager", "Recruitment Manager", "Talent Acquisition Manager", "People Operations Manager"],
    "specialist": ["Talent Acquisition Specialist", "HR Business Partner", "Senior Recruiter"],
    "director": ["Head of HR", "HR Director", "CHRO (Chief Human Resources Officer)"],
    "leader": ["Head of HR", "HR Director", "CHRO (Chief Human Resources Officer)", "Executive HR Leader (15+ years)"],
    # Industry
    "tech": ["AI/ML HR Manager", "Tech Startup Recruiter", "SaaS Company HR Manager", "Cybersecurity HR Manager"],
    "ai": ["AI/ML HR Manager"],
    "ml": ["AI/ML HR Manager"],
    "fintech": ["Fintech HR Manager"],
    "health": ["Healthcare HR Specialist"],
    "ecommerce": ["E-commerce Talent Acquisition"],
    "saas": ["SaaS Company HR Manager"],
    "crypto": ["Blockchain/Crypto HR"],
    "blockchain": ["Blockchain/Crypto HR"],
    "edtech": ["EdTech Recruiter"],
    "gaming": ["Gaming Industry HR"],
    "cybersecurity": ["Cybersecurity HR Manager"],
    # Location
    "bangalore": ["Bangalore HR Manager"],
    "mumbai": ["Mumbai Talent Acquisition"],
    "delhi": ["Delhi/NCR HR Director"],
    "ncr": ["Delhi/NCR HR Director"],
    "hyderabad": ["Hyderabad Tech Recruiter"],
    "pune": ["Pune IT HR Manager"],
    "chennai": ["Chennai Software HR"],
    # Experience
    "entry": ["Entry Level HR Coordinator"],
    "junior": ["Entry Level HR Coordinator"],
    "mid": ["Mid-level HR Manager (3-7 years)"],
    "senior": ["Senior HR Manager (7-15 years)", "Senior Recruiter"],
    "executive": ["Executive HR Leader (15+ years)", "CHRO (Chief Human Resources Officer)"],
    # Hot Keywords
    "remote": ["Remote Work HR Specialist"],
    "diversity": ["Diversity & Inclusion Manager"],
    "inclusion": ["Diversity & Inclusion Manager"],
    "experience": ["Employee Experience Manager"],
    "analytics": ["HR Analytics Manager"],
    "compensation": ["Compensation & Benefits Manager"],
    "benefits": ["Compensation & Benefits Manager"],
    "learning": ["Learning & Development HR"],
    "development": ["Learning & Development HR"],
}

def search_csv_repository_for_contacts(user_id: int, keywords: List[str]) -> int:
    """
    Iterates through all CSV files in the repository and adds matching contacts.
    """
    added_count = 0
    outreach_db = OutreachDatabase()
    
    if not os.path.exists(CSV_UPLOAD_FOLDER):
        logger.warning(f"CSV repository folder not found: {CSV_UPLOAD_FOLDER}")
        return 0

    # Flexible header mapping
    header_map = {
        'full_name': ['full_name', 'name', 'contact name', 'contact', 'full name', 'customer name', 'first name', 'last name'],
        'title': ['title', 'job title', 'position', 'role', 'job_title', 'job position'],
        'company': ['company', 'company name', 'organization', 'employer', 'business', 'corp'],
        'email': ['email', 'email address', 'e-mail', 'mail', 'email_address', 'contact email'],
        'linkedin_url': ['linkedin_url', 'linkedin', 'linkedin profile', 'profile url', 'linkedin_profile', 'social profile']
    }

    logger.info(f"Searching CSV repository for keywords: {keywords}")

    for filename in os.listdir(CSV_UPLOAD_FOLDER):
        if not filename.endswith('.csv'):
            continue
            
        filepath = os.path.join(CSV_UPLOAD_FOLDER, filename)
        try:
            with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv.DictReader(f)
                
                file_headers = reader.fieldnames
                if not file_headers:
                    continue
                    
                # Create a mapping for this specific file
                file_map = {}
                normalized_headers = {h.lower().strip().replace(' ', '_'): h for h in file_headers}
                
                for db_key, possible_matches in header_map.items():
                    for match in possible_matches:
                        if match in normalized_headers:
                            file_map[db_key] = normalized_headers[match]
                            break
                
                has_split_name = 'full_name' not in file_map and \
                                 ('first_name' in normalized_headers or 'last_name' in normalized_headers)
                
                for row in reader:
                    contact_data = {}
                    for db_key, file_header in file_map.items():
                        if file_header in row and row[file_header]:
                            contact_data[db_key] = row[file_header].strip()
                    
                    if 'full_name' not in contact_data and has_split_name:
                        first = row.get(normalized_headers.get('first_name', ''), '').strip()
                        last = row.get(normalized_headers.get('last_name', ''), '').strip()
                        if first or last:
                            contact_data['full_name'] = f"{first} {last}".strip()

                    search_text = f"{contact_data.get('full_name', '')} {contact_data.get('title', '')} {contact_data.get('company', '')}".lower()
                    
                    if any(k.lower() in search_text for k in keywords):
                        if contact_data.get('full_name') or contact_data.get('email'):
                            if not contact_data.get('linkedin_url'):
                                if contact_data.get('email'):
                                    email_hash = hashlib.md5(contact_data['email'].lower().encode()).hexdigest()
                                    contact_data['linkedin_url'] = f"csv-repo-{email_hash}"
                                else:
                                    unique_str = f"{contact_data.get('full_name')}-{contact_data.get('company')}"
                                    unique_hash = hashlib.md5(unique_str.encode()).hexdigest()
                                    contact_data['linkedin_url'] = f"csv-repo-noemail-{unique_hash}"
                            
                            contact_data['source'] = f'CSV Repo: {filename}'
                            if outreach_db.add_contact(user_id, contact_data):
                                added_count += 1
        except Exception as e:
            logger.error(f"Error searching CSV {filename}: {e}")
            continue
    return added_count

# --- Celery Tasks for Outreach ---
@celery.task(bind=True, name='app.scrape_contacts_task')
def scrape_contacts_task(self, user_id: int, keywords: List[str]):
    """Celery task to scrape contacts from various platforms."""
    logger.info(f"Celery task {self.request.id} started for user {user_id} to find contacts with keywords: {keywords}.")
    try:
        outreach_db = OutreachDatabase()
        
        # 1. Search Internal Database
        self.update_state(state='PROGRESS', meta={'status': 'Searching internal database...', 'progress': 20})
        internal_count = outreach_db.search_and_copy_global_contacts(user_id, keywords)

        # 2. Search CSV Repository
        self.update_state(state='PROGRESS', meta={'status': 'Searching CSV repository...', 'progress': 30})
        csv_count = search_csv_repository_for_contacts(user_id, keywords)

        # 3. Research External Sources (Public Directories & Career Pages)
        self.update_state(state='PROGRESS', meta={'status': 'Researching public sources...', 'progress': 50})
        researcher = LegalContactResearcher(user_id, task=self)
        
        # Use keywords to drive the external research
        search_query = " ".join(keywords)
        
        # Try to extract potential company names from keywords
        # Companies are typically capitalized or multi-word terms
        potential_companies = []
        for keyword in keywords:
            # If keyword has capital letters or multiple words, it might be a company
            if any(c.isupper() for c in keyword) or len(keyword.split()) > 1:
                potential_companies.append(keyword)
        
        # If no companies detected, pass None to let the researcher work with just keywords
        companies_arg = potential_companies if potential_companies else None
        
        logger.info(f"Extracted potential companies: {companies_arg}")
        external_results = researcher.run(keywords=search_query, companies=companies_arg)
        
        external_count = 0
        if external_results.get('contacts'):
            # Process found contacts
            for contact in external_results['contacts']:
                # Map scraper result to DB schema
                emails = contact.get('contact_info', {}).get('emails', [])
                if emails:
                    contact_data = {
                        'full_name': 'HR Contact', # Placeholder as scraper might not get names
                        'title': 'Recruiter / HR',
                        'company': contact.get('company', 'Unknown'),
                        'email': emails[0],
                        'linkedin_url': contact.get('source_url', f"scraped-{time.time()}"),
                        'source': contact.get('source_type', 'External Research')
                    }
                    if outreach_db.add_contact(user_id, contact_data):
                        external_count += 1

        total_count = internal_count + csv_count + external_count
        message = f"Search complete. Found {internal_count} internal, {csv_count} from CSVs, and {external_count} external contacts."
        logger.info(message)
        
        # Final update
        self.update_state(state='PROGRESS', meta={'status': message, 'progress': 100})
        
        return {'status': 'Complete', 'message': message, 'progress': 100, 'added_count': total_count}
    except Exception as e:
        logger.error(f"Celery task {self.request.id} failed for user {user_id}: {e}", exc_info=True)
        self.update_state(state='FAILURE', meta={'status': f'Task failed: {str(e)}', 'error': str(e), 'progress': 100})
        raise


@app.route('/scrape_contacts', methods=['POST'])
@login_required
def scrape_contacts():
    data = request.get_json()
    keywords_str = data.get('keywords', '') if data else ''
    keywords = [term.strip() for term in keywords_str.split(',') if term.strip()]
    task = scrape_contacts_task.delay(current_user.id, keywords)
    return jsonify({'success': True, 'task_id': task.id})

@app.route('/verify_email', methods=['POST'])
@login_required
def verify_email():
    """Verifies if an email address exists using SMTP checks."""
    data = request.get_json()
    email = data.get('email')
    
    if not email:
        return jsonify({'exists': False, 'message': 'Email is required.'}), 400

    # 1. Basic Syntax Check
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        return jsonify({'exists': False, 'message': 'Invalid email format.'})

    # 2. SMTP Verification
    try:
        if 'dns.resolver' not in sys.modules:
             # Fallback if dnspython is missing
             return jsonify({'exists': True, 'message': 'Syntax valid (SMTP check skipped - dnspython missing).'})

        domain = email.split('@')[-1]
        
        # Get MX Record
        try:
            records = dns.resolver.resolve(domain, 'MX')
            mx_record = str(records[0].exchange)
        except Exception:
            return jsonify({'exists': False, 'message': f'Could not resolve MX records for {domain}.'})

        # Connect to SMTP Server
        # Note: Port 25 is often blocked by residential ISPs.
        server = smtplib.SMTP(timeout=5)
        server.set_debuglevel(0)
        
        server.connect(mx_record)
        server.helo(server.local_hostname or 'localhost')
        server.mail('verify@example.com') # Use a generic sender
        code, message = server.rcpt(email)
        server.quit()

        if code == 250:
            return jsonify({'exists': True, 'message': 'Email address is valid and deliverable.'})
        else:
            return jsonify({'exists': False, 'message': f'Email rejected by server (Code {code}).'})
            
    except Exception as e:
        logger.warning(f"SMTP verification failed for {email}: {e}")
        return jsonify({'exists': False, 'message': f'Verification failed: {str(e)}'})

@app.route('/add_contact', methods=['POST'])
@login_required
def add_contact():
    """Manually adds a contact to the database."""
    data = request.get_json()
    
    if not data or not data.get('name') or not data.get('email'):
        return jsonify({'success': False, 'message': 'Name and Email are required.'}), 400

    outreach_db = OutreachDatabase()
    
    # Generate a placeholder LinkedIn URL if not provided (DB requires uniqueness)
    linkedin_url = data.get('linkedin_url')
    if not linkedin_url:
        email_hash = hashlib.md5(data['email'].lower().encode()).hexdigest()
        linkedin_url = f"manual-entry-{email_hash}"

    contact_data = {
        'full_name': data['name'],
        'title': data.get('role', ''),
        'company': data.get('company', ''),
        'email': data['email'],
        'linkedin_url': linkedin_url,
        'source': 'Manual Entry'
    }

    if outreach_db.add_contact(current_user.id, contact_data):
        return jsonify({'success': True, 'message': 'Contact added successfully.'})
    else:
        return jsonify({'success': False, 'message': 'Failed to add contact. It may already exist.'})

@app.route('/import_contacts', methods=['POST'])
@login_required
def import_contacts():
    """Bulk imports contacts from a CSV file for the current user."""
    if 'contacts_csv' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded.'}), 400
    
    file = request.files['contacts_csv']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected.'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'message': 'File must be a CSV.'}), 400

    try:
        # Save the file to the repository
        filename = secure_filename(file.filename)
        timestamp = int(time.time())
        save_path = os.path.join(CSV_UPLOAD_FOLDER, f"import_{timestamp}_{filename}")
        file.save(save_path)
        logger.info(f"Saved user import CSV to {save_path}")

        with open(save_path, 'r', encoding='utf-8-sig') as f:
            csv_text = f.read()

        reader = csv.DictReader(csv_text.splitlines())
        
        header_map = {
            'full_name': ['full_name', 'name', 'contact name', 'contact', 'full name', 'customer name'],
            'title': ['title', 'job title', 'position', 'role', 'job_title', 'job position'],
            'company': ['company', 'company name', 'organization', 'employer', 'business', 'corp'],
            'email': ['email', 'email address', 'e-mail', 'mail', 'email_address', 'contact email'],
            'linkedin_url': ['linkedin_url', 'linkedin', 'linkedin profile', 'profile url', 'linkedin_profile', 'social profile']
        }

        outreach_db = OutreachDatabase()
        added_count = 0
        global_contacts_to_add = []
        
        for row in reader:
            processed_row = {}
            normalized_row = {k.lower().strip().replace(' ', '_'): v.strip() for k, v in row.items() if k}
            
            for db_key, possible_headers in header_map.items():
                for header in possible_headers:
                    if header in normalized_row:
                        processed_row[db_key] = normalized_row[header]
                        break
            
            if 'full_name' not in processed_row:
                first = normalized_row.get('first_name', '')
                last = normalized_row.get('last_name', '')
                if first or last: processed_row['full_name'] = f"{first} {last}".strip()

            if processed_row.get('full_name'):
                if not processed_row.get('linkedin_url') and processed_row.get('email'):
                    email_hash = hashlib.md5(processed_row['email'].lower().encode()).hexdigest()
                    processed_row['linkedin_url'] = f"csv-import-{email_hash}"
                
                processed_row['source'] = 'CSV Import'
                if outreach_db.add_contact(current_user.id, processed_row):
                    added_count += 1
                
                # Prepare for global database update
                global_entry = processed_row.copy()
                global_entry['added_date'] = datetime.datetime.now().isoformat()
                global_contacts_to_add.append(global_entry)

        # Update the original database (global_hiring_contacts)
        if global_contacts_to_add:
            global_count = outreach_db.add_global_contacts(global_contacts_to_add)
            logger.info(f"Imported {global_count} new contacts to global database from user upload.")

        return jsonify({'success': True, 'message': f'Successfully imported {added_count} contacts.'})

    except Exception as e:
        logger.error(f"CSV Import Error: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Error processing CSV: {str(e)}'}), 500

def add_status_colors(job):
    """Add status color attributes to job dictionary for UI display"""
    status_colors = {
        'new': {'bg': '#f3f4f6', 'text': '#374151'},
        'applied': {'bg': '#dcfce7', 'text': '#166534'},
        'interview': {'bg': '#e0e7ff', 'text': '#3730a3'},
        'offered': {'bg': '#fef3c7', 'text': '#92400e'},
        'rejected': {'bg': '#fee2e2', 'text': '#991b1b'},
        'withdrawn': {'bg': '#f3f4f6', 'text': '#6b7280'}
    }
    status = job.get('status', 'new')
    colors = status_colors.get(status, status_colors['new'])
    job['status_color'] = colors['bg']
    job['status_text_color'] = colors['text']
    
    # Add status class and text for template
    status_class_map = {
        'new': 'bg-gray-100 text-gray-700',
        'applied': 'bg-green-100 text-green-700',
        'interview': 'bg-blue-100 text-blue-700',
        'offered': 'bg-yellow-100 text-yellow-700',
        'rejected': 'bg-red-100 text-red-700',
        'withdrawn': 'bg-gray-100 text-gray-600'
    }
    status_text_map = {
        'new': 'New',
        'applied': 'Applied',
        'interview': 'Interview',
        'offered': 'Offered',
        'rejected': 'Rejected',
        'withdrawn': 'Withdrawn'
    }
    job['status_class'] = status_class_map.get(status, 'bg-gray-100 text-gray-700')
    job['status_text'] = status_text_map.get(status, status.title() if status else 'New')
    return job

@app.route('/dashboard')
@login_required
def dashboard():
    db = JobDatabase()
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 10, type=int)
    offset = (page - 1) * limit
    sort_by = request.args.get('sort_by', 'found_date')
    sort_order = request.args.get('sort_order', 'DESC')
    search_query = request.args.get('search_query', '')
    status_filter = request.args.get('status_filter', 'all')
    location_filter = request.args.get('location_filter', '')
    job_type_filter = request.args.get('job_type_filter', 'all')
    view_mode = 'detailed'  # Always use detailed view

    jobs, total_jobs = db.get_jobs_for_user(
        current_user.id,
        limit=limit,
        offset=offset,
        sort_by=sort_by,
        sort_order=sort_order,
        search_query=search_query,
        status_filter=status_filter,
        location_filter=location_filter,
        job_type_filter=job_type_filter
    )
    
    # Add status colors to jobs
    jobs = [add_status_colors(job) for job in jobs]
    
    saved_profiles = db.get_search_profiles(current_user.id)
    user_settings = db.get_user_settings(current_user.id)
    
    total_pages = (total_jobs + limit - 1) // limit if limit > 0 else 1

    # Generate a window of page numbers for pagination controls
    page_window = 2  # Number of pages to show before and after the current page
    start_page = max(1, page - page_window)
    end_page = min(total_pages, page + page_window)
    if end_page < start_page: # Handle case where total_pages is very small
        end_page = start_page
    pagination_pages = range(start_page, end_page + 1)

    return render_template('dashboard.html', 
                           jobs=jobs, 
                           saved_profiles=saved_profiles,
                           user_settings=user_settings,
                           current_page=page,
                           total_pages=total_pages,
                           limit=limit,
                           pagination_pages=pagination_pages,
                           sort_by=sort_by,
                           sort_order=sort_order,
                           search_query=search_query,
                           status_filter=status_filter,
                           location_filter=location_filter,
                           job_type_filter=job_type_filter,
                           total_jobs=total_jobs,
                           view_mode=view_mode,
                           now=datetime.datetime.now())

@app.route('/job/<int:job_id>')
@login_required
def job_details(job_id):
    """Display full page for job details"""
    db = JobDatabase()
    jobs, _ = db.get_jobs_for_user(current_user.id, limit=1, search_query=f"id:{job_id}")
    
    if not jobs:
        flash('Job not found or you do not have permission to view it.', 'error')
        return redirect(url_for('dashboard'))
    
    job = jobs[0]
    # Add status colors
    job = add_status_colors(job)
    
    return render_template('job_details.html', job=job, now=datetime.datetime.now())

@app.route('/search_jobs', methods=['POST'])
@login_required
def search_jobs():
    search_terms_str = request.form.get('search_terms', '')
    location = request.form.get('location', '')
    experience = request.form.get('experience', '')
    job_type = request.form.get('job_type', '')

    search_terms = [term.strip() for term in search_terms_str.split(',') if term.strip()]

    if not search_terms:
        flash('Please enter at least one search term.', 'warning')
        return redirect(url_for('dashboard'))

    # Trigger the Celery task for job scraping
    try:
        task = run_job_hunt_task.delay(current_user.id, search_terms, location, experience, job_type)
        logger.info(f"Celery task {task.id} enqueued for user {current_user.id} with terms '{search_terms_str}'.")
        return jsonify({'status': 'success', 'task_id': task.id})
    except Exception as e:
        logger.error(f"Failed to enqueue Celery task for user {current_user.id} with terms '{search_terms_str}': {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to start search. Celery worker might be down.'}), 500

@app.route('/save_profile', methods=['POST'])
@login_required
def save_profile():
    # Handle both JSON and form data
    if request.is_json:
        data = request.get_json()
        profile_name = data.get('profile_name')
        search_terms = data.get('search_terms')
        location = data.get('location', '')
        experience = data.get('experience', '')
        job_type = data.get('job_type', '')
    else:
        profile_name = request.form['profile_name']
        search_terms = request.form['search_terms']
        location = request.form['location']
        experience = request.form.get('experience', '')
        job_type = request.form.get('job_type', '')

    if not profile_name or not search_terms:
        if request.is_json:
            return jsonify({'error': 'Profile name and search terms are required'}), 400
        flash('Profile name and search terms are required', 'danger')
        return redirect(url_for('dashboard'))

    db = JobDatabase()
    profile_id = db.save_search_profile(current_user.id, profile_name, search_terms, location, experience, job_type)
    
    if profile_id:
        if request.is_json:
            return jsonify({
                'success': True,
                'message': f"Search profile '{profile_name}' saved successfully!",
                'profile_id': profile_id
            })
        flash(f"Search profile '{profile_name}' saved successfully!", 'success')
    else:
        if request.is_json:
            return jsonify({'error': f"Failed to save profile. A profile with name '{profile_name}' might already exist."}), 400
        flash(f"Failed to save profile. A profile with name '{profile_name}' might already exist.", 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/delete_profile/<int:profile_id>', methods=['POST'])
@login_required
def delete_profile(profile_id):
    db = JobDatabase()
    if db.delete_search_profile(profile_id, current_user.id):
        flash('Search profile deleted successfully!', 'success')
    else:
        flash('Failed to delete search profile.', 'danger')
    return redirect(url_for('dashboard'))

@app.route('/apply_profile/<int:profile_id>', methods=['POST'])
@login_required
def apply_profile(profile_id):
    db = JobDatabase()
    try:
        profile = db.get_search_profile_by_id(profile_id, current_user.id)
        if not profile:
            return jsonify({'error': 'Profile not found or access denied.'}), 404
        
        # Always return JSON for apply_profile since it's used by JavaScript
        return jsonify({
            'success': True,
            'profile': {
                'profile_name': profile['profile_name'],
                'search_terms': profile['search_terms'],
                'location': profile['location'] or '',
                'experience': profile['experience'] or '',
                'job_type': profile['job_type'] or ''
            }
        })
        
        # Note: Original behavior for starting search task is now handled by the search form
    except Exception as e:
        logger.error(f"Failed to apply profile for user {current_user.id} from profile {profile_id}: {e}", exc_info=True)
        return jsonify({'error': 'Failed to apply profile.'}), 500

@app.route('/delete_job/<int:job_id>', methods=['POST'])
@login_required
def delete_job(job_id):
    db = JobDatabase()
    if db.delete_job(job_id, current_user.id):
        flash('Job deleted successfully.', 'success')
    else:
        flash('Job not found or could not be deleted.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/delete_selected', methods=['POST'])
@login_required
def delete_selected():
    job_ids = request.form.getlist('job_ids')
    job_ids = [int(jid) for jid in job_ids if jid.isdigit()]
    if not job_ids:
        flash('No jobs selected for deletion.', 'warning')
        return redirect(url_for('dashboard'))
    
    db = JobDatabase()
    deleted_count = db.delete_jobs(job_ids, current_user.id)
    if deleted_count > 0:
        flash(f'Successfully deleted {deleted_count} job(s).', 'success')
    else:
        flash('No jobs were deleted. They may not exist or belong to you.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/delete_all_jobs', methods=['POST'])
@login_required
def delete_all_jobs():
    db = JobDatabase()
    deleted_count = db.delete_all_jobs(current_user.id)
    if deleted_count > 0:
        flash(f'Successfully deleted all {deleted_count} jobs.', 'success')
    else:
        flash('No jobs to delete.', 'info')
    return redirect(url_for('dashboard'))

@app.route('/update_settings', methods=['POST'])
@login_required
def update_settings():
    email_recipients = request.form.get('email_recipients', '').strip()
    send_excel_attachment = 'send_excel_attachment' in request.form

    db = JobDatabase()
    if db.update_user_settings(current_user.id, email_recipients, send_excel_attachment):
        flash('Email settings updated successfully!', 'success')
    else:
        flash('Failed to update email settings.', 'danger')
    return redirect(url_for('dashboard'))

@app.route('/update_job_status', methods=['POST'])
@app.route('/update_job_status/<int:job_id>', methods=['POST'])
@login_required
def update_job_status(job_id=None):
    # Handle both URL parameter and form data
    if job_id is None:
        job_id = request.form.get('job_id', type=int)
    
    status = request.form.get('status')
    notes = request.form.get('notes', '')
    sync_to_calendar = request.form.get('sync_to_calendar') == '1'  # Check if checkbox is checked

    if not job_id or not status:
        if request.is_json or 'application/json' in request.headers.get('Accept', ''):
            return jsonify({'error': 'Job ID and status are required'}), 400
        flash('Job ID and status are required', 'danger')
        return redirect(url_for('dashboard'))

    db = JobDatabase()
    
    if db.update_job_status_and_notes(job_id, current_user.id, status, notes):
        success_message = 'Job status updated successfully!'
        
        # Only sync to calendar if user requested it
        if sync_to_calendar:
            # Get job details for calendar integration
            jobs, _ = db.get_jobs_for_user(current_user.id, limit=1, offset=0, 
                                           search_query=f"id:{job_id}")
            
            if jobs:
                job_data = jobs[0]
                calendar_manager = GoogleCalendarManager()
                
                try:
                    calendar_synced = False
                    
                    if status == 'interview':
                        result = calendar_manager.create_interview_event(job_data)
                        if result:
                            success_message = 'Job status updated and interview event created in calendar!'
                            calendar_synced = True
                    elif status == 'applied':
                        result = calendar_manager.create_followup_reminder(job_data, status)
                        if result:
                            success_message = 'Job status updated and follow-up reminder created!'
                            calendar_synced = True
                    elif status in ['offered', 'rejected']:
                        result = calendar_manager.create_followup_reminder(job_data, status)
                        if result:
                            success_message = f'Job status updated and {status} follow-up reminder created!'
                            calendar_synced = True
                    elif job_data.get('deadline') and job_data['deadline'] != 'N/A':
                        # Create deadline reminder for any status if deadline exists
                        result = calendar_manager.create_deadline_reminder(job_data, job_data['deadline'])
                        if result:
                            success_message = 'Job status updated and deadline reminder created!'
                            calendar_synced = True
                    
                    if not calendar_synced and sync_to_calendar:
                        success_message += ' (No calendar event needed for this status)'
                        
                except Exception as e:
                    logger.error(f"Calendar integration failed: {e}")
                    if request.is_json or 'application/json' in request.headers.get('Accept', ''):
                        return jsonify({'error': 'Job status updated, but calendar sync failed. Check your Google Calendar connection.'}), 500
                    flash('Job status updated, but calendar sync failed. Check your Google Calendar connection.', 'warning')
                    return redirect(url_for('dashboard'))
        
        # Return JSON response for AJAX requests
        if request.is_json or 'application/json' in request.headers.get('Accept', ''):
            # Get updated job data
            jobs, _ = db.get_jobs_for_user(current_user.id, limit=1, offset=0, 
                                           search_query=f"id:{job_id}")
            if jobs:
                job_data = jobs[0]
                return jsonify({
                    'success': True,
                    'message': success_message,
                    'job': {
                        'id': job_data['id'],
                        'status': job_data['status'],
                        'title': job_data['title'],
                        'company': job_data['company']
                    }
                })
            else:
                return jsonify({'error': 'Job not found after update'}), 404
        
        flash(success_message, 'success')
    else:
        if request.is_json or 'application/json' in request.headers.get('Accept', ''):
            return jsonify({'error': 'Failed to update job status'}), 500
        flash('Failed to update job status.', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/sync_job_to_calendar/<int:job_id>', methods=['POST'])
@login_required
def sync_job_to_calendar(job_id):
    """Sync a specific job to Google Calendar."""
    try:
        db = JobDatabase()
        calendar_manager = GoogleCalendarManager()
        
        if not calendar_manager.calendar_service:
            return jsonify({'error': 'Calendar not connected'}), 400
        
        # Get job details
        jobs, _ = db.get_jobs_for_user(current_user.id, limit=1, offset=0, 
                                       search_query=f"id:{job_id}")
        
        if not jobs:
            return jsonify({'error': 'Job not found'}), 404
        
        job_data = jobs[0]
        events_created = []
        
        # Create appropriate calendar events based on job status and data
        if job_data.get('status') == 'interview':
            result = calendar_manager.create_interview_event(job_data)
            if result:
                events_created.append('Interview event')
        
        if job_data.get('status') == 'applied':
            result = calendar_manager.create_followup_reminder(job_data, 'applied')
            if result:
                events_created.append('Follow-up reminder')
        
        if job_data.get('status') in ['offered', 'rejected']:
            result = calendar_manager.create_followup_reminder(job_data, job_data['status'])
            if result:
                events_created.append(f'{job_data["status"].title()} follow-up')
        
        if job_data.get('deadline') and job_data['deadline'] != 'N/A':
            result = calendar_manager.create_deadline_reminder(job_data, job_data['deadline'])
            if result:
                events_created.append('Deadline reminder')
        
        if events_created:
            message = f'Successfully created: {", ".join(events_created)}'
            return jsonify({'success': True, 'message': message, 'events_created': events_created})
        else:
            return jsonify({'success': True, 'message': 'No calendar events needed for this job', 'events_created': []})
        
    except Exception as e:
        logger.error(f"Failed to sync job {job_id} to calendar: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/feedback', methods=['POST'])
@login_required
def feedback():
    job_id = request.form.get('job_id', type=int)
    feedback_type = request.form.get('feedback_type') # 'like' or 'dislike'
    
    db = JobDatabase()
    if db.record_job_feedback(current_user.id, job_id, feedback_type):
        flash(f"Feedback recorded: {feedback_type}!", 'success')
        # Here, you might also trigger an update to custom scores based on feedback
        # For simplicity, let's assume feedback directly updates the job entry.
    else:
        flash('Failed to record feedback.', 'danger')
    return redirect(url_for('dashboard'))

@app.route('/assisted_apply/<int:job_id>', methods=['POST'])
@login_required
def assisted_apply(job_id):
    """Triggers the assisted apply Celery task."""
    try:
        # You could add a check here to see if the user has filled out their profile
        task = assisted_apply_task.delay(current_user.id, job_id)
        logger.info(f"Enqueued assisted_apply_task {task.id} for user {current_user.id}, job {job_id}.")
        return jsonify({'status': 'success', 'message': 'Task started! A browser window should open shortly.'})
    except Exception as e:
        logger.error(f"Failed to enqueue assisted_apply_task for user {current_user.id}, job {job_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to start task. Is the Celery worker running?'}), 500

@app.route('/analyze_traffic/<int:job_id>', methods=['POST'])
@login_required
def analyze_traffic(job_id):
    """Placeholder for traffic analysis feature."""
    db = JobDatabase()
    # In a real implementation, you would fetch job/company details
    # and use an external API (like SimilarWeb, etc.) to get traffic data.
    jobs, _ = db.get_jobs_for_user(current_user.id, limit=1, search_query=f"id:{job_id}")

    if not jobs:
        return jsonify({'status': 'error', 'message': 'Job not found.'}), 404
        
    company_name = jobs[0].get('company')
    logger.info(f"Traffic analysis requested for job {job_id} (Company: {company_name}) by user {current_user.id}.")

    # Placeholder response with fake data for the modal
    # In a real scenario, this would come from an API call.
    fake_analysis_data = {
        "monthly_visits": f"{random.randint(50, 500)}K",
        "bounce_rate": f"{random.randint(30, 70)}%",
        "pages_per_visit": f"{random.uniform(1.5, 5.0):.2f}",
    }

    return jsonify({
        'status': 'success',
        'company_name': company_name,
        'message': f'Traffic analysis for {company_name}.',
        'analysis': fake_analysis_data
    })

@app.route('/api/job_status_data')
@login_required
def job_status_data():
    """API endpoint to provide job status data for charts."""
    db = JobDatabase()
    status_counts = db.get_job_status_counts(current_user.id)
    return jsonify(status_counts)

@app.route('/sync_calendar', methods=['POST'])
@login_required
def sync_calendar():
    """Manually sync existing jobs with Google Calendar (with duplicate prevention)."""
    try:
        db = JobDatabase()
        calendar_manager = GoogleCalendarManager()
        
        if not calendar_manager.calendar_service or not calendar_manager.job_alerts_calendar_id:
            flash('Google Calendar not connected or Job Alerts calendar not created. Please check your connection.', 'danger')
            return redirect(url_for('dashboard'))
        
        # Get all jobs for the user
        jobs, _ = db.get_jobs_for_user(current_user.id, limit=1000)
        
        synced_count = 0
        skipped_count = 0
        
        flash(f'Starting sync of {len(jobs)} jobs to "{calendar_manager.user_name}\'s Job Alerts" calendar...', 'info')
        
        for job in jobs:
            try:
                events_created = 0
                
                # Create deadline reminders for jobs with deadlines
                if job.get('deadline') and job['deadline'] != 'N/A':
                    result = calendar_manager.create_deadline_reminder(job, job['deadline'])
                    if result:
                        events_created += 1
                
                # Create interview events for jobs in interview status
                if job.get('status') == 'interview':
                    result = calendar_manager.create_interview_event(job)
                    if result:
                        events_created += 1
                
                # Create follow-up reminders for applied/offered/rejected jobs
                if job.get('status') in ['applied', 'offered', 'rejected']:
                    result = calendar_manager.create_followup_reminder(job, job['status'])
                    if result:
                        events_created += 1
                
                if events_created > 0:
                    synced_count += events_created
                else:
                    skipped_count += 1
                    
            except Exception as e:
                logger.warning(f"Failed to sync job {job.get('id')}: {e}")
                skipped_count += 1
                continue
        
        if synced_count > 0:
            flash(f'✅ Successfully synced {synced_count} calendar events! {skipped_count} jobs skipped (no events needed or duplicates prevented).', 'success')
        else:
            flash(f'No new calendar events created. {skipped_count} jobs processed (may already have events or no events needed).', 'info')
        
    except Exception as e:
        logger.error(f"Calendar sync failed: {e}")
        flash('Calendar sync failed. Please check your Google Calendar connection.', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/test_calendar', methods=['POST'])
@login_required
def test_calendar():
    """Test Google Calendar connection."""
    try:
        calendar_manager = GoogleCalendarManager()
        
        if not calendar_manager.calendar_service:
            flash('Google Calendar connection failed. Please check your credentials.json file.', 'danger')
            return redirect(url_for('dashboard'))
        
        # Try to list calendars to test connection
        calendars = calendar_manager.calendar_service.calendarList().list().execute()
        primary_calendar = next((cal for cal in calendars['items'] if cal.get('primary')), None)
        
        if primary_calendar:
            flash(f'✅ Google Calendar connected successfully! Primary calendar: {primary_calendar.get("summary", "Unknown")}', 'success')
        else:
            flash('Google Calendar connected but no primary calendar found.', 'warning')
            
    except Exception as e:
        logger.error(f"Calendar test failed: {e}")
        error_str = str(e).lower()
        if "insufficient authentication scopes" in error_str:
            flash('❌ Calendar permissions missing. Click "Re-authorize Google APIs" to fix this.', 'danger')
        elif "accessnotconfigured" in error_str or "has not been used" in error_str:
            flash('❌ Google Calendar API not enabled. Please enable it in Google Cloud Console for project: job-alert-automation-465708', 'danger')
        else:
            flash(f'Google Calendar connection test failed: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/reauth_google', methods=['POST'])
@login_required
def reauth_google():
    """Force re-authentication with Google APIs to get all required scopes."""
    try:
        # Delete existing token to force re-auth
        if os.path.exists('token.pickle'):
            os.remove('token.pickle')
            logger.info("Deleted existing token.pickle for re-authentication")
        
        # Force new authentication
        emailer = SmartEmailer()
        if emailer.gmail_service:
            flash('✅ Google APIs re-authorized successfully! You now have Gmail and Calendar access.', 'success')
        else:
            flash('❌ Re-authorization failed. Please check your credentials.json file.', 'danger')
            
    except Exception as e:
        logger.error(f"Re-authorization failed: {e}")
        flash(f'Re-authorization failed: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/calendar_events')
@login_required
def get_calendar_events():
    """Get all job-related calendar events."""
    try:
        calendar_manager = GoogleCalendarManager()
        
        if not calendar_manager.calendar_service:
            return jsonify({'error': 'Calendar not connected'}), 400
        
        # Get events from the last 30 days to next 90 days
        from datetime import datetime, timedelta
        import pytz
        
        now = datetime.now(pytz.UTC)
        time_min = (now - timedelta(days=30)).isoformat()
        time_max = (now + timedelta(days=90)).isoformat()
        
        # Search for events in our Job Alerts calendar
        calendar_id = calendar_manager.job_alerts_calendar_id or 'primary'
        
        events_result = calendar_manager.calendar_service.events().list(
            calendarId=calendar_id,
            timeMin=time_min,
            timeMax=time_max,
            maxResults=100,
            singleEvents=True,
            orderBy='startTime'
        ).execute()
        
        events = events_result.get('items', [])
        
        # Format events for display
        formatted_events = []
        for event in events:
            start = event['start'].get('dateTime', event['start'].get('date'))
            # Parse datetime for better display
            try:
                if 'T' in start:
                    dt = datetime.fromisoformat(start.replace('Z', '+00:00'))
                    display_date = dt.strftime('%Y-%m-%d %H:%M')
                else:
                    display_date = start
            except:
                display_date = start
                
            formatted_events.append({
                'id': event['id'],
                'summary': event.get('summary', 'No Title'),
                'description': event.get('description', ''),
                'start': display_date,
                'htmlLink': event.get('htmlLink', ''),
                'created': event.get('created', ''),
                'calendar_id': calendar_id
            })
        
        return jsonify({'events': formatted_events})
        
    except Exception as e:
        logger.error(f"Failed to get calendar events: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/delete_calendar_event/<event_id>', methods=['DELETE'])
@login_required
def delete_calendar_event(event_id):
    """Delete a specific calendar event."""
    try:
        calendar_manager = GoogleCalendarManager()
        
        if not calendar_manager.calendar_service:
            return jsonify({'error': 'Calendar not connected'}), 400
        
        # Use the Job Alerts calendar or primary calendar
        calendar_id = calendar_manager.job_alerts_calendar_id or 'primary'
        
        # Delete the event
        calendar_manager.calendar_service.events().delete(
            calendarId=calendar_id,
            eventId=event_id
        ).execute()
        
        return jsonify({'success': True, 'message': 'Event deleted successfully'})
        
    except Exception as e:
        logger.error(f"Failed to delete calendar event {event_id}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/calendar_sync_status')
@login_required
def calendar_sync_status():
    """Check if calendar is in sync and return last sync time."""
    try:
        calendar_manager = GoogleCalendarManager()
        
        if not calendar_manager.calendar_service:
            return jsonify({'error': 'Calendar not connected', 'synced': False}), 400
        
        # Get current event count
        from datetime import datetime, timedelta
        import pytz
        
        now = datetime.now(pytz.UTC)
        time_min = (now - timedelta(days=30)).isoformat()
        time_max = (now + timedelta(days=90)).isoformat()
        
        calendar_id = calendar_manager.job_alerts_calendar_id or 'primary'
        
        events_result = calendar_manager.calendar_service.events().list(
            calendarId=calendar_id,
            timeMin=time_min,
            timeMax=time_max,
            maxResults=100,
            singleEvents=True,
            orderBy='startTime'
        ).execute()
        
        events = events_result.get('items', [])
        
        return jsonify({
            'synced': True,
            'event_count': len(events),
            'last_sync': now.isoformat(),
            'calendar_id': calendar_id
        })
        
    except Exception as e:
        logger.error(f"Failed to check calendar sync status: {e}")
        return jsonify({'error': str(e), 'synced': False}), 500


# --- Missing Routes for Dashboard JavaScript ---

@app.route('/start_scrape', methods=['POST'])
@login_required
def start_scrape():
    """Start job scraping task and return task ID."""
    try:
        search_terms = request.form.get('search_terms', '')
        location = request.form.get('location', '')
        experience = request.form.get('experience', '')
        job_type = request.form.get('job_type', '')
        
        if not search_terms:
            return jsonify({'error': 'Search terms are required'}), 400
        
        # Convert search terms to list
        search_terms_list = [term.strip() for term in search_terms.split(',') if term.strip()]
        
        # Start the Celery task
        task = run_job_hunt_task.delay(
            current_user.id, search_terms_list, location, experience, job_type
        )
        
        return jsonify({
            'task_id': task.id,
            'status': 'started',
            'message': 'Job search started successfully'
        })
        
    except Exception as e:
        logger.error(f"Failed to start scraping task: {e}")
        return jsonify({'error': 'Failed to start search. Celery worker might be down.'}), 500


@app.route('/status/<task_id>')
@login_required
def status(task_id):
    """Get status of a Celery task."""
    try:
        task = run_job_hunt_task.AsyncResult(task_id)
        
        if task.state == 'PENDING':
            response = {
                'status': 'Pending...',
                'progress': 0,
                'details': 'Task is waiting to be processed...'
            }
        elif task.state == 'PROGRESS':
            response = {
                'status': task.info.get('status', 'In Progress...'),
                'progress': task.info.get('progress', 0),
                'details': task.info.get('status', 'Processing...')  # Use status as details for now
            }
        elif task.state == 'SUCCESS':
            response = {
                'status': 'SUCCESS',
                'progress': 100,
                'details': 'Job search completed successfully!',
                'message': task.info.get('message', 'Search completed')
            }
        else:  # FAILURE
            response = {
                'status': 'FAILURE',
                'progress': 100,
                'details': f'Task failed: {str(task.info)}',
                'message': 'Search failed'
            }
            
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Failed to get task status: {e}")
        return jsonify({
            'status': 'ERROR',
            'progress': 0,
            'details': f'Error checking task status: {str(e)}'
        }), 500


@app.route('/jobs')
@login_required
def jobs():
    """Get jobs for current user as JSON."""
    try:
        db = JobDatabase()
        jobs_data, total_count = db.get_jobs_for_user(current_user.id, limit=100)
        
        # Format jobs for frontend
        def get_status_colors(status):
            status_colors = {
                'new': {'bg': '#f3f4f6', 'text': '#374151'},
                'applied': {'bg': '#dcfce7', 'text': '#166534'},
                'interview': {'bg': '#e0e7ff', 'text': '#3730a3'},
                'offered': {'bg': '#fef3c7', 'text': '#92400e'},
                'rejected': {'bg': '#fee2e2', 'text': '#991b1b'},
                'withdrawn': {'bg': '#f3f4f6', 'text': '#6b7280'}
            }
            return status_colors.get(status, status_colors['new'])
        
        formatted_jobs = []
        for job in jobs_data:
            colors = get_status_colors(job.get('status', 'new'))
            formatted_job = {
                'id': job['id'],
                'title': job['title'],
                'company': job['company'],
                'location': job['location'],
                'salary': job['salary'] or 'Not specified',
                'link': job['link'],
                'post_date': job['posted_date'],
                'status': job['status'],
                'status_color': colors['bg'],
                'status_text_color': colors['text'],
                'description': job['description'][:200] + '...' if len(job.get('description', '')) > 200 else job.get('description', '')
            }
            formatted_jobs.append(formatted_job)
        
        return jsonify(formatted_jobs)
        
    except Exception as e:
        logger.error(f"Failed to fetch jobs: {e}")
        return jsonify({'error': 'Failed to fetch jobs'}), 500

@app.route('/api/analyze_skill_gap', methods=['POST'])
@login_required
def analyze_skill_gap():
    try:
        data = request.get_json()
        target_role = data.get('target_role')
        target_company = data.get('target_company', '')
        
        if not target_role:
            return jsonify({'error': 'Target role is required'}), 400
            
        # Fetch user resume from DB
        db = JobDatabase()
        user_details = db.get_user_details(current_user.id)
        resume_text = ""
        if user_details:
            # Try parsed data first
            if user_details.get('resume_parsed_data'):
                try:
                    parsed = json.loads(user_details['resume_parsed_data'])
                    resume_text = json.dumps(parsed)
                except:
                    resume_text = user_details['resume_parsed_data']
        
        if not resume_text:
             resume_text = "No resume data found. Please upload a resume in your profile."

        system_prompt = """
        You are an expert Career Coach. Analyze the gap between the candidate's resume and the target job role.
        If a Target Company is provided, provide specific insights about that company's culture, interview process, or specific tech stack preferences.
        
        Return a JSON object with exactly these keys:
        - readiness_score: (integer 0-100)
        - strong_skills: (list of matching skills)
        - missing_skills: (list of critical skills missing)
        - company_insights: (string or null) - Actionable advice on how to learn about the company and prepare for them specifically.
        - company_specific_skills: (list of strings) - Skills specifically valued by this company (e.g. "Amazon Leadership Principles" for Amazon, or specific tech stack).
        """
        
        user_prompt = f"TARGET ROLE: {target_role}\nTARGET COMPANY: {target_company}\nRESUME:\n{resume_text}"
        
        result = safe_ai_request(system_prompt, user_prompt)
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error in analyze_skill_gap: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/learning-path/generate', methods=['POST'])
@login_required
def generate_roadmap():
    try:
        data = request.get_json()
        target_role = data.get('target_role')
        missing_skills = data.get('missing_skills')
        current_skill_level = data.get('current_skill_level', 'Beginner')
        availability_hours = data.get('availability_hours', 10)
        learning_style = data.get('learning_style', 'Project-based')

        try:
            availability_hours = int(availability_hours)
        except (ValueError, TypeError):
            availability_hours = 10

        service = RoadmapService()
        result = service.generate(target_role, current_skill_level, missing_skills, availability_hours, learning_style)
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error in generate_roadmap: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate_combined_roadmap', methods=['POST'])
@login_required
def generate_combined_roadmap():
    try:
        data = request.get_json()
        target_role = data.get('target_role')
        user_skills = data.get('user_skills', [])
        current_skill_level = data.get('current_skill_level', 'Beginner')
        availability_hours = data.get('availability_hours', 10)
        learning_style = data.get('learning_style', 'Project-based')

        if not target_role:
            return jsonify({'error': 'Target role is required'}), 400

        try:
            availability_hours = int(availability_hours)
        except (ValueError, TypeError):
            availability_hours = 10
        
        # Fallback: fetch skills from resume if not provided
        if not user_skills:
            db = JobDatabase()
            user_details = db.get_user_details(current_user.id)
            if user_details and user_details.get('resume_parsed_data'):
                try:
                    parsed = json.loads(user_details['resume_parsed_data'])
                    user_skills = parsed.get('skills', [])
                except Exception as e:
                    logger.warning(f"Failed to parse resume skills for combined roadmap: {e}")

        service = RoadmapService()
        result = service.generate_combined(target_role, user_skills, current_skill_level, availability_hours, learning_style)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error in generate_combined_roadmap: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/save_roadmap', methods=['POST'])
@login_required
def save_roadmap():
    try:
        data = request.get_json()
        target_role = data.get('target_role')
        roadmap_data = data.get('roadmap')
        path_id = data.get('path_id')

        if not target_role or not roadmap_data:
            return jsonify({'error': 'Target role and roadmap data are required'}), 400

        weekly_breakdown = roadmap_data.get('weekly_breakdown', [])

        if path_id:
            # Update existing path
            path = LearningPath.query.filter_by(id=path_id, user_id=current_user.id).first()
            if not path:
                return jsonify({'error': 'Path not found'}), 404
            
            path.last_accessed = datetime.datetime.utcnow()
            
            # Update progress for existing modules
            for idx, week in enumerate(weekly_breakdown):
                module = LearningModule.query.filter_by(path_id=path.id, order_index=idx).first()
                if module:
                    module.topics_data = week # Save granular progress
                    # Determine completion: All topics AND mini-project must be done
                    mini_project = week.get('mini_project', {})
                    is_project_completed = mini_project.get('completed', False) if isinstance(mini_project, dict) else False
                    
                    focus_topics = week.get('focus_topics', [])
                    all_topics_completed = True
                    if isinstance(focus_topics, list):
                        for t in focus_topics:
                            if isinstance(t, dict) and not t.get('completed'):
                                all_topics_completed = False
                                break
                    
                    is_completed = is_project_completed and all_topics_completed
                    
                    progress = UserLearningProgress.query.filter_by(user_id=current_user.id, module_id=module.id).first()
                    if not progress:
                        progress = UserLearningProgress(user_id=current_user.id, module_id=module.id)
                        db.session.add(progress)
                    
                    progress.is_completed = is_completed
                    if is_completed and not progress.completed_at:
                        progress.completed_at = datetime.datetime.utcnow()
                    elif not is_completed:
                        progress.completed_at = None
            
            db.session.commit()
            return jsonify({'success': True, 'path_id': path.id})

        else:
            # Create New Path
            new_path = LearningPath(
                user_id=current_user.id,
                target_role=target_role,
                last_accessed=datetime.datetime.utcnow()
            )
            db.session.add(new_path)
            db.session.flush()

            for idx, week in enumerate(weekly_breakdown):
                week_num = week.get('week_number', idx + 1)
                focus_topics = week.get('focus_topics', [])
                
                # Handle focus_topics whether they are strings or objects
                topics_list = []
                if isinstance(focus_topics, list):
                    for t in focus_topics:
                        if isinstance(t, dict):
                            topics_list.append(t.get('title', ''))
                        else:
                            topics_list.append(str(t))
                topics = ", ".join(topics_list)
                
                title = f"Week {week_num}: {topics}"
                if len(title) > 255:
                    title = title[:252] + "..."
                
                # Handle mini_project
                mini_project = week.get('mini_project', {})
                project_title = mini_project.get('title', '') if isinstance(mini_project, dict) else str(mini_project)
                    
                description = f"Why it matters: {week.get('why_this_matters', '')}\n\nProject: {project_title}"
                
                new_module = LearningModule(
                    path_id=new_path.id,
                    title=title,
                    description=description,
                    order_index=idx,
                    topics_data=week # Save initial state
                )
                db.session.add(new_module)
                db.session.flush() # Get ID for progress

                # Check initial completion status
                is_project_completed = mini_project.get('completed', False) if isinstance(mini_project, dict) else False
                all_topics_completed = True
                if isinstance(focus_topics, list):
                    for t in focus_topics:
                        if isinstance(t, dict) and not t.get('completed'):
                            all_topics_completed = False
                            break
                
                if is_project_completed and all_topics_completed:
                    prog = UserLearningProgress(user_id=current_user.id, module_id=new_module.id, is_completed=True, completed_at=datetime.datetime.utcnow())
                    db.session.add(prog)
        
            db.session.commit()
            return jsonify({'success': True, 'path_id': new_path.id})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving roadmap: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/analyze_job_match', methods=['POST'])
@login_required
def analyze_job_match():
    try:
        data = request.get_json()
        job_description = data.get('job_description')

        if not job_description:
            return jsonify({'error': 'Job description is required'}), 400

        # Fetch user resume from DB
        db = JobDatabase()
        user_details = db.get_user_details(current_user.id)
        resume_text = ""
        if user_details:
            if user_details.get('resume_parsed_data'):
                try:
                    parsed = json.loads(user_details['resume_parsed_data'])
                    resume_text = json.dumps(parsed)
                except:
                    resume_text = user_details['resume_parsed_data']
        
        if not resume_text:
             resume_text = "No resume data found. Please upload a resume in your profile."

        system_prompt = """
        You are an expert ATS (Applicant Tracking System) scanner. 
        Compare the Resume against the Job Description.
        Return a JSON object with exactly these keys:
        - match_score: (integer 0-100)
        - matched_skills: (list of strings)
        - missing_skills: (list of strings)
        - missing_keywords: (list of strings found in JD but missing in resume)
        - resume_improvement_suggestions: (list of actionable strings)
        """

        user_prompt = f"RESUME:\n{resume_text}\n\nJOB DESCRIPTION:\n{job_description}"

        result = safe_ai_request(system_prompt, user_prompt)
        return jsonify(result)

    except Exception as e:
        logger.error(f"Error in analyze_job_match: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

# --- Learning Path APIs ---

@app.route('/learning-path/save', methods=['POST'])
@login_required
def save_learning_path():
    try:
        data = request.get_json()
        target_role = data.get('target_role')
        modules_data = data.get('modules', []) # Expecting list of {title, description}

        if not target_role or not modules_data:
            return jsonify({'error': 'Target role and modules are required'}), 400

        # Create Path
        new_path = LearningPath(
            user_id=current_user.id,
            target_role=target_role,
            last_accessed=datetime.datetime.utcnow()
        )
        db.session.add(new_path)
        db.session.flush() # Get ID

        # Create Modules
        for idx, mod in enumerate(modules_data):
            new_module = LearningModule(
                path_id=new_path.id,
                title=mod.get('title'),
                description=mod.get('description', ''),
                order_index=idx
            )
            db.session.add(new_module)
        
        db.session.commit()
        return jsonify({'success': True, 'path_id': new_path.id})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving learning path: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate-project-guide', methods=['POST'])
@login_required
def generate_project_guide():
    """Generate a PDF guide for a mini project using Cerebras AI"""
    try:
        data = request.get_json()
        project_title = data.get('project_title', '').strip()
        
        if not project_title:
            return jsonify({'error': 'Project title is required'}), 400
        
        # Sanitize filename
        safe_filename = re.sub(r'[^a-zA-Z0-9_-]', '_', project_title.lower())[:50]
        pdf_filename = f'project_{safe_filename}.pdf'
        pdf_path = os.path.join('static', 'learning_resources', pdf_filename)
        
        # Check if PDF already exists (cache)
        if os.path.exists(pdf_path):
            logger.info(f"PDF already exists for {project_title}, serving from cache")
            return jsonify({
                'success': True,
                'pdf_url': f'/static/learning_resources/{pdf_filename}',
                'cached': True
            })
        
        # Generate guide content using Cerebras
        logger.info(f"Generating PDF guide for: {project_title}")
        
        system_prompt = """You are an expert technical educator creating step-by-step project guides.
Create comprehensive, beginner-friendly tutorials with clear instructions, code examples, and best practices."""
        
        user_prompt = f"""Create a detailed step-by-step tutorial guide for the following mini project:

"{project_title}"

Structure the guide with these sections:
1. **Project Overview**: Brief description (2-3 sentences) of what this project accomplishes
2. **Prerequisites**: Required tools, technologies, and prerequisite knowledge
3. **Step-by-Step Instructions**: Detailed, numbered steps to complete the project
4. **Code Examples**: Include relevant code snippets with explanations
5. **Common Pitfalls**: List of common mistakes and how to avoid them
6. **Expected Outcome**: What the finished project should look like/do
7. **Next Steps**: Suggestions for extending or improving the project

Format the response as valid JSON with a single key 'content' containing HTML-formatted guide content.
Use proper HTML tags: <h2>, <h3>, <p>, <code>, <pre>, <ul>, <ol>, <li>, etc.
Make it practical, detailed, and beginner-friendly."""
        
        # Call Cerebras API
        response_data = safe_ai_request(system_prompt, user_prompt, model="llama-3.3-70b", retries=2)
        
        if not response_data or 'content' not in response_data:
            return jsonify({'error': 'Failed to generate guide content'}), 500
        
        html_content = response_data['content']
        
        # Handle if content is an array of strings (join them)
        if isinstance(html_content, list):
            html_content = ''.join(html_content)
        
        # Clean up escaped newlines and extra whitespace
        html_content = html_content.replace('\\n', '\n').replace('\n\n\n', '\n\n')
        
        # Create full HTML document for PDF
        html_doc = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>{project_title} - Project Guide</title>
<style>
@page {{ size: A4; margin: 2cm; }}
body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 800px; margin: 0 auto; }}
h1 {{ color: #2563eb; border-bottom: 3px solid #2563eb; padding-bottom: 10px; }}
h2 {{ color: #1e40af; margin-top: 25px; }}
h3 {{ color: #3b82f6; margin-top: 15px; }}
code {{ background: #f3f4f6; padding: 2px 6px; border-radius: 3px; font-family: monospace; font-size: 0.9em; color: #1f2937; }}
pre {{ background: #f9fafb; color: #1f2937; padding: 15px; border-radius: 8px; margin: 15px 0; border: 1px solid #e5e7eb; }}
pre code {{ background: none; color: #1f2937; padding: 0; }}
ul, ol {{ margin: 10px 0; padding-left: 30px; }}
li {{ margin: 8px 0; }}
</style></head><body>
<div style="text-align: center; margin-bottom: 30px; border-bottom: 2px solid #e5e7eb; padding-bottom: 15px;">
<h1>{project_title}</h1>
<p style="color: #6b7280;">A Step-by-Step Project Guide</p>
</div>
{html_content}
<div style="margin-top: 40px; padding-top: 15px; border-top: 2px solid #e5e7eb; text-align: center; color: #6b7280; font-size: 0.85em;">
<p>Generated by JobSnap Learning Path</p>
</div></body></html>"""
        
        # Convert HTML to PDF using xhtml2pdf (Windows-friendly)
        try:
            from xhtml2pdf import pisa
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
            
            # Generate PDF
            with open(pdf_path, 'wb') as pdf_file:
                pisa_status = pisa.CreatePDF(html_doc, dest=pdf_file)
            
            if pisa_status.err:
                raise Exception(f"PDF generation failed with {pisa_status.err} errors")
            
            logger.info(f"Successfully generated PDF: {pdf_filename}")
            
            return jsonify({
                'success': True,
                'pdf_url': f'/static/learning_resources/{pdf_filename}',
                'cached': False
            })
            
        except ImportError:
            logger.error("xhtml2pdf not installed")
            return jsonify({'error': 'PDF generation library not available'}), 500
        except Exception as pdf_error:
            logger.error(f"PDF generation error: {pdf_error}")
            return jsonify({'error': f'Failed to generate PDF: {str(pdf_error)}'}), 500
            
    except Exception as e:
        logger.error(f"Error generating project guide: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/learning-paths', methods=['GET'])
@login_required
def get_learning_paths():
    try:
        paths = LearningPath.query.filter_by(user_id=current_user.id).order_by(LearningPath.last_accessed.desc()).all()
        result = []
        
        for path in paths:
            total_modules = len(path.modules)
            
            # Get completed modules for this path
            completed_progress = UserLearningProgress.query.join(LearningModule).filter(
                UserLearningProgress.user_id == current_user.id,
                UserLearningProgress.is_completed == True,
                LearningModule.path_id == path.id
            ).all()
            
            completed_module_ids = {p.module_id for p in completed_progress}
            completed_count = len(completed_module_ids)
            
            progress = (completed_count / total_modules * 100) if total_modules > 0 else 0
            
            # Determine the next module to learn (Resume feature)
            next_module = None
            for mod in path.modules:
                if mod.id not in completed_module_ids:
                    next_module = {
                        'id': mod.id,
                        'title': mod.title,
                        'description': mod.description
                    }
                    break
            
            result.append({
                'path_id': path.id,
                'target_role': path.target_role,
                'progress': round(progress),
                'total_modules': total_modules,
                'completed_modules': completed_count,
                'last_accessed': path.last_accessed.isoformat(),
                'next_module': next_module,
                'is_completed': completed_count == total_modules and total_modules > 0
            })
            
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching paths: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/learning-path/resume', methods=['GET'])
@login_required
def resume_learning():
    """
    Resume the most recently accessed learning path.
    Returns the path details and the specific module to continue with.
    """
    try:
        # Find the last accessed path
        last_path = LearningPath.query.filter_by(user_id=current_user.id).order_by(LearningPath.last_accessed.desc()).first()
        
        if not last_path:
            return jsonify({'has_path': False, 'message': 'No learning paths started yet.'})
            
        # Calculate progress to find next module
        completed_progress = UserLearningProgress.query.join(LearningModule).filter(
            UserLearningProgress.user_id == current_user.id,
            UserLearningProgress.is_completed == True,
            LearningModule.path_id == last_path.id
        ).all()
        
        completed_module_ids = {p.module_id for p in completed_progress}
        
        next_module = None
        for mod in last_path.modules:
            if mod.id not in completed_module_ids:
                next_module = {
                    'id': mod.id,
                    'title': mod.title,
                    'description': mod.description
                }
                break
        
        return jsonify({
            'has_path': True,
            'path_id': last_path.id,
            'target_role': last_path.target_role,
            'next_module': next_module,
            'is_completed': next_module is None
        })
        
    except Exception as e:
        logger.error(f"Error resuming learning: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/learning-path/<int:path_id>', methods=['DELETE'])
@login_required
def delete_learning_path(path_id):
    try:
        path = LearningPath.query.filter_by(id=path_id, user_id=current_user.id).first_or_404()
        
        # Manually delete progress records first to avoid foreign key constraints
        # if the database schema doesn't have ON DELETE CASCADE configured.
        module_ids = [m.id for m in path.modules]
        if module_ids:
            UserLearningProgress.query.filter(UserLearningProgress.module_id.in_(module_ids)).delete(synchronize_session=False)
            
        db.session.delete(path)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Learning path deleted successfully'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting learning path {path_id}: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/learning-path/<int:path_id>', methods=['GET'])
@login_required
def get_learning_path_detail(path_id):
    path = LearningPath.query.filter_by(id=path_id, user_id=current_user.id).first_or_404()
    
    # Update last accessed
    path.last_accessed = datetime.datetime.utcnow()
    db.session.commit()

    modules_data = []
    for mod in path.modules:
        progress = UserLearningProgress.query.filter_by(user_id=current_user.id, module_id=mod.id).first()
        modules_data.append({
            'id': mod.id,
            'title': mod.title,
            'description': mod.description,
            'is_completed': progress.is_completed if progress else False,
            'topics_data': mod.topics_data
        })

    return jsonify({
        'path_id': path.id,
        'target_role': path.target_role,
        'modules': modules_data
    })

@app.route('/learning-path/<int:path_id>/progress', methods=['POST'])
@login_required
def update_path_progress(path_id):
    data = request.get_json()
    module_id = data.get('module_id')
    is_completed = data.get('is_completed')

    # Verify ownership via path
    module = LearningModule.query.join(LearningPath).filter(
        LearningModule.id == module_id,
        LearningPath.id == path_id,
        LearningPath.user_id == current_user.id
    ).first_or_404()

    progress = UserLearningProgress.query.filter_by(user_id=current_user.id, module_id=module_id).first()
    
    if not progress:
        progress = UserLearningProgress(user_id=current_user.id, module_id=module_id)
        db.session.add(progress)
    
    progress.is_completed = is_completed
    progress.completed_at = datetime.datetime.utcnow() if is_completed else None
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/learning-path/<int:path_id>/export', methods=['GET'])
@login_required
def export_learning_path(path_id):
    fmt = request.args.get('format', 'md')
    path = LearningPath.query.filter_by(id=path_id, user_id=current_user.id).first_or_404()
    
    if fmt == 'certificate':
        if not HAS_REPORTLAB:
            return jsonify({'error': 'PDF generation library (reportlab) not installed.'}), 501
            
        # Check completion
        total_modules = len(path.modules)
        completed_progress = UserLearningProgress.query.join(LearningModule).filter(
            UserLearningProgress.user_id == current_user.id,
            UserLearningProgress.is_completed == True,
            LearningModule.path_id == path.id
        ).count()
        
        if total_modules == 0 or completed_progress < total_modules:
             return jsonify({'error': 'Cannot generate certificate. Learning path is not 100% complete.'}), 400

        from reportlab.lib.pagesizes import landscape
        from reportlab.lib.units import inch
        
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=landscape(letter))
        width, height = landscape(letter)
        
        # Decorative Border
        c.setStrokeColorRGB(0.25, 0.25, 0.7)
        c.setLineWidth(4)
        c.rect(0.5*inch, 0.5*inch, width-1*inch, height-1*inch)
        c.setStrokeColorRGB(0.8, 0.8, 0.8)
        c.setLineWidth(1)
        c.rect(0.6*inch, 0.6*inch, width-1.2*inch, height-1.2*inch)

        # Content
        c.setFillColorRGB(0.2, 0.2, 0.2)
        c.setFont("Helvetica-Bold", 42)
        c.drawCentredString(width/2, height - 2.5*inch, "Certificate of Completion")
        
        c.setFont("Helvetica", 20)
        c.drawCentredString(width/2, height - 3.5*inch, "This is to certify that")
        
        # User Name
        user_name = current_user.username
        try:
            db_instance = JobDatabase()
            user_details = db_instance.get_user_details(current_user.id)
            if user_details and user_details.get('full_name'):
                user_name = user_details.get('full_name')
        except: pass
            
        c.setFont("Helvetica-Bold", 32)
        c.setFillColorRGB(0.1, 0.1, 0.6)
        c.drawCentredString(width/2, height - 4.5*inch, user_name)
        
        c.setFillColorRGB(0.2, 0.2, 0.2)
        c.setFont("Helvetica", 20)
        c.drawCentredString(width/2, height - 5.5*inch, "has successfully completed the learning path")
        
        c.setFont("Helvetica-Bold", 28)
        c.drawCentredString(width/2, height - 6.5*inch, path.target_role)
        
        # Footer
        c.setFont("Helvetica", 12)
        date_str = datetime.datetime.now().strftime("%B %d, %Y")
        c.drawString(1.5*inch, 1.5*inch, f"Date: {date_str}")
        c.drawRightString(width - 1.5*inch, 1.5*inch, "AI Job Agent Platform")
        
        c.showPage()
        c.save()
        
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"Certificate_{path.target_role.replace(' ', '_')}.pdf", mimetype='application/pdf')

    if fmt == 'pdf':
        if not HAS_REPORTLAB:
            return jsonify({'error': 'PDF generation library (reportlab) not installed. Please install it or export as Markdown.'}), 501
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        story.append(Paragraph(f"Learning Path: {path.target_role}", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Modules
        for mod in path.modules:
            story.append(Paragraph(f"Module {mod.order_index + 1}: {mod.title}", styles['Heading2']))
            # Simple handling of newlines for PDF
            desc_text = mod.description.replace('\n', '<br/>') if mod.description else ""
            story.append(Paragraph(desc_text, styles['BodyText']))
            story.append(Spacer(1, 12))
            
        doc.build(story)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"learning_path_{path.id}.pdf", mimetype='application/pdf')
        
    else:
        # Markdown Export
        content = f"# Learning Path: {path.target_role}\n\n"
        for mod in path.modules:
            content += f"## Module {mod.order_index + 1}: {mod.title}\n\n"
            if mod.description:
                content += f"{mod.description}\n\n"
            content += "---\n\n"
        
        buffer = BytesIO()
        buffer.write(content.encode('utf-8'))
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"learning_path_{path.id}.md", mimetype='text/markdown')

# This part runs the Flask app if executed directly
if __name__ == "__main__":
    # Initialize database
    db_instance = JobDatabase()
    db_instance.init_db() # Ensure tables are created/migrated

    # Setup signal handlers for graceful shutdown
    signal.signal(signal.SIGINT, lambda s, f: sys.exit(0))
    signal.signal(signal.SIGTERM, lambda s, f: sys.exit(0))
    
    logger.info("Starting Flask application...")
    app.run(debug=True, host='0.0.0.0', port=5000)
